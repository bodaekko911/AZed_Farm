from typing import Optional

from datetime import date
import io

from fastapi import APIRouter, Depends
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.permissions import require_action, require_permission
from app.database import get_async_session
from app.core.navigation import render_app_header
from app.models.product import Product
from app.models.user import User
from app.services.receive_service import (
    BatchReceiptCreate,
    ReceiptCreate,
    ReceiptUpdate,
    create_receipt,
    create_receipt_batch,
    delete_receipt,
    list_receipts,
    update_receipt,
)

router = APIRouter(
    prefix="/receive",
    tags=["Receive Products"],
    dependencies=[Depends(require_permission("page_receive_products"))],
)


def to_xlsx(headers, rows, sheet_name="Receive Products"):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    hfill = PatternFill("solid", fgColor="2a7a2a")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5FAF5")
    for column in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in column), default=10)
        ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 40)
    ws.row_dimensions[1].height = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── API ───────────────────────────────────────────────────────────────────────

@router.get("/api/products")
async def get_products(db: AsyncSession = Depends(get_async_session)):
    result = await db.execute(
        select(Product)
        .where(or_(Product.is_active.is_(True), Product.is_active.is_(None)))
        .order_by(Product.name)
    )
    return [
        {
            "id":    p.id,
            "sku":   p.sku,
            "name":  p.name,
            "unit":  p.unit,
            "cost":  float(p.cost)  if p.cost  is not None else 0.0,
            "stock": float(p.stock) if p.stock is not None else 0.0,
        }
        for p in result.scalars().all()
    ]


@router.get("/api/suppliers")
async def get_suppliers_list(db: AsyncSession = Depends(get_async_session)):
    """Lightweight supplier list for the receive form dropdown."""
    from app.models.supplier import Supplier
    result = await db.execute(select(Supplier).order_by(Supplier.name))
    return [
        {
            "id":      s.id,
            "name":    s.name,
            "balance": float(s.balance or 0),
        }
        for s in result.scalars().all()
    ]


@router.get("/api/expense-categories")
async def get_expense_categories(db: AsyncSession = Depends(get_async_session)):
    """Active expense categories — lets the receive form override the
    auto-derived category (Products / Packaging Materials) and route the
    posted expense to any chosen category instead."""
    from app.models.expense import ExpenseCategory
    result = await db.execute(
        select(ExpenseCategory)
        .where(ExpenseCategory.is_active == "1")
        .order_by(ExpenseCategory.account_code, ExpenseCategory.name)
    )
    return [
        {
            "id":           c.id,
            "name":         c.name,
            "account_code": c.account_code,
        }
        for c in result.scalars().all()
    ]


@router.get("/api/locations")
async def get_locations_list(db: AsyncSession = Depends(get_async_session)):
    """Active storage locations for the receive form's destination dropdown.

    Proxied here (gated only by page_receive_products) so users who can
    receive stock don't also need page_inventory just to load the picker.
    """
    from app.models.inventory import StockLocation
    result = await db.execute(
        select(StockLocation)
        .where(StockLocation.is_active == True)
        .order_by(StockLocation.name)
    )
    return {
        "items": [
            {
                "id":   loc.id,
                "name": loc.name,
                "code": loc.code or "",
            }
            for loc in result.scalars().all()
        ],
    }


@router.get("/api/farms")
async def get_farms_list(db: AsyncSession = Depends(get_async_session)):
    """Active farms for the receive form's cost-allocation dropdown.

    Proxied here (gated only by page_receive_products) so users who can
    receive stock don't also need page_farm just to load the picker.
    """
    from app.models.farm import Farm
    result = await db.execute(
        select(Farm).where(Farm.is_active == 1).order_by(Farm.name)
    )
    return [
        {
            "id":       f.id,
            "name":     f.name,
            "location": f.location or "",
        }
        for f in result.scalars().all()
    ]


@router.post("/api/receive-batch", status_code=201)
async def receive_products_batch(
    data: BatchReceiptCreate,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(require_action("receive_products", "receipts", "create")),
):
    return await create_receipt_batch(db, data, current_user)


# Single-product endpoint kept for backward compatibility / API clients.
@router.post("/api/receive", status_code=201)
async def receive_products(
    data: ReceiptCreate,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(require_action("receive_products", "receipts", "create")),
):
    return await create_receipt(db, data, current_user)


@router.get("/api/history")
async def get_receipt_history(
    skip:       int           = 0,
    limit:      int           = 50,
    product_id: Optional[int] = None,
    db: AsyncSession = Depends(get_async_session),
):
    return await list_receipts(db, skip=skip, limit=limit, product_id=product_id)


@router.put("/api/receipt/{receipt_id}")
async def edit_receipt(
    receipt_id: int,
    data: ReceiptUpdate,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(require_action("receive_products", "receipts", "update")),
):
    return await update_receipt(db, receipt_id, data, current_user)


@router.delete("/api/receipt/{receipt_id}")
async def remove_receipt(
    receipt_id: int,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(require_action("receive_products", "receipts", "delete")),
):
    return await delete_receipt(db, receipt_id, current_user)


@router.get("/api/export.xlsx")
async def export_receipts_excel(
    product_id: Optional[int] = None,
    db: AsyncSession = Depends(get_async_session),
    _: User = Depends(require_action("receive_products", "receipts", "export")),
):
    data = await list_receipts(db, skip=0, limit=10000, product_id=product_id)
    headers = [
        "Receipt #",
        "Receive Date",
        "Product",
        "SKU",
        "Qty",
        "Unit Cost",
        "Total Cost",
        "Expense Ref",
        "Supplier",
        "Amount Paid",
        "Amount Unpaid",
        "Payment Method",
        "Supplier Ref",
        "Notes",
        "Received By",
        "Created At",
    ]
    rows = [
        [
            item["ref_number"],
            item["receive_date"],
            item["product_name"],
            item["product_sku"],
            item["qty"],
            item["unit_cost"],
            item["total_cost"],
            item["expense_ref"],
            item.get("supplier_name"),
            item.get("amount_paid"),
            item.get("amount_unpaid"),
            item.get("payment_method"),
            item["supplier_ref"],
            item["notes"],
            item["received_by"],
            item["created_at"],
        ]
        for item in data["items"]
    ]
    buf = to_xlsx(headers, rows, "Receive Products")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=receive_products_{date.today()}.xlsx"},
    )


# ── UI ────────────────────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
def receive_ui(current_user: User = Depends(require_permission("page_receive_products"))):
    return """<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<script src="/static/theme-init.js"></script>
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Receive Products — AZed ERP</title>
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#060810;--surface:#0a0d18;--card:#0f1424;--card2:#151c30;
  --border:rgba(255,255,255,0.06);--border2:rgba(255,255,255,0.11);
  --green:#00ff9d;--blue:#4d9fff;--amber:#ffb547;--danger:#ff4d6d;--purple:#a855f7;
  --text:#f0f4ff;--sub:#8899bb;--muted:#445066;
  --sans:'Outfit',sans-serif;--mono:'JetBrains Mono',monospace;
}
body.light{
  --bg:#f4f5ef;--surface:#f1f3eb;--card:#eceee6;--card2:#e4e6de;
  --border:rgba(0,0,0,0.08);--border2:rgba(0,0,0,0.14);
  --green:#0f8a43;
  --text:#1a1e14;--sub:#4a5040;--muted:#7b816f;
}
body.light nav{background:rgba(244,245,239,.92);}
body.light .picker-list{background:#fff;border-color:rgba(0,0,0,.12);}
body.light .picker-item:hover,.picker-item.active{background:rgba(77,159,255,.1);}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:var(--sans);background:var(--bg);color:var(--text);min-height:100vh;font-size:14px}

/* ── nav ── */
nav{position:sticky;top:0;z-index:200;display:flex;align-items:center;
  justify-content:space-between;gap:12px;padding:0 24px;height:58px;
  background:rgba(6,8,16,.92);backdrop-filter:blur(20px);border-bottom:1px solid var(--border)}
.nav-left{display:flex;align-items:center;gap:16px}
.nav-logo{font-size:17px;font-weight:900;text-decoration:none;
  background:linear-gradient(135deg,var(--green),var(--blue));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}
.nav-title{font-size:14px;font-weight:600;color:var(--sub)}
.nav-right{display:flex;align-items:center;gap:10px}
.mode-btn{display:flex;align-items:center;justify-content:center;width:36px;height:36px;
  border-radius:10px;border:1px solid var(--border);background:var(--card);
  color:var(--sub);font-size:16px;cursor:pointer;transition:all .2s;font-family:var(--sans)}
.mode-btn:hover{border-color:var(--border2);transform:scale(1.06)}
.account-menu{position:relative}
.user-pill{display:flex;align-items:center;gap:10px;background:var(--card);
  border:1px solid var(--border);border-radius:40px;padding:7px 16px 7px 10px;
  cursor:pointer;transition:all .2s}
.user-pill:hover,.user-pill.open{border-color:var(--border2)}
.user-avatar{width:28px;height:28px;background:linear-gradient(135deg,#7ecb6f,#d4a256);
  border-radius:50%;display:flex;align-items:center;justify-content:center;
  font-size:12px;font-weight:700;color:#0a0c08}
.user-name{font-size:13px;font-weight:500;color:var(--sub)}
.menu-caret{font-size:11px;color:var(--muted)}
.account-dropdown{position:absolute;right:0;top:calc(100% + 10px);min-width:220px;
  background:var(--card);border:1px solid var(--border2);border-radius:14px;padding:8px;
  box-shadow:0 24px 50px rgba(0,0,0,.35);display:none;z-index:500}
.account-dropdown.open{display:block}
.account-head{padding:10px 12px 8px;border-bottom:1px solid var(--border);margin-bottom:6px}
.account-label{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:1px}
.account-email{font-size:12px;color:var(--sub);margin-top:4px;word-break:break-word}
.account-item{width:100%;display:flex;align-items:center;gap:10px;padding:10px 12px;border:none;
  background:transparent;border-radius:10px;color:var(--sub);font-family:var(--sans);
  font-size:13px;text-decoration:none;cursor:pointer;text-align:left}
.account-item:hover{background:var(--card2);color:var(--text)}
.account-item.danger:hover{color:var(--danger)}
.logout-btn{background:transparent;border:1px solid var(--border);color:var(--muted);
  font-family:var(--sans);font-size:12px;padding:8px 16px;border-radius:8px;
  cursor:pointer;transition:all .2s}
.logout-btn:hover{border-color:var(--danger);color:var(--danger)}

/* ── layout ── */
.page{max-width:1100px;margin:0 auto;padding:32px 24px 60px}
.page-header{margin-bottom:28px}
.page-header h1{font-size:24px;font-weight:700;margin-bottom:4px}
.page-header p{font-size:13px;color:var(--muted)}
.card{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:24px;margin-bottom:28px}
.section-title{font-size:11px;font-weight:600;letter-spacing:2px;text-transform:uppercase;
  color:var(--muted);margin-bottom:20px;display:flex;align-items:center;gap:12px}
.section-title::after{content:'';flex:1;height:1px;background:linear-gradient(90deg,var(--border2),transparent)}
.section-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:20px}
.section-head .section-title{margin-bottom:0;flex:1}

/* ── meta fields (date / supplier / notes) ── */
.meta-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:20px}
@media(max-width:600px){.meta-grid{grid-template-columns:1fr}}
.field{display:flex;flex-direction:column;gap:6px}
.field.full{grid-column:1/-1}
label{font-size:11px;font-weight:600;color:var(--sub);letter-spacing:.5px;text-transform:uppercase}
input[type=text],input[type=number],input[type=date],textarea{
  background:var(--surface);border:1px solid var(--border);border-radius:10px;
  color:var(--text);font-family:var(--sans);font-size:14px;padding:9px 13px;
  transition:border-color .2s;outline:none;width:100%}
input:focus,textarea:focus{border-color:var(--blue)}
textarea{resize:vertical;min-height:60px}
.field-help{font-size:12px;color:var(--muted);line-height:1.45}
.field-error{display:none;font-size:12px;color:var(--danger);font-weight:700}
.field-error.show{display:block}
.required-pill{display:inline-flex;align-items:center;padding:4px 9px;border-radius:999px;background:rgba(255,181,71,.12);color:var(--amber);font-size:10px;font-weight:800;letter-spacing:.8px;text-transform:uppercase}

/* ── product rows table ── */
.rows-wrap{overflow:visible}
.rows-table{width:100%;border-collapse:collapse;margin-bottom:12px}
.rows-table th{
  text-align:left;padding:9px 12px;color:var(--muted);font-size:11px;
  font-weight:600;letter-spacing:1px;text-transform:uppercase;
  border-bottom:1px solid var(--border);white-space:nowrap}
.rows-table td{padding:8px 8px;vertical-align:middle}
.rows-table tr.data-row:last-child td{border-bottom:none}
.rows-table tr.data-row td{border-bottom:1px solid var(--border)}

/* ── searchable picker ── */
.picker{position:relative}
.picker-input{width:100%;min-width:280px}
.picker-list{
  position:absolute;top:calc(100% + 4px);left:0;z-index:300;
  min-width:380px;width:max-content;max-width:520px;
  background:var(--card2);border:1px solid var(--border2);border-radius:10px;
  max-height:280px;overflow-y:auto;display:none;box-shadow:0 8px 32px rgba(0,0,0,.5)}
.picker-list.open{display:block}
.picker-item{padding:10px 16px;cursor:pointer;font-size:13px;border-radius:6px;display:flex;align-items:center;gap:8px}
.picker-item:hover,.picker-item.highlighted{background:rgba(77,159,255,.12);color:var(--blue)}
.picker-item .sku{font-family:var(--mono);font-size:11px;color:var(--muted)}
.picker-item .stock{margin-left:auto;font-family:var(--mono);font-size:11px;color:var(--muted);white-space:nowrap}
.picker-empty{padding:12px 16px;color:var(--muted);font-size:13px}

/* ── fancy combobox (fpicker) — used for supplier/storage/farm/category dropdowns ── */
.fpicker{position:relative;width:100%}
.fpicker-trigger{
  display:flex;align-items:center;gap:10px;width:100%;
  background:var(--surface);border:1px solid var(--border);border-radius:10px;
  color:var(--text);font-family:var(--sans);font-size:14px;
  padding:9px 13px;cursor:pointer;transition:all .15s;text-align:left;
  min-height:40px;
}
.fpicker-trigger:hover{border-color:var(--border2)}
.fpicker-trigger:focus,.fpicker-trigger.open{outline:none;border-color:var(--blue);box-shadow:0 0 0 3px rgba(77,159,255,.10)}
.fpicker-trigger.invalid{border-color:rgba(255,77,109,.55);box-shadow:0 0 0 3px rgba(255,77,109,.10)}
.fpicker-label{flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.fpicker-label.placeholder{color:var(--muted)}
.fpicker-meta{font-family:var(--mono);font-size:11px;color:var(--muted);flex-shrink:0}
.fpicker-icon{flex-shrink:0;width:16px;height:16px;color:var(--muted);transition:transform .2s}
.fpicker-trigger.open .fpicker-icon{transform:rotate(180deg);color:var(--blue)}
.fpicker-clear{
  flex-shrink:0;width:18px;height:18px;border:none;background:transparent;
  color:var(--muted);font-size:14px;font-weight:700;cursor:pointer;
  border-radius:50%;display:none;align-items:center;justify-content:center;
  padding:0;line-height:1;transition:all .15s;
}
.fpicker-clear:hover{background:rgba(255,77,109,.15);color:var(--danger)}
.fpicker:hover .fpicker-clear.has-value{display:flex}
.fpicker-trigger.open .fpicker-clear.has-value{display:flex}

.fpicker-panel{
  position:absolute;top:calc(100% + 6px);left:0;right:0;z-index:600;
  background:var(--card2);border:1px solid var(--border2);border-radius:12px;
  box-shadow:0 12px 40px rgba(0,0,0,.5);
  max-height:340px;display:none;flex-direction:column;overflow:hidden;
  animation:fpickerOpen .15s ease-out;
}
.fpicker-panel.open{display:flex}
@keyframes fpickerOpen{from{opacity:0;transform:translateY(-4px)}to{opacity:1;transform:translateY(0)}}
.fpicker-search{
  flex-shrink:0;padding:10px 12px;border-bottom:1px solid var(--border);
  display:flex;align-items:center;gap:8px;background:var(--card);
}
.fpicker-search-icon{width:14px;height:14px;color:var(--muted);flex-shrink:0}
.fpicker-search input{
  flex:1;background:transparent;border:none;outline:none;
  color:var(--text);font-family:var(--sans);font-size:13px;padding:2px 0;
}
.fpicker-search input::placeholder{color:var(--muted)}
.fpicker-list{flex:1;overflow-y:auto;padding:6px 0;min-height:0}
.fpicker-item{
  display:flex;align-items:center;gap:10px;padding:9px 14px;
  cursor:pointer;font-size:13px;color:var(--text);
  transition:background .12s;border-left:2px solid transparent;
}
.fpicker-item:hover,.fpicker-item.highlighted{background:rgba(77,159,255,.10)}
.fpicker-item.selected{border-left-color:var(--blue);background:rgba(77,159,255,.06)}
.fpicker-item.selected::after{content:"✓";margin-left:auto;color:var(--blue);font-weight:700}
.fpicker-item-emoji{font-size:16px;line-height:1;flex-shrink:0}
.fpicker-item-main{flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.fpicker-item-sub{font-family:var(--mono);font-size:11px;color:var(--muted);flex-shrink:0}
.fpicker-empty{padding:18px 14px;text-align:center;color:var(--muted);font-size:13px;font-style:italic}
.fpicker-section-label{
  padding:8px 14px 4px;font-size:10px;font-weight:700;
  color:var(--muted);letter-spacing:1px;text-transform:uppercase;
}

body.light .fpicker-trigger{background:#fff}
body.light .fpicker-panel{background:#fff;border-color:rgba(0,0,0,.12)}
body.light .fpicker-search{background:#fafafa}

/* ── row inputs ── */
.qty-input,.cost-input{width:100px;text-align:right;font-family:var(--mono)}
.unit-cell{font-size:12px;color:var(--sub);white-space:nowrap;padding:0 8px}
.row-total{font-family:var(--mono);font-size:13px;color:var(--amber);
  text-align:right;white-space:nowrap;padding:0 8px;min-width:80px}
.remove-btn{background:transparent;border:1px solid var(--border);color:var(--muted);
  width:28px;height:28px;border-radius:8px;cursor:pointer;font-size:15px;
  display:flex;align-items:center;justify-content:center;transition:all .2s}
.remove-btn:hover{border-color:var(--danger);color:var(--danger)}
.remove-btn:disabled{opacity:.3;cursor:default}

/* ── footer row ── */
.add-row-btn{
  display:inline-flex;align-items:center;gap:6px;
  background:transparent;border:1px dashed var(--border2);color:var(--sub);
  font-family:var(--sans);font-size:13px;font-weight:500;
  padding:8px 16px;border-radius:10px;cursor:pointer;transition:all .2s}
.add-row-btn:hover{border-color:var(--blue);color:var(--blue)}
.grand-total{
  display:flex;align-items:center;justify-content:flex-end;gap:12px;
  margin-top:16px;padding-top:14px;border-top:1px solid var(--border)}
.grand-total-label{font-size:12px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:1px}
.grand-total-value{font-family:var(--mono);font-size:20px;font-weight:700;color:var(--amber)}
.submit-btn{
  width:100%;padding:13px;border-radius:12px;border:none;cursor:pointer;
  background:var(--green);color:#0a0c08;font-family:var(--sans);
  font-size:14px;font-weight:700;letter-spacing:.3px;transition:all .2s;margin-top:16px}
.submit-btn:hover:not(:disabled){filter:brightness(1.1);transform:translateY(-1px)}
.submit-btn:active:not(:disabled){transform:translateY(0)}
.submit-btn:disabled{opacity:.4;cursor:not-allowed}

/* ── history table ── */
.table-wrap{overflow-x:auto}
table.hist{width:100%;border-collapse:collapse;font-size:13px}
table.hist th{text-align:left;padding:9px 13px;color:var(--muted);font-size:11px;
  font-weight:600;letter-spacing:1px;text-transform:uppercase;
  border-bottom:1px solid var(--border);white-space:nowrap}
table.hist td{padding:10px 13px;border-bottom:1px solid var(--border);vertical-align:middle}
table.hist tr:last-child td{border-bottom:none}
table.hist tr:hover td{background:rgba(255,255,255,.025)}
body.light table.hist tr:hover td{background:rgba(0,0,0,.03)}
.badge{display:inline-block;padding:3px 9px;border-radius:6px;
  font-family:var(--mono);font-size:11px;font-weight:600;
  background:rgba(77,159,255,.12);color:var(--blue);border:1px solid rgba(77,159,255,.2)}
.badge-exp{background:rgba(0,255,157,.1);color:var(--green);border-color:rgba(0,255,157,.2)}
.badge-none{background:rgba(68,80,102,.2);color:var(--muted);border-color:transparent}
.empty-row{text-align:center;padding:40px;color:var(--muted)}
.history-actions{display:flex;gap:8px;flex-wrap:wrap}
.action-btn{background:transparent;border:1px solid var(--border2);color:var(--sub);padding:6px 10px;border-radius:8px;cursor:pointer;font-size:12px;font-family:var(--sans);transition:all .2s}
.action-btn:hover{border-color:var(--blue);color:var(--blue)}
.action-btn.danger:hover{border-color:var(--danger);color:var(--danger)}
.action-btn.export{border-color:rgba(0,255,157,.25);color:var(--green)}
.action-btn.export:hover{border-color:var(--green)}

/* modal */
.modal-wrap{position:fixed;inset:0;background:rgba(4,7,14,.72);backdrop-filter:blur(8px);display:none;align-items:center;justify-content:center;padding:24px;z-index:999}
.modal-wrap.open{display:flex}
.modal-card{width:min(560px,100%);background:var(--card);border:1px solid var(--border2);border-radius:18px;padding:24px;box-shadow:0 24px 50px rgba(0,0,0,.45)}
.modal-title{font-size:18px;font-weight:800;margin-bottom:6px}
.modal-sub{font-size:13px;color:var(--muted);margin-bottom:18px}
.modal-actions{display:flex;justify-content:flex-end;gap:10px;margin-top:18px}
.modal-btn{border:none;border-radius:10px;padding:10px 16px;font-family:var(--sans);font-size:13px;font-weight:700;cursor:pointer}
.modal-btn.secondary{background:var(--card2);color:var(--sub);border:1px solid var(--border2)}
.modal-btn.primary{background:var(--green);color:#0a0c08}

/* ── toast ── */
.toast{position:fixed;bottom:24px;right:24px;z-index:9999;
  padding:12px 20px;border-radius:12px;font-size:13px;font-weight:500;
  box-shadow:0 8px 32px rgba(0,0,0,.4);transition:opacity .3s;opacity:0;pointer-events:none}
.toast.show{opacity:1;pointer-events:auto}
.toast.ok{background:#0f2918;border:1px solid var(--green);color:var(--green)}
.toast.err{background:#240f14;border:1px solid var(--danger);color:var(--danger)}
</style>
    <script src="/static/auth-guard.js"></script>
</head>
<body>

""" + render_app_header(current_user, "page_receive_products") + """

<div class="page">

  <div class="page-header">
    <h1>&#128507; Receive Products</h1>
    <p>Add one or more products, enter quantities and costs, then submit. Costs are posted automatically as expenses.</p>
  </div>

  <!-- ── Receive form ── -->
  <div class="card" id="receive-form-card">
    <div class="section-title">New Receipt</div>

    <form id="receive-form" onsubmit="submitBatch(event)">

      <!-- batch-level fields -->
      <div class="meta-grid">
        <div class="field">
          <label>Receive Date *</label>
          <input type="date" id="receive-date" required>
        </div>
        <div class="field">
          <label>Supplier <span style="color:var(--muted);font-weight:400">(optional — for credit tracking)</span></label>
          <div class="fpicker" id="fp-supplier">
            <input type="hidden" id="supplier-select" value="">
          </div>
        </div>
        <div class="field">
          <label>Supplier Ref / Invoice <span style="color:var(--muted);font-weight:400">(optional)</span></label>
          <input type="text" id="supplier-ref" maxlength="150" placeholder="e.g. INV-2026-001">
        </div>
        <div class="field">
          <label>Storage <span style="color:var(--muted);font-weight:400">(where to put the received stock)</span></label>
          <div class="fpicker" id="fp-location">
            <input type="hidden" id="location-select" value="">
          </div>
        </div>
        <div class="field">
          <label>Cost Allocation <span style="color:var(--muted);font-weight:400">(which farm or "🐾 Animals" this expense belongs to)</span></label>
          <div class="fpicker" id="fp-farm">
            <input type="hidden" id="farm-select" value="">
          </div>
        </div>
        <div class="field">
          <label>Expense Category <span class="required-pill" style="margin-left:6px">Required</span></label>
          <div class="fpicker" id="fp-expense-category">
            <input type="hidden" id="expense-category-select" value="">
          </div>
          <div class="field-help" id="expense-category-help">Pick which category in Expenses this purchase posts to.</div>
          <div class="field-error" id="expense-category-error">Pick an expense category before receiving stock.</div>
        </div>
        <div class="field full" id="payment-block" style="display:none">
          <label>Payment</label>
          <div style="display:flex;flex-wrap:wrap;gap:14px;align-items:center;margin-top:6px">
            <label style="display:flex;align-items:center;gap:6px;font-size:13px;color:var(--text);cursor:pointer">
              <input type="radio" name="pay-mode" value="cash" onchange="onPayModeChange()" checked> Cash now
            </label>
            <label style="display:flex;align-items:center;gap:6px;font-size:13px;color:var(--text);cursor:pointer">
              <input type="radio" name="pay-mode" value="credit" onchange="onPayModeChange()"> On account (settle later)
            </label>
            <label style="display:flex;align-items:center;gap:6px;font-size:13px;color:var(--text);cursor:pointer">
              <input type="radio" name="pay-mode" value="partial" onchange="onPayModeChange()"> Partial payment
            </label>
          </div>
          <div id="partial-row" style="display:none;margin-top:10px;align-items:center;gap:10px">
            <label style="font-size:12px;color:var(--muted);min-width:120px">Amount paid now</label>
            <input type="number" id="amount-paid" step="0.01" min="0" placeholder="0.00" style="max-width:180px" oninput="updateRemainingHint()">
            <span style="font-size:12px;color:var(--muted)" id="remaining-hint"></span>
          </div>
        </div>
        <div class="field full">
          <label>Notes <span style="color:var(--muted);font-weight:400">(optional)</span></label>
          <textarea id="notes" maxlength="500" placeholder="Any additional details…" rows="2"></textarea>
        </div>
      </div>

      <!-- product rows -->
      <div class="rows-wrap">
        <table class="rows-table">
          <thead>
            <tr>
              <th style="min-width:320px">Product</th>
              <th>Qty</th>
              <th></th><!-- unit label -->
              <th>Unit Cost</th>
              <th style="text-align:right">Total</th>
              <th></th><!-- remove -->
            </tr>
          </thead>
          <tbody id="rows-body"></tbody>
        </table>
      </div>

      <button type="button" class="add-row-btn" id="add-row-btn" onclick="addRow()">&#43; Add Product</button>

      <div class="grand-total">
        <span class="grand-total-label">Grand Total</span>
        <span class="grand-total-value" id="grand-total">—</span>
      </div>

      <button type="submit" class="submit-btn" id="submit-btn" disabled>
        &#10003; Receive Stock
      </button>

    </form>
  </div>

  <!-- ── Recent received products ── -->
  <div class="card">
    <div class="section-head">
      <div class="section-title">Recent Received Products</div>
      <button type="button" class="action-btn export" id="export-btn" onclick="exportReceipts()" style="display:none">Export Excel</button>
    </div>
    <div class="table-wrap">
      <table class="hist">
        <thead>
          <tr>
            <th>Receipt #</th><th>Date</th><th>Product</th>
            <th>Qty</th><th>Unit Cost</th><th>Total</th>
            <th>Storage</th>
            <th>Expense</th><th>Supplier</th><th>Payment</th><th>Supplier Ref</th><th>Notes</th><th>By</th><th>Actions</th>
          </tr>
        </thead>
        <tbody id="history-body">
          <tr><td colspan="14" class="empty-row">Loading…</td></tr>
        </tbody>
      </table>
    </div>
  </div>

</div>

<div class="modal-wrap" id="edit-modal">
  <div class="modal-card">
    <div class="modal-title">Edit Receipt</div>
    <div class="modal-sub" id="edit-modal-sub">Update this recent received product entry.</div>
    <div class="meta-grid">
      <div class="field">
        <label>Product</label>
        <input type="text" id="edit-product" readonly>
      </div>
      <div class="field">
        <label>Receive Date *</label>
        <input type="date" id="edit-date" required>
      </div>
      <div class="field">
        <label>Quantity *</label>
        <input type="number" id="edit-qty" min="0.001" step="0.001" required>
      </div>
      <div class="field">
        <label>Unit Cost</label>
        <input type="number" id="edit-cost" min="0" step="0.01">
      </div>
      <div class="field full">
        <label>Supplier / Reference</label>
        <input type="text" id="edit-supplier" maxlength="150">
      </div>
      <div class="field full">
        <label>Notes</label>
        <textarea id="edit-notes" maxlength="500" rows="3"></textarea>
      </div>
    </div>
    <div class="modal-actions">
      <button type="button" class="modal-btn secondary" onclick="closeEditModal()">Cancel</button>
      <button type="button" class="modal-btn primary" id="edit-save-btn" onclick="saveEditReceipt()">Save Changes</button>
    </div>
  </div>
</div>

<div class="toast" id="toast"></div>

<script>

// ── State ───────────────────────────────────────────────────────────────────
let _products  = [];   // [{id, sku, name, unit, cost, stock}, …]
let _rowSeq    = 0;    // monotonic counter for unique row IDs
let _currentUserRole = '';
let _currentUserPermissions = new Set();
let _historyItems = [];
let _editingReceipt = null;

function hasPermission(permission) {
  return _currentUserRole === 'admin' || _currentUserPermissions.has(permission);
}

function applyReceivePermissions() {
  const canCreate = hasPermission('action_receive_products_create');
  const canExport = hasPermission('action_receive_products_export');
  const formCard = document.getElementById('receive-form-card');
  const addRowBtn = document.getElementById('add-row-btn');
  const submitBtn = document.getElementById('submit-btn');
  const exportBtn = document.getElementById('export-btn');
  if (formCard) formCard.style.display = canCreate ? '' : 'none';
  if (addRowBtn) addRowBtn.style.display = canCreate ? '' : 'none';
  if (submitBtn) submitBtn.style.display = canCreate ? '' : 'none';
  if (exportBtn) exportBtn.style.display = canExport ? '' : 'none';
}

// ── Bootstrap ───────────────────────────────────────────────────────────────
async function init() {
  if (localStorage.getItem('colorMode') === 'light') {
    document.body.classList.add('light');
    document.getElementById('mode-btn').innerHTML = '&#9728;&#65039;';
  }
  // Build the four main-form pickers before loaders populate them
  _fpickers.supplier = new FPicker('fp-supplier', 'supplier-select', {
    placeholder: '— No supplier (cash purchase) —',
    searchPlaceholder: 'Search suppliers…',
    emptyText: 'No suppliers found',
    onChange: () => onSupplierChange(),
  });
  _fpickers.location = new FPicker('fp-location', 'location-select', {
    placeholder: 'Loading storages…',
    searchPlaceholder: 'Search storages…',
    emptyText: 'No storages',
    clearable: false,
  });
  _fpickers.farm = new FPicker('fp-farm', 'farm-select', {
    placeholder: '— General expense —',
    searchPlaceholder: 'Search farms…',
    emptyText: 'No matches',
  });
  _fpickers.category = new FPicker('fp-expense-category', 'expense-category-select', {
    placeholder: '— Select expense category —',
    searchPlaceholder: 'Search categories…',
    emptyText: 'No categories',
    clearable: false,
    onChange: () => onExpenseCategoryChange(),
  });

  await Promise.all([initUser(), loadProducts(), loadSuppliers(), loadLocations(), loadFarms(), loadExpenseCategories()]);
  document.getElementById('receive-date').value = todayIso();
  addRow();          // start with one empty row
  await loadHistory();
}

let _suppliers = [];
let _locations = [];
let _farms = [];

async function loadFarms() {
  let failed = false;
  try {
    const r = await fetch('/receive/api/farms');
    if (r.ok) {
      const data = await r.json();
      _farms = Array.isArray(data) ? data : (data && data.items) ? data.items : [];
    } else {
      _farms = [];
      failed = true;
    }
  } catch (_) { _farms = []; failed = true; }
  if (_fpickers.farm) {
    if (failed) {
      // Surface failure instead of letting the picker stay at its initial label.
      _fpickers.farm.setOptions([
        { value: '',           label: 'General expense', emoji: '🏷️', meta: 'no allocation' },
        { value: '__animals__', label: 'Animals',        emoji: '🐾', meta: 'animal bucket' },
      ]);
    } else {
      _fpickers.farm.setOptions([
        { value: '',           label: 'General expense', emoji: '🏷️', meta: 'no allocation' },
        ..._farms.map(f => ({
          value: String(f.id),
          label: f.name || ('Farm #' + f.id),
          emoji: '🌾',
          meta: f.location || '',
        })),
        { value: '__animals__', label: 'Animals',        emoji: '🐾', meta: 'animal bucket' },
      ]);
    }
  }
}

async function loadLocations() {
  let failed = false;
  try {
    const r = await fetch('/receive/api/locations');
    if (r.ok) {
      const data = await r.json();
      _locations = (data && data.items) ? data.items : (Array.isArray(data) ? data : []);
    } else {
      _locations = [];
      failed = true;
    }
  } catch (_) { _locations = []; failed = true; }
  if (_fpickers.location) {
    if (failed) {
      // Don't leave the picker stuck on "Loading storages…" — give the user
      // a clear failure state with no options.
      _fpickers.location.setOptions([]);
      _fpickers.location.placeholder = 'Could not load storages — refresh to retry';
      _fpickers.location.setValue('');
      showToast('Could not load storages. Refresh the page to retry.', 'err');
    } else if (!_locations.length) {
      _fpickers.location.setOptions([]);
      _fpickers.location.placeholder = 'No storages defined — ask an admin to create one';
      _fpickers.location.setValue('');
    } else {
      _fpickers.location.setOptions(_locations.map(l => ({
        value: String(l.id),
        label: l.name,
        emoji: '📦',
      })));
      // Default to Main Warehouse if present; otherwise first location.
      let defaultId = '';
      const main = _locations.find(l => (l.name || '').toLowerCase().includes('main'));
      if (main) defaultId = String(main.id);
      else if (_locations[0]) defaultId = String(_locations[0].id);
      if (defaultId) _fpickers.location.setValue(defaultId);
    }
  }
}

async function loadSuppliers() {
  try {
    const r = await fetch('/receive/api/suppliers');
    if (r.ok) {
      _suppliers = await r.json();
    } else {
      _suppliers = [];
    }
  } catch (_) { _suppliers = []; }
  if (_fpickers.supplier) {
    _fpickers.supplier.setOptions([
      { value: '', label: 'No supplier (cash purchase)', emoji: '💵' },
      ..._suppliers.map(s => {
        const bal = Number(s.balance || 0);
        return {
          value: String(s.id),
          label: s.name,
          emoji: '🏢',
          meta:  bal > 0 ? ('owed ' + bal.toFixed(2)) : '',
        };
      }),
    ]);
  }
}

let _expenseCategories = [];

async function loadExpenseCategories() {
  try {
    const r = await fetch('/receive/api/expense-categories');
    if (r.ok) {
      _expenseCategories = await r.json();
    } else {
      _expenseCategories = [];
    }
  } catch (_) { _expenseCategories = []; }
  if (_fpickers.category) {
    _fpickers.category.setOptions(_expenseCategories.map(c => ({
      value: String(c.id),
      label: c.name,
      emoji: '💸',
      meta:  c.account_code || '',
    })));
  }
}

function onExpenseCategoryChange() {
  // Clear validation styling once a category is picked.
  const val = document.getElementById('expense-category-select').value;
  const errEl = document.getElementById('expense-category-error');
  if (val) {
    if (_fpickers.category) _fpickers.category.setInvalid(false);
    if (errEl) errEl.classList.remove('show');
  }
  validateSubmit();
}

function escHtml(s){
  return String(s ?? '')
    .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')
    .replace(/"/g,'&quot;').replace(/'/g,'&#39;');
}

// ─────────────────────────────────────────────────────────────────────────────
// FPicker — reusable searchable combobox.
//   Drives a hidden <input> with the selected value (so existing
//   document.getElementById('xxx').value lookups still work).
//   Options: { id, value, label, emoji?, meta?, group?, searchText? }
// ─────────────────────────────────────────────────────────────────────────────
class FPicker {
  constructor(hostId, hiddenInputId, opts = {}) {
    this.host = document.getElementById(hostId);
    this.hidden = document.getElementById(hiddenInputId);
    if (!this.host || !this.hidden) throw new Error('FPicker host/hidden missing: ' + hostId);
    this.placeholder = opts.placeholder || 'Select…';
    this.searchPlaceholder = opts.searchPlaceholder || 'Type to search…';
    this.emptyText = opts.emptyText || 'No results';
    this.clearable = opts.clearable !== false;   // default true
    this.options = [];                            // [{value,label,emoji,meta,group,searchText}]
    this.onChange = opts.onChange || (()=>{});
    this._q = '';
    this._highlight = 0;
    this._build();
  }

  _build() {
    this.host.innerHTML = `
      <button type="button" class="fpicker-trigger" id="${this.host.id}-trig">
        <span class="fpicker-label placeholder" id="${this.host.id}-lbl">${escHtml(this.placeholder)}</span>
        <span class="fpicker-meta" id="${this.host.id}-meta" style="display:none"></span>
        <button type="button" class="fpicker-clear" id="${this.host.id}-clr" tabindex="-1" aria-label="Clear">×</button>
        <svg class="fpicker-icon" viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="2"><polyline points="3,6 8,11 13,6"/></svg>
      </button>
      <div class="fpicker-panel" id="${this.host.id}-panel">
        <div class="fpicker-search">
          <svg class="fpicker-search-icon" viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="2"><circle cx="7" cy="7" r="5"/><line x1="11" y1="11" x2="14" y2="14"/></svg>
          <input type="text" id="${this.host.id}-search" placeholder="${escHtml(this.searchPlaceholder)}" autocomplete="off" spellcheck="false">
        </div>
        <div class="fpicker-list" id="${this.host.id}-list"></div>
      </div>
    `;
    // Re-append hidden so the inner HTML didn't drop it
    if (!this.host.contains(this.hidden)) this.host.appendChild(this.hidden);

    this.trig  = document.getElementById(this.host.id + '-trig');
    this.lbl   = document.getElementById(this.host.id + '-lbl');
    this.metaEl= document.getElementById(this.host.id + '-meta');
    this.clrBtn= document.getElementById(this.host.id + '-clr');
    this.panel = document.getElementById(this.host.id + '-panel');
    this.sBox  = document.getElementById(this.host.id + '-search');
    this.list  = document.getElementById(this.host.id + '-list');

    this.trig.addEventListener('click', e => { e.stopPropagation(); this.toggle(); });
    this.trig.addEventListener('keydown', e => {
      if (e.key === 'Enter' || e.key === ' ' || e.key === 'ArrowDown') {
        e.preventDefault();
        this.open();
      }
    });
    this.clrBtn.addEventListener('click', e => {
      e.stopPropagation();
      this.setValue('', { fireChange: true });
    });
    this.sBox.addEventListener('input', () => { this._q = this.sBox.value.trim().toLowerCase(); this._highlight = 0; this._render(); });
    this.sBox.addEventListener('keydown', e => this._onKey(e));
    this.list.addEventListener('mousemove', e => {
      const item = e.target.closest('.fpicker-item');
      if (!item) return;
      const idx = parseInt(item.dataset.idx, 10);
      if (!isNaN(idx) && idx !== this._highlight) {
        this._highlight = idx;
        this._updateHighlight();
      }
    });
    this.list.addEventListener('click', e => {
      const item = e.target.closest('.fpicker-item');
      if (!item) return;
      const idx = parseInt(item.dataset.idx, 10);
      this._pick(idx);
    });
    document.addEventListener('click', e => {
      if (!this.host.contains(e.target)) this.close();
    });
  }

  setOptions(opts) {
    this.options = (opts || []).map(o => ({
      value: String(o.value ?? ''),
      label: o.label ?? '',
      emoji: o.emoji || '',
      meta:  o.meta  || '',
      group: o.group || '',
      searchText: (o.searchText || (o.label + ' ' + (o.meta || ''))).toLowerCase(),
    }));
    // Re-sync current display in case the selected value's metadata changed
    this.setValue(this.hidden.value, { fireChange: false });
  }

  _filtered() {
    if (!this._q) return this.options.map((o, i) => ({ o, i }));
    return this.options
      .map((o, i) => ({ o, i }))
      .filter(({ o }) => o.searchText.includes(this._q));
  }

  open() {
    this.panel.classList.add('open');
    this.trig.classList.add('open');
    this._q = '';
    this.sBox.value = '';
    // Highlight the currently selected item if any
    const cur = this.options.findIndex(o => o.value === this.hidden.value);
    this._highlight = cur >= 0 ? cur : 0;
    this._render();
    setTimeout(() => this.sBox.focus(), 50);
  }

  close() {
    this.panel.classList.remove('open');
    this.trig.classList.remove('open');
  }

  toggle() {
    if (this.panel.classList.contains('open')) this.close();
    else this.open();
  }

  _onKey(e) {
    const filtered = this._filtered();
    if (e.key === 'ArrowDown') {
      e.preventDefault();
      this._highlight = Math.min(this._highlight + 1, filtered.length - 1);
      this._updateHighlight(true);
    } else if (e.key === 'ArrowUp') {
      e.preventDefault();
      this._highlight = Math.max(this._highlight - 1, 0);
      this._updateHighlight(true);
    } else if (e.key === 'Enter') {
      e.preventDefault();
      const visible = filtered[this._highlight];
      if (visible) this._pick(visible.i);
    } else if (e.key === 'Escape') {
      e.preventDefault();
      this.close();
      this.trig.focus();
    }
  }

  _pick(idx) {
    const o = this.options[idx];
    if (!o) return;
    this.setValue(o.value, { fireChange: true });
    this.close();
    this.trig.focus();
  }

  _render() {
    const filtered = this._filtered();
    if (!filtered.length) {
      this.list.innerHTML = `<div class="fpicker-empty">${escHtml(this.emptyText)}</div>`;
      return;
    }
    let html = '';
    let lastGroup = null;
    filtered.forEach(({ o, i }, vIdx) => {
      if (o.group && o.group !== lastGroup) {
        html += `<div class="fpicker-section-label">${escHtml(o.group)}</div>`;
        lastGroup = o.group;
      } else if (!o.group) {
        lastGroup = null;
      }
      const selected = o.value === this.hidden.value;
      const highlight = vIdx === this._highlight;
      html += `
        <div class="fpicker-item${selected ? ' selected' : ''}${highlight ? ' highlighted' : ''}" data-idx="${i}">
          ${o.emoji ? `<span class="fpicker-item-emoji">${escHtml(o.emoji)}</span>` : ''}
          <span class="fpicker-item-main">${escHtml(o.label)}</span>
          ${o.meta ? `<span class="fpicker-item-sub">${escHtml(o.meta)}</span>` : ''}
        </div>
      `;
    });
    this.list.innerHTML = html;
    // Scroll the highlighted item into view
    const hEl = this.list.querySelector('.highlighted');
    if (hEl && hEl.scrollIntoView) hEl.scrollIntoView({ block: 'nearest' });
  }

  _updateHighlight(scroll) {
    const items = this.list.querySelectorAll('.fpicker-item');
    items.forEach((el, idx) => el.classList.toggle('highlighted', idx === this._highlight));
    if (scroll) {
      const hEl = items[this._highlight];
      if (hEl && hEl.scrollIntoView) hEl.scrollIntoView({ block: 'nearest' });
    }
  }

  setValue(value, { fireChange = false } = {}) {
    const v = String(value ?? '');
    this.hidden.value = v;
    const o = this.options.find(x => x.value === v);
    if (o) {
      this.lbl.textContent = o.label;
      this.lbl.classList.remove('placeholder');
      if (o.emoji) this.lbl.textContent = o.emoji + ' ' + o.label;
      if (o.meta) {
        this.metaEl.textContent = o.meta;
        this.metaEl.style.display = '';
      } else {
        this.metaEl.style.display = 'none';
      }
      this.clrBtn.classList.add('has-value');
    } else {
      this.lbl.textContent = this.placeholder;
      this.lbl.classList.add('placeholder');
      this.metaEl.style.display = 'none';
      this.clrBtn.classList.remove('has-value');
    }
    // Make the clear button visible only when clearable AND there's a value
    if (!this.clearable) this.clrBtn.style.display = 'none';

    if (fireChange) this.onChange(v, o);
  }

  getValue() { return this.hidden.value; }

  setInvalid(on) {
    this.trig.classList.toggle('invalid', !!on);
  }
}

// Holder for the four main-form pickers
const _fpickers = {};

function onSupplierChange(){
  const v = document.getElementById('supplier-select').value;
  const block = document.getElementById('payment-block');
  if (v) {
    block.style.display = '';
    // default to "credit" when a supplier is selected — it's the whole point
    const credit = document.querySelector('input[name="pay-mode"][value="credit"]');
    if (credit) credit.checked = true;
    onPayModeChange();
  } else {
    block.style.display = 'none';
  }
}

function onPayModeChange(){
  const mode = (document.querySelector('input[name="pay-mode"]:checked') || {}).value || 'cash';
  const partial = document.getElementById('partial-row');
  partial.style.display = (mode === 'partial') ? 'flex' : 'none';
  if (mode !== 'partial') document.getElementById('amount-paid').value = '';
  updateRemainingHint();
}

function updateRemainingHint(){
  const total = computeGrandTotalRaw();
  const hint = document.getElementById('remaining-hint');
  if (!hint) return;
  const paid = parseFloat(document.getElementById('amount-paid').value || 0) || 0;
  const remaining = Math.max(total - paid, 0);
  if (total > 0) {
    hint.textContent = `Total ${total.toFixed(2)} · Remaining on account: ${remaining.toFixed(2)}`;
  } else {
    hint.textContent = '';
  }
}

function computeGrandTotalRaw(){
  let total = 0;
  document.querySelectorAll('.data-row').forEach(tr => {
    const id = tr.dataset.row;
    const qty = parseFloat((document.getElementById(`qty-${id}`)||{}).value) || 0;
    const cost = parseFloat((document.getElementById(`cost-${id}`)||{}).value) || 0;
    if (qty > 0 && cost > 0) total += qty * cost;
  });
  return total;
}

async function initUser() {
  try {
    const r = await fetch('/auth/me');
    if (!r.ok) { _redirectToLogin(); return; }
    const u = await r.json();
    _currentUserRole = u.role || '';
    _currentUserPermissions = new Set(
      (typeof u.permissions === 'string' ? u.permissions.split(',') : (u.permissions || []))
        .map(v => v.trim())
        .filter(Boolean)
    );
    document.getElementById('user-name').innerText   = u.name;
    document.getElementById('user-avatar').innerText = u.name.charAt(0).toUpperCase();
    const emailEl = document.getElementById('user-email');
    if (emailEl) emailEl.innerText = u.email;
    applyReceivePermissions();
  } catch { _redirectToLogin(); }
}

function toggleAccountMenu(event){
  event.stopPropagation();
  const trigger = document.getElementById('account-trigger');
  const dropdown = document.getElementById('account-dropdown');
  const open = dropdown.classList.toggle('open');
  trigger.classList.toggle('open', open);
  trigger.setAttribute('aria-expanded', open ? 'true' : 'false');
}

document.addEventListener('click', e => {
  const menu = document.getElementById('account-dropdown');
  const trigger = document.getElementById('account-trigger');
  if(!menu || !trigger) return;
  if(menu.contains(e.target) || trigger.contains(e.target)) return;
  menu.classList.remove('open');
  trigger.classList.remove('open');
  trigger.setAttribute('aria-expanded', 'false');
});

async function loadProducts() {
  const r = await fetch('/receive/api/products');
  if (!r.ok) return;
  _products = await r.json();
}

// ── Row management ──────────────────────────────────────────────────────────
function addRow() {
  if (!hasPermission('action_receive_products_create')) return;
  const id  = ++_rowSeq;
  const tr  = document.createElement('tr');
  tr.className = 'data-row';
  tr.dataset.row = id;
  tr.innerHTML = `
    <td>
      <div class="picker" id="picker-${id}">
        <input type="text" class="picker-input" id="psearch-${id}"
               autocomplete="off" placeholder="Search product…"
               oninput="onPickerInput(${id})"
               onfocus="onPickerFocus(${id})"
               onkeydown="onPickerKey(event,${id})">
        <input type="hidden" id="pid-${id}">
        <div class="picker-list" id="plist-${id}"></div>
      </div>
    </td>
    <td><input type="number" class="qty-input" id="qty-${id}"
               min="0.001" step="0.001" placeholder="0.000"
               oninput="recalcRow(${id})"></td>
    <td class="unit-cell" id="unit-${id}">—</td>
    <td><input type="number" class="cost-input" id="cost-${id}"
               min="0" step="0.01" placeholder="0.00"
               oninput="recalcRow(${id})"></td>
    <td class="row-total" id="total-${id}">—</td>
    <td>
      <button type="button" class="remove-btn" id="rem-${id}"
              onclick="removeRow(${id})" title="Remove row">&#215;</button>
    </td>`;
  document.getElementById('rows-body').appendChild(tr);
  refreshRemoveButtons();
  document.getElementById(`psearch-${id}`).focus();
}

function removeRow(id) {
  const tr = document.querySelector(`tr[data-row="${id}"]`);
  if (tr) tr.remove();
  refreshRemoveButtons();
  updateGrandTotal();
  validateSubmit();
}

function refreshRemoveButtons() {
  const rows = document.querySelectorAll('#rows-body tr.data-row');
  rows.forEach(r => {
    const btn = r.querySelector('.remove-btn');
    if (btn) btn.disabled = rows.length === 1;
  });
}

// ── Picker logic ────────────────────────────────────────────────────────────
function onPickerInput(id) {
  const query = document.getElementById(`psearch-${id}`).value.trim().toLowerCase();
  renderPickerList(id, query);
  // Clear selection if text was changed
  document.getElementById(`pid-${id}`).value = '';
  recalcRow(id);
  validateSubmit();
}

function onPickerFocus(id) {
  const query = document.getElementById(`psearch-${id}`).value.trim().toLowerCase();
  renderPickerList(id, query);
}

function renderPickerList(id, query) {
  const list = document.getElementById(`plist-${id}`);
  const hits  = query
    ? _products.filter(p =>
        p.name.toLowerCase().includes(query) ||
        p.sku.toLowerCase().includes(query))
    : _products;

  if (hits.length === 0) {
    list.innerHTML = `<div class="picker-empty">No products found</div>`;
  } else {
    list.innerHTML = hits.map(p => `
      <div class="picker-item" data-id="${p.id}"
           onmousedown="selectProduct(event,${id},${p.id},'${esc(p.name)}','${esc(p.sku)}','${esc(p.unit)}',${p.cost},${p.stock})">
        <span>${esc(p.name)}</span>
        <span class="sku">${esc(p.sku)}</span>
        <span class="stock">${p.stock.toFixed(3)}&thinsp;${esc(p.unit)}</span>
      </div>`).join('');
  }
  list.classList.add('open');
}

function selectProduct(e, rowId, productId, name, sku, unit, cost, stock) {
  e.preventDefault();
  document.getElementById(`psearch-${rowId}`).value  = `${name}`;
  document.getElementById(`pid-${rowId}`).value      = productId;
  document.getElementById(`unit-${rowId}`).textContent = unit;
  document.getElementById(`plist-${rowId}`).classList.remove('open');

  // Pre-fill cost with last known cost if field is empty or zero
  const costInput = document.getElementById(`cost-${rowId}`);
  if ((!costInput.value || parseFloat(costInput.value) === 0) && cost > 0) {
    costInput.value = cost.toFixed(2);
  }

  recalcRow(rowId);
  validateSubmit();
  // Move focus to qty
  const qtyInput = document.getElementById(`qty-${rowId}`);
  if (!qtyInput.value) qtyInput.focus();
}

// Close picker when clicking outside
document.addEventListener('click', e => {
  document.querySelectorAll('.picker-list.open').forEach(list => {
    if (!list.closest('.picker').contains(e.target)) {
      list.classList.remove('open');
    }
  });
});

// Basic keyboard navigation in picker
function onPickerKey(e, id) {
  const list  = document.getElementById(`plist-${id}`);
  const items = list.querySelectorAll('.picker-item');
  if (!items.length) return;

  let cur = Array.from(items).findIndex(i => i.classList.contains('highlighted'));

  if (e.key === 'ArrowDown') {
    e.preventDefault();
    cur = Math.min(cur + 1, items.length - 1);
    items.forEach((item, i) => item.classList.toggle('highlighted', i === cur));
    items[cur]?.scrollIntoView({block:'nearest'});
  } else if (e.key === 'ArrowUp') {
    e.preventDefault();
    cur = Math.max(cur - 1, 0);
    items.forEach((item, i) => item.classList.toggle('highlighted', i === cur));
    items[cur]?.scrollIntoView({block:'nearest'});
  } else if (e.key === 'Enter') {
    e.preventDefault();
    if (cur >= 0) items[cur]?.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
  } else if (e.key === 'Escape') {
    list.classList.remove('open');
  }
}

// ── Totals ──────────────────────────────────────────────────────────────────
function recalcRow(id) {
  const qty  = parseFloat(document.getElementById(`qty-${id}`)?.value)  || 0;
  const cost = parseFloat(document.getElementById(`cost-${id}`)?.value) || 0;
  const el   = document.getElementById(`total-${id}`);
  if (!el) return;
  if (qty > 0 && cost > 0) {
    el.textContent = (qty * cost).toFixed(2);
    el.style.color = 'var(--amber)';
  } else {
    el.textContent = '—';
    el.style.color = 'var(--muted)';
  }
  updateGrandTotal();
  validateSubmit();
}

function updateGrandTotal() {
  let sum = 0;
  document.querySelectorAll('#rows-body tr.data-row').forEach(tr => {
    const id    = tr.dataset.row;
    const qty   = parseFloat(document.getElementById(`qty-${id}`)?.value)  || 0;
    const cost  = parseFloat(document.getElementById(`cost-${id}`)?.value) || 0;
    if (qty > 0 && cost > 0) sum += qty * cost;
  });
  const el = document.getElementById('grand-total');
  if (sum > 0) { el.textContent = sum.toFixed(2); el.style.color = 'var(--amber)'; }
  else         { el.textContent = '—';             el.style.color = 'var(--muted)'; }
  updateRemainingHint();
}

function validateSubmit() {
  const rows = document.querySelectorAll('#rows-body tr.data-row');
  const categoryId = document.getElementById('expense-category-select')?.value;
  const valid = Array.from(rows).some(tr => {
    const id  = tr.dataset.row;
    const pid = document.getElementById(`pid-${id}`)?.value;
    const qty = parseFloat(document.getElementById(`qty-${id}`)?.value) || 0;
    return pid && qty > 0;
  });
  document.getElementById('submit-btn').disabled = !(categoryId && valid);
}

// ── Submit ──────────────────────────────────────────────────────────────────
async function submitBatch(e) {
  e.preventDefault();
  if (!hasPermission('action_receive_products_create')) {
    showToast('Permission denied: action_receive_products_create', 'err');
    return;
  }
  const btn = document.getElementById('submit-btn');
  btn.disabled = true;
  btn.textContent = 'Receiving…';

  const rows    = document.querySelectorAll('#rows-body tr.data-row');
  const items   = [];

  rows.forEach(tr => {
    const id   = tr.dataset.row;
    const pid  = document.getElementById(`pid-${id}`)?.value;
    const qty  = parseFloat(document.getElementById(`qty-${id}`)?.value);
    const cost = document.getElementById(`cost-${id}`)?.value.trim();
    if (!pid || !qty || qty <= 0) return;   // skip incomplete rows
    const item = { product_id: parseInt(pid), qty };
    if (cost) item.unit_cost = parseFloat(cost);
    items.push(item);
  });

  // Validate expense category (required)
  const expenseCatVal = document.getElementById('expense-category-select').value;
  if (!expenseCatVal) {
    const errEl = document.getElementById('expense-category-error');
    if (errEl) errEl.classList.add('show');
    if (_fpickers.category) _fpickers.category.setInvalid(true);
    showToast('Pick an expense category before receiving stock.', 'err');
    btn.disabled = false; btn.textContent = '✓ Receive Stock';
    return;
  }
  if (items.length === 0) {
    showToast('Add at least one product with a quantity.', 'err');
    btn.disabled = false; btn.textContent = '✓ Receive Stock';
    return;
  }

  // Supplier & payment fields
  const supplierVal = document.getElementById('supplier-select').value;
  const supplierId = supplierVal ? parseInt(supplierVal, 10) : null;
  let amountPaid = null;
  if (supplierId) {
    const mode = (document.querySelector('input[name="pay-mode"]:checked') || {}).value || 'cash';
    const grandTotal = computeGrandTotalRaw();
    if (mode === 'cash')         amountPaid = grandTotal;
    else if (mode === 'credit')  amountPaid = 0;
    else if (mode === 'partial') {
      amountPaid = parseFloat(document.getElementById('amount-paid').value) || 0;
      if (amountPaid <= 0) {
        showToast('Enter the amount paid for a partial payment.', 'err');
        btn.disabled = false; btn.textContent = '✓ Receive Stock';
        return;
      }
      if (amountPaid > grandTotal) {
        showToast('Amount paid cannot exceed the grand total.', 'err');
        btn.disabled = false; btn.textContent = '✓ Receive Stock';
        return;
      }
    }
  }

  // Cost allocation: farm or "🐾 Animals" sentinel
  const farmSelVal = document.getElementById('farm-select').value;
  const isAnimalExp = farmSelVal === '__animals__';
  const farmIdVal   = isAnimalExp ? null : (parseInt(farmSelVal, 10) || null);

  // Expense category (required — validated above)
  const expenseCategoryId = parseInt(expenseCatVal, 10);

  const payload = {
    receive_date: document.getElementById('receive-date').value,
    supplier_ref: document.getElementById('supplier-ref').value.trim() || null,
    supplier_id:  supplierId,
    amount_paid:  amountPaid,
    notes:        document.getElementById('notes').value.trim() || null,
    location_id:  parseInt(document.getElementById('location-select').value, 10) || null,
    farm_id:           farmIdVal,
    is_animal_expense: isAnimalExp,
    expense_category_id: expenseCategoryId,
    items,
  };

  try {
    const r = await fetch('/receive/api/receive-batch', {
      method:  'POST',
      headers: {'Content-Type': 'application/json'},
      body:    JSON.stringify(payload),
    });
    if (!r.ok) {
      const err = await r.json().catch(() => ({}));
      showToast(err.detail || 'Receive failed', 'err');
    } else {
      const data = await r.json();
      const expCount = data.receipts.filter(r => r.expense_ref).length;
      const paid   = Number(data.total_paid   || 0);
      const unpaid = Number(data.total_unpaid || 0);
      let msg = `${data.count} product${data.count > 1 ? 's' : ''} received`;
      if (data.total_cost) msg += ` · Total ${data.total_cost.toFixed(2)}`;
      if (expCount)        msg += ` · ${expCount} expense${expCount > 1 ? 's' : ''} posted`;
      if (unpaid > 0)      msg += ` · ${unpaid.toFixed(2)} on account`;
      else if (paid > 0)   msg += ` · paid ${paid.toFixed(2)}`;
      showToast(msg, 'ok');
      resetForm();
      await loadProducts();
      await loadSuppliers();
      await loadHistory();
    }
  } catch { showToast('Network error', 'err'); }

  btn.disabled = false; btn.textContent = '✓ Receive Stock';
}

function resetForm() {
  document.getElementById('supplier-ref').value = '';
  if (_fpickers.supplier) _fpickers.supplier.setValue('');
  if (_fpickers.farm)     _fpickers.farm.setValue('');
  if (_fpickers.category) {
    _fpickers.category.setValue('');
    _fpickers.category.setInvalid(false);
  }
  const errEl = document.getElementById('expense-category-error');
  if (errEl) errEl.classList.remove('show');
  // Storage stays defaulted to Main Warehouse — don't clear it.
  document.getElementById('amount-paid').value = '';
  const cashRadio = document.querySelector('input[name="pay-mode"][value="cash"]');
  if (cashRadio) cashRadio.checked = true;
  document.getElementById('payment-block').style.display = 'none';
  document.getElementById('partial-row').style.display = 'none';
  document.getElementById('notes').value = '';
  document.getElementById('receive-date').value = todayIso();
  document.getElementById('rows-body').innerHTML = '';
  _rowSeq = 0;
  addRow();
  updateGrandTotal();
}

// ── History ─────────────────────────────────────────────────────────────────
async function loadHistory() {
  const r     = await fetch('/receive/api/history?limit=100');
  const tbody = document.getElementById('history-body');
  if (!r.ok) { tbody.innerHTML = `<tr><td colspan="14" class="empty-row">Could not load.</td></tr>`; return; }
  const data  = await r.json();
  _historyItems = data.items || [];
  const canUpdate = hasPermission('action_receive_products_update');
  const canDelete = hasPermission('action_receive_products_delete');
  const canManage = canUpdate || canDelete;
  if (!_historyItems.length) {
    tbody.innerHTML = `<tr><td colspan="14" class="empty-row">No receipts yet.</td></tr>`;
    return;
  }
  tbody.innerHTML = _historyItems.map(row => {
    let payCell = '<span style="color:var(--muted)">—</span>';
    if (row.total_cost != null && row.total_cost > 0) {
      const total  = parseFloat(row.total_cost);
      const paid   = parseFloat(row.amount_paid || 0);
      const unpaid = parseFloat(row.amount_unpaid || 0);
      if (unpaid > 0 && paid > 0) {
        payCell = `<span style="color:var(--amber)" title="Partial: paid ${paid.toFixed(2)} / unpaid ${unpaid.toFixed(2)}">Partial · ${unpaid.toFixed(2)} owed</span>`;
      } else if (unpaid > 0) {
        payCell = `<span style="color:var(--amber)">On account · ${unpaid.toFixed(2)}</span>`;
      } else {
        payCell = `<span style="color:#3dd06a">Paid</span>`;
      }
    }
    return `<tr>
    <td><span class="badge">${esc(row.ref_number)}</span></td>
    <td>${esc(row.receive_date||'')}</td>
    <td>
      <div style="font-weight:600">${esc(row.product_name||'')}</div>
      <div style="font-family:var(--mono);font-size:11px;color:var(--muted)">${esc(row.product_sku||'')}</div>
    </td>
    <td style="font-family:var(--mono)">${parseFloat(row.qty).toFixed(3)}</td>
    <td style="font-family:var(--mono)">${row.unit_cost!=null ? parseFloat(row.unit_cost).toFixed(2) : '<span style="color:var(--muted)">—</span>'}</td>
    <td style="font-family:var(--mono);color:var(--amber)">${row.total_cost!=null ? parseFloat(row.total_cost).toFixed(2) : '<span style="color:var(--muted)">—</span>'}</td>
    <td style="color:var(--sub)">${row.location_name ? esc(row.location_name) : '<span style="color:var(--muted)">—</span>'}</td>
    <td>${row.expense_ref ? `<span class="badge badge-exp">${esc(row.expense_ref)}</span>` : '<span class="badge badge-none">—</span>'}</td>
    <td style="color:var(--sub)">${row.supplier_name ? esc(row.supplier_name) : '<span style="color:var(--muted)">—</span>'}</td>
    <td>${payCell}</td>
    <td style="color:var(--sub)">${esc(row.supplier_ref||'—')}</td>
    <td style="color:var(--sub);max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap"
        title="${esc(row.notes||'')}">${esc(row.notes||'—')}</td>
    <td style="color:var(--muted)">${esc(row.received_by||'—')}</td>
    <td>${canManage ? `<div class="history-actions">
      ${canUpdate ? `<button type="button" class="action-btn" onclick="openEditModal(${row.id})">Edit</button>` : ''}
      ${canDelete ? `<button type="button" class="action-btn danger" onclick="deleteReceipt(${row.id})">Delete</button>` : ''}
    </div>` : '<span style="color:var(--muted)">No action permission</span>'}</td>
  </tr>`;
  }).join('');
}

// ── Utils ───────────────────────────────────────────────────────────────────
function openEditModal(receiptId) {
  if (!hasPermission('action_receive_products_update')) return;
  const receipt = _historyItems.find(item => item.id === receiptId);
  if (!receipt) {
    showToast('Receipt not found', 'err');
    return;
  }
  _editingReceipt = receipt;
  document.getElementById('edit-modal-sub').textContent = `Editing ${receipt.ref_number}`;
  document.getElementById('edit-product').value = receipt.product_name || '';
  document.getElementById('edit-date').value = receipt.receive_date || todayIso();
  document.getElementById('edit-qty').value = parseFloat(receipt.qty || 0).toFixed(3);
  document.getElementById('edit-cost').value = receipt.unit_cost != null ? parseFloat(receipt.unit_cost).toFixed(2) : '';
  document.getElementById('edit-supplier').value = receipt.supplier_ref || '';
  document.getElementById('edit-notes').value = receipt.notes || '';
  document.getElementById('edit-modal').classList.add('open');
}

function closeEditModal() {
  _editingReceipt = null;
  document.getElementById('edit-modal').classList.remove('open');
}

async function saveEditReceipt() {
  if (!_editingReceipt || !hasPermission('action_receive_products_update')) return;
  const btn = document.getElementById('edit-save-btn');
  btn.disabled = true;
  btn.textContent = 'Saving...';
  const payload = {
    receive_date: document.getElementById('edit-date').value,
    qty: parseFloat(document.getElementById('edit-qty').value),
    supplier_ref: document.getElementById('edit-supplier').value.trim() || null,
    notes: document.getElementById('edit-notes').value.trim() || null,
  };
  const costValue = document.getElementById('edit-cost').value.trim();
  payload.unit_cost = costValue === '' ? null : parseFloat(costValue);

  try {
    const r = await fetch(`/receive/api/receipt/${_editingReceipt.id}`, {
      method: 'PUT',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(payload),
    });
    const data = await r.json().catch(() => ({}));
    if (!r.ok) {
      showToast(data.detail || 'Could not update receipt', 'err');
      return;
    }
    closeEditModal();
    showToast(`Receipt ${data.ref_number} updated`, 'ok');
    await loadProducts();
    await loadHistory();
  } catch {
    showToast('Network error', 'err');
  } finally {
    btn.disabled = false;
    btn.textContent = 'Save Changes';
  }
}

async function deleteReceipt(receiptId) {
  if (!hasPermission('action_receive_products_delete')) return;
  const receipt = _historyItems.find(item => item.id === receiptId);
  if (!receipt) {
    showToast('Receipt not found', 'err');
    return;
  }
  if (!confirm(`Delete receipt ${receipt.ref_number}? This will reverse its stock receipt.`)) return;
  try {
    const r = await fetch(`/receive/api/receipt/${receiptId}`, {method: 'DELETE'});
    const data = await r.json().catch(() => ({}));
    if (!r.ok) {
      showToast(data.detail || 'Could not delete receipt', 'err');
      return;
    }
    if (_editingReceipt && _editingReceipt.id === receiptId) closeEditModal();
    showToast(`Receipt ${receipt.ref_number} deleted`, 'ok');
    await loadProducts();
    await loadHistory();
  } catch {
    showToast('Network error', 'err');
  }
}

function exportReceipts() {
  if (!hasPermission('action_receive_products_export')) return;
  showToast('Preparing Excel export...', 'ok');
  window.location.href = '/receive/api/export.xlsx';
}

function todayIso() { return new Date().toISOString().slice(0, 10); }

function esc(s) {
  return String(s)
    .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')
    .replace(/"/g,'&quot;').replace(/'/g,'&#039;');
}

function showToast(msg, type) {
  const t = document.getElementById('toast');
  t.textContent = msg; t.className = `toast ${type} show`;
  setTimeout(() => t.className = 'toast', 3800);
}

function toggleMode() {
  const light = document.body.classList.toggle('light');
  document.getElementById('mode-btn').innerHTML = light ? '&#9728;&#65039;' : '&#127769;';
  localStorage.setItem('colorMode', light ? 'light' : 'dark');
}

async function logout() {
  await fetch('/auth/logout', {method:'POST'});
  window.location.href = '/';
}

init();
</script>
</body>
</html>"""
