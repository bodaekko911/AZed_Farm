from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy import func, inspect, literal, or_, select
from datetime import datetime, date, timedelta, timezone
from collections import defaultdict
from types import SimpleNamespace
from typing import Optional, Any
import io
import re

from app.core.permissions import require_permission
from app.database import get_async_session
from app.core.navigation import render_app_header
from app.core.product_types import is_stock_tracked_product
from app.models.product import Product
from app.models.invoice import Invoice, InvoiceItem
from app.models.b2b import B2BClient, B2BInvoice, B2BInvoiceItem, B2BRefund, B2BRefundItem
from app.models.inventory import StockMove
from app.models.farm import Farm, FarmDelivery, FarmDeliveryItem
from app.models.animal import AnimalGroup, FeedingLog, MortalityLog, AnimalIntakeLog
from app.models.spoilage import SpoilageRecord
from app.models.refund import RetailRefund, RetailRefundItem
from app.models.production import ProductionBatch, BatchInput, BatchOutput
from app.models.drying import DryingBatch, DryingBatchStage, DryingBatchStageInput, DryingBatchStageOutput
from app.models.accounting import Account, Journal, JournalEntry
from app.models.receipt import ProductReceipt
from app.models.expense import Expense, ExpenseCategory
from app.models.hr import (
    Attendance,
    Employee,
    EmployeeLoan,
    EmployeeLoanRepayment,
    EmployeePayrollDeduction,
    Payroll,
)
from app.models.user import User
from app.services.expense_service import SALARY_CATEGORY_NAME

router = APIRouter(
    prefix="/reports",
    tags=["Reports"],
    dependencies=[Depends(require_permission("page_reports"))],
)


# ── EXCEL HELPER ───────────────────────────────────────
def _excel_dependencies():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter
    except ImportError:
        raise Exception("Run: pip install openpyxl --break-system-packages")


def _coerce_excel_value(value, fmt):
    if value in (None, ""):
        return value
    if fmt == "date" and isinstance(value, str):
        try:
            return datetime.fromisoformat(value[:10]).date()
        except ValueError:
            return value
    if fmt == "datetime" and isinstance(value, str):
        normalized = value.replace("T", " ")
        for parser in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(normalized, parser)
            except ValueError:
                continue
    return value


def _apply_excel_number_format(cell, fmt):
    if fmt == "money":
        cell.number_format = '#,##0.00'
    elif fmt == "qty":
        cell.number_format = '#,##0.00'
    elif fmt == "int":
        cell.number_format = '#,##0'
    elif fmt == "percent":
        cell.number_format = '0.00%'
    elif fmt == "percent_value":
        cell.number_format = '0.00"%"'
    elif fmt == "date":
        cell.number_format = 'yyyy-mm-dd'
    elif fmt == "datetime":
        cell.number_format = 'yyyy-mm-dd hh:mm'


def _autosize_report_sheet(ws, get_column_letter, min_width=10, max_width=42):
    for col_idx in range(1, ws.max_column + 1):
        values = []
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                values.append(str(value))
        max_len = max((len(v) for v in values), default=min_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min(max_len + 3, max_width), min_width)


def add_report_sheet(
    wb,
    *,
    sheet_name,
    report_title,
    headers,
    rows,
    metadata=None,
    column_formats=None,
    wrap_columns=None,
    total_row_indices=None,
    tab_color="1F4E78",
):
    openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter = _excel_dependencies()
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = tab_color

    title_fill = PatternFill("solid", fgColor="1F4E78")
    meta_fill = PatternFill("solid", fgColor="EAF1FB")
    header_fill = PatternFill("solid", fgColor="2F6F4F")
    alt_fill = PatternFill("solid", fgColor="F7FAFC")
    total_fill = PatternFill("solid", fgColor="E3F2E8")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row = 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max(len(headers), 2))
    title_cell = ws.cell(row=current_row, column=1, value=report_title)
    title_cell.fill = title_fill
    title_cell.font = Font(bold=True, color="FFFFFF", size=15)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = border
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    generated_cell = ws.cell(row=current_row, column=1, value=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    generated_cell.font = Font(italic=True, color="5B6B7A", size=10)
    current_row += 1

    for label, value in (metadata or []):
        label_cell = ws.cell(row=current_row, column=1, value=label)
        value_cell = ws.cell(row=current_row, column=2, value=value)
        label_cell.font = Font(bold=True, color="334E68")
        label_cell.fill = meta_fill
        value_cell.fill = meta_fill
        label_cell.border = border
        value_cell.border = border
        current_row += 1

    current_row += 1
    header_row = current_row
    for col_no, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_no, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[header_row].height = 20

    column_formats = column_formats or {}
    wrap_columns = set(wrap_columns or [])
    total_row_indices = set(total_row_indices or [])
    for row_idx, row in enumerate(rows, start=1):
        excel_row = header_row + row_idx
        is_total_row = row_idx in total_row_indices
        for col_idx, value in enumerate(row, 1):
            header = headers[col_idx - 1]
            fmt = column_formats.get(header)
            cell = ws.cell(row=excel_row, column=col_idx, value=_coerce_excel_value(value, fmt))
            cell.border = border
            if is_total_row:
                cell.fill = total_fill
                cell.font = Font(bold=True)
            elif row_idx % 2 == 1:
                cell.fill = alt_fill
            horizontal = "left"
            if fmt in {"money", "qty", "int", "percent", "percent_value"}:
                horizontal = "right"
            elif fmt in {"date", "datetime"}:
                horizontal = "center"
            cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=(header in wrap_columns))
            _apply_excel_number_format(cell, fmt)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
    _autosize_report_sheet(ws, get_column_letter)
    return ws


def build_report_workbook(sheet_specs):
    openpyxl, *_ = _excel_dependencies()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for spec in sheet_specs:
        add_report_sheet(wb, **spec)
    return wb


def workbook_to_buffer(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def to_xlsx(headers, rows, sheet_name="Report", report_title=None, metadata=None, column_formats=None, wrap_columns=None, total_row_indices=None):
    wb = build_report_workbook([
        {
            "sheet_name": sheet_name,
            "report_title": report_title or sheet_name,
            "headers": headers,
            "rows": rows,
            "metadata": metadata or [],
            "column_formats": column_formats or {},
            "wrap_columns": wrap_columns or set(),
            "total_row_indices": total_row_indices or set(),
        }
    ])
    return workbook_to_buffer(wb)


def parse_dates(date_from, date_to):
    """
    Convert a local date range (YYYY-MM-DD strings) to UTC datetime bounds that
    match what the Dashboard uses.  Relies on APP_TIMEZONE so that a transaction
    at 11 pm local time is never shifted into the wrong calendar day.
    """
    from app.core.time_utils import utc_bounds, today_local
    from datetime import date as _date

    try:
        if date_from and date_to:
            d_from_local = _date.fromisoformat(date_from)
            d_to_local   = _date.fromisoformat(date_to)
        else:
            today = today_local()
            d_from_local = today.replace(day=1)
            d_to_local   = today
    except Exception:
        today = today_local()
        d_from_local = today.replace(day=1)
        d_to_local   = today

    return utc_bounds(d_from_local, d_to_local)


def _plain_date_range(date_from, date_to):
    """Plain calendar-date range (no UTC conversion) for reports that filter
    DATE columns directly (e.g. animal intake/death/feed dates). Defaults to the
    current month-to-date. Avoids the off-by-one a UTC datetime bound would cause
    when compared against a DATE column."""
    from app.core.time_utils import today_local
    from datetime import date as _date
    today = today_local()
    try:
        today = today.date()
    except AttributeError:
        pass
    try:
        if date_from and date_to:
            return _date.fromisoformat(date_from), _date.fromisoformat(date_to)
    except Exception:
        pass
    return today.replace(day=1), today


def _resolve_pagination(skip, limit, default_limit=100):
    skip_value = getattr(skip, "default", skip)
    limit_value = getattr(limit, "default", limit)
    try:
        skip_value = max(int(skip_value or 0), 0)
    except (TypeError, ValueError):
        skip_value = 0
    try:
        limit_value = max(int(limit_value or default_limit), 0)
    except (TypeError, ValueError):
        limit_value = default_limit
    return skip_value, limit_value


def _num(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _product_category(product) -> str:
    category = getattr(product, "category", None)
    return category or "—"


async def _schema_has_columns(db: AsyncSession, required: dict[str, set[str]]) -> bool:
    def inspect_schema(sync_session):
        inspector = inspect(sync_session.get_bind())
        table_names = set(inspector.get_table_names())
        for table_name, column_names in required.items():
            if table_name not in table_names:
                return False
            existing_columns = {column["name"] for column in inspector.get_columns(table_name)}
            if column_names - existing_columns:
                return False
        return True

    try:
        if hasattr(db, "run_sync"):
            return await db.run_sync(inspect_schema)
        sync_session = getattr(db, "session", None)
        if sync_session is not None:
            return inspect_schema(sync_session)
    except Exception:
        return False
    return False


def _transaction_sort_value(value) -> datetime:
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        normalized = value.replace("T", " ")
        for candidate in (normalized, normalized[:10]):
            for parser in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
                try:
                    return datetime.strptime(candidate, parser)
                except ValueError:
                    continue
    return datetime.min


PRODUCTION_STOCK_MOVE_REF_TYPES = {
    "production",
    "production_reversal",
    "drying_batch",
    "drying_batch_spoilage",
    "spoilage",
    "spoilage_reversal",
}


def _production_move_label(move: StockMove) -> tuple[str, str, str]:
    ref_type = (move.ref_type or "").lower()
    note = move.note or ""
    if ref_type == "drying_batch":
        return "Processing Stock Move", "Processing", "processed"
    if ref_type == "drying_batch_spoilage":
        return "Processing Spoilage", "Processing", "spoilage"
    if ref_type in {"spoilage", "spoilage_reversal"}:
        return "Production Spoilage", "Production", "spoilage" if ref_type == "spoilage" else "reversed"
    if ref_type == "production_reversal" or "reversal" in note.lower() or "deleted batch" in note.lower():
        return "Production Reversal", "Production", "reversed"
    return "Production Stock Move", "Production", "completed"


def _production_move_reference(move: StockMove) -> str:
    note_match = re.search(r"\b(?:BATCH|PKG|DRY|SPL)-\d{4,}\b", move.note or "")
    if note_match:
        return note_match.group(0)
    if move.ref_id:
        prefix = {
            "production": "BATCH",
            "production_reversal": "BATCH",
            "drying_batch": "DRY",
            "drying_batch_spoilage": "DRY-SPL",
            "spoilage": "SPL",
            "spoilage_reversal": "SPL",
        }.get((move.ref_type or "").lower(), "MOVE")
        return f"{prefix}-{int(move.ref_id):04d}"
    return f"MOVE-{int(move.id):04d}" if move.id else "MOVE"


def _paginate_rows(rows, skip, limit, include_all=False):
    if include_all:
        return rows
    if limit == 0:
        return []
    return rows[skip : skip + limit]


def _channel_totals():
    return {
        "gross_sales": 0.0,
        "cash_collected": 0.0,
        "outstanding": 0.0,
        "count": 0,
    }


async def _load_b2b_client_payment_records(
    db: AsyncSession,
    *,
    d_from: datetime,
    d_to: datetime,
):
    payment_ref_types = ("consignment_client_payment", "consignment_payment", "b2b_payment", "b2b_collection", "b2b")
    payment_result = await db.execute(
        select(Journal)
        .where(
            Journal.ref_type.in_(payment_ref_types),
            Journal.created_at >= d_from,
            Journal.created_at <= d_to,
        )
        .options(selectinload(Journal.entries).selectinload(JournalEntry.account), selectinload(Journal.user))
        .order_by(Journal.created_at.desc(), Journal.id.desc())
    )
    journals = payment_result.scalars().all()
    client_ids = {journal.ref_id for journal in journals if journal.ref_type == "consignment_client_payment" and journal.ref_id}
    invoice_ids = set()
    invoice_numbers = set()
    invoice_pattern = re.compile(r"\b([A-Z]*B2B-\d{5,})\b", re.IGNORECASE)
    for journal in journals:
        if journal.ref_type == "consignment_client_payment":
            continue
        if journal.ref_id:
            invoice_ids.add(journal.ref_id)
        match = invoice_pattern.search(journal.description or "")
        if match:
            invoice_numbers.add(match.group(1).upper())

    invoice_map_by_id = {}
    invoice_map_by_number = {}
    if invoice_ids or invoice_numbers:
        conditions = []
        if invoice_ids:
            conditions.append(B2BInvoice.id.in_(invoice_ids))
        if invoice_numbers:
            conditions.append(func.upper(B2BInvoice.invoice_number).in_(invoice_numbers))
        invoice_result = await db.execute(
            select(B2BInvoice)
            .where(or_(*conditions))
            .options(selectinload(B2BInvoice.client))
        )
        invoices = invoice_result.scalars().all()
        invoice_map_by_id = {invoice.id: invoice for invoice in invoices}
        invoice_map_by_number = {str(invoice.invoice_number or "").upper(): invoice for invoice in invoices}
        client_ids.update(invoice.client_id for invoice in invoices if invoice.client_id)

    client_map = {}
    if client_ids:
        client_result = await db.execute(select(B2BClient).where(B2BClient.id.in_(client_ids)))
        client_map = {client.id: client for client in client_result.scalars().all()}

    payment_records = []
    for journal in journals:
        invoice = None
        amount = 0.0
        has_cash_debit = False
        for entry in journal.entries:
            if entry.account and entry.account.code == "1000" and _num(entry.debit) > 0:
                amount = _num(entry.debit)
                has_cash_debit = True
                break
        if journal.ref_type == "b2b" and not has_cash_debit:
            continue
        if amount <= 0:
            amount = max((_num(entry.debit) for entry in journal.entries), default=0.0)
        client = client_map.get(journal.ref_id)
        reference = f"BCP-{journal.id}"
        if journal.ref_type != "consignment_client_payment":
            invoice = invoice_map_by_id.get(journal.ref_id) if journal.ref_id else None
            if not invoice:
                match = invoice_pattern.search(journal.description or "")
                if match:
                    reference = match.group(1).upper()
                    invoice = invoice_map_by_number.get(reference)
            if invoice:
                reference = invoice.invoice_number or reference
                client = invoice.client or client_map.get(invoice.client_id)
        collection_type = "consignment" if journal.ref_type in {"consignment_client_payment", "consignment_payment"} else ((invoice.invoice_type or "b2b") if invoice else "b2b")
        cash_amount = amount if collection_type == "cash" else 0.0
        full_payment_amount = amount if collection_type == "full_payment" else 0.0
        consignment_amount = amount if collection_type == "consignment" else 0.0
        payment_records.append({
            "journal_id": journal.id,
            "invoice_id": invoice.id if invoice else None,
            "reference": reference,
            "client_id": client.id if client else journal.ref_id,
            "client": client.name if client else "—",
            "datetime": journal.created_at.strftime("%Y-%m-%d %H:%M") if journal.created_at else "—",
            "date": journal.created_at.strftime("%Y-%m-%d") if journal.created_at else "",
            "user_name": journal.user.name if journal.user else "—",
            "amount": round(amount, 2),
            "cash_amount": round(cash_amount, 2),
            "full_payment_amount": round(full_payment_amount, 2),
            "consignment_amount": round(consignment_amount, 2),
            "collection_type": collection_type,
            "notes": journal.description or "",
            "payment_method": "cash",
            "status": "posted",
            "journal_ref_type": journal.ref_type or "—",
        })
    return payment_records


async def _load_b2b_issued_invoice_records(
    db: AsyncSession,
    *,
    d_from: datetime,
    d_to: datetime,
):
    result = await db.execute(
        select(B2BInvoice)
        .where(B2BInvoice.created_at >= d_from, B2BInvoice.created_at <= d_to)
        .options(
            selectinload(B2BInvoice.items).selectinload(B2BInvoiceItem.product),
            selectinload(B2BInvoice.client),
            selectinload(B2BInvoice.user),
        )
        .order_by(B2BInvoice.created_at.desc(), B2BInvoice.id.desc())
    )
    records = []
    for inv in result.scalars().all():
        amount_paid = _num(inv.amount_paid)
        total = _num(inv.total)
        records.append(
            {
                "invoice_number": inv.invoice_number,
                "client": inv.client.name if inv.client else "-",
                "datetime": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "-",
                "user_name": inv.user.name if inv.user else "-",
                "invoice_type": inv.invoice_type,
                "status": inv.status or "-",
                "items": [
                    {
                        "name": item.product.name if item.product else "-",
                        "qty": _num(item.qty),
                        "unit_price": _num(item.unit_price),
                        "total": _num(item.total),
                    }
                    for item in inv.items
                ],
                "total": total,
                "amount_paid": amount_paid,
                "balance_due": max(total - amount_paid, 0.0),
            }
        )
    return records


async def _build_sales_report(
    db: AsyncSession,
    *,
    d_from: datetime,
    d_to: datetime,
    skip: int = 0,
    limit: int = 100,
    include_all: bool = False,
):
    """
    Operational Sales Report using the same revenue definition as Dashboard and P&L.

    Revenue / Net Sales = paid POS invoices
                        + B2B collections posted in the period
                        - retail refunds
                        - B2B refunds

    Notes:
    - B2B invoices are issued on one date and collected on another. Revenue for
      B2B reporting follows the collection journal date, not invoice.created_at.
    - Uncollected B2B invoices are outstanding, not current-period revenue.
    - Top products are calculated from POS + B2B sold items minus refunded items.
    """
    b2b_payment_records = await _load_b2b_client_payment_records(db, d_from=d_from, d_to=d_to)
    b2b_issued_invoice_records = await _load_b2b_issued_invoice_records(db, d_from=d_from, d_to=d_to)

    pos_result = await db.execute(
        select(Invoice)
        .where(
            Invoice.created_at >= d_from,
            Invoice.created_at <= d_to,
            Invoice.status == "paid",
        )
        .options(
            selectinload(Invoice.items).selectinload(InvoiceItem.product),
            selectinload(Invoice.user),
            selectinload(Invoice.customer),
        )
    )
    pos_invoices = pos_result.scalars().all()

    b2b_invoice_ids = {
        int(record["invoice_id"])
        for record in b2b_payment_records
        if record.get("invoice_id") is not None
    }
    b2b_invoices = []
    if b2b_invoice_ids:
        b2b_result = await db.execute(
            select(B2BInvoice)
            .where(B2BInvoice.id.in_(b2b_invoice_ids))
            .options(
                selectinload(B2BInvoice.items).selectinload(B2BInvoiceItem.product),
                selectinload(B2BInvoice.client),
                selectinload(B2BInvoice.user),
            )
        )
        b2b_invoices = b2b_result.scalars().all()

    retail_refund_result = await db.execute(
        select(RetailRefund)
        .where(RetailRefund.created_at >= d_from, RetailRefund.created_at <= d_to)
        .options(
            selectinload(RetailRefund.items).selectinload(RetailRefundItem.product),
            selectinload(RetailRefund.customer),
            selectinload(RetailRefund.user),
        )
    )
    retail_refunds = retail_refund_result.scalars().all()

    b2b_refund_result = await db.execute(
        select(B2BRefund)
        .where(B2BRefund.created_at >= d_from, B2BRefund.created_at <= d_to)
        .options(
            selectinload(B2BRefund.items).selectinload(B2BRefundItem.product),
            selectinload(B2BRefund.client),
            selectinload(B2BRefund.user),
        )
    )
    b2b_refunds = b2b_refund_result.scalars().all()

    channels = {"pos": _channel_totals(), "b2b": _channel_totals()}
    daily = defaultdict(lambda: {"gross_sales": 0.0, "refunds": 0.0, "cash_collected": 0.0})
    product_sales: dict[str, dict[str, Any]] = {}
    sold_item_records: list[dict[str, Any]] = []

    def product_key(product_id: Any, name: str) -> str:
        return f"product:{product_id}" if product_id is not None else f"name:{name}"

    def add_product(
        product_id: Any,
        name: str | None,
        qty: Any,
        revenue: Any,
        multiplier: int = 1,
    ) -> None:
        product_name = name or "-"
        key = product_key(product_id, product_name)
        if key not in product_sales:
            product_sales[key] = {"name": product_name, "qty": 0.0, "revenue": 0.0}
        if product_sales[key]["name"] == "-" and product_name != "-":
            product_sales[key]["name"] = product_name
        product_sales[key]["qty"] += _num(qty) * multiplier
        product_sales[key]["revenue"] += _num(revenue) * multiplier

    def add_sold_item_record(
        *,
        source: str,
        reference: str | None,
        datetime_value: str,
        counterparty: str,
        user_name: str,
        product: Any,
        product_name: str | None,
        qty: Any,
        unit_price: Any,
        line_total: Any,
        payment_method: str = "-",
        status: str = "-",
        line_type: str = "sale",
    ) -> None:
        sold_item_records.append(
            {
                "source": source,
                "reference": reference or "-",
                "datetime": datetime_value,
                "counterparty": counterparty,
                "user_name": user_name,
                "sku": getattr(product, "sku", None) or "-",
                "product": product_name or getattr(product, "name", None) or "-",
                "category": _product_category(product) if product else "-",
                "qty": round(_num(qty), 3),
                "unit_price": round(_num(unit_price), 3),
                "line_total": round(_num(line_total), 2),
                "payment_method": payment_method or "-",
                "status": status or "-",
                "line_type": line_type,
            }
        )

    pos_records = []
    for inv in sorted(
        pos_invoices,
        key=lambda x: x.created_at or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    ):
        total = _num(inv.total)
        collected = total
        outstanding = 0.0
        day_key = inv.created_at.strftime("%Y-%m-%d") if inv.created_at else ""

        channels["pos"]["gross_sales"] += total
        channels["pos"]["cash_collected"] += collected
        channels["pos"]["outstanding"] += outstanding
        channels["pos"]["count"] += 1
        daily[day_key]["gross_sales"] += total
        daily[day_key]["cash_collected"] += collected

        for item in inv.items:
            add_product(item.product_id, item.name or "-", item.qty, item.total, 1)
            add_sold_item_record(
                source="POS",
                reference=inv.invoice_number or f"POS-{inv.id}",
                datetime_value=inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "-",
                counterparty=inv.customer.name if inv.customer else "Walk-in",
                user_name=inv.user.name if inv.user else "-",
                product=item.product,
                product_name=item.name,
                qty=item.qty,
                unit_price=item.unit_price,
                line_total=item.total,
                payment_method=inv.payment_method or "-",
                status=inv.status or "-",
            )

        pos_records.append(
            {
                "invoice_number": inv.invoice_number or f"POS-{inv.id}",
                "datetime": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "-",
                "customer": inv.customer.name if inv.customer else "Walk-in",
                "user_name": inv.user.name if inv.user else "-",
                "payment": inv.payment_method or "-",
                "status": inv.status or "-",
                "items": [
                    {
                        "name": item.name,
                        "qty": _num(item.qty),
                        "unit_price": _num(item.unit_price),
                        "total": _num(item.total),
                    }
                    for item in inv.items
                ],
                "total": total,
                "cash_collected": collected,
                "outstanding": outstanding,
            }
        )

    b2b_invoices_by_id = {inv.id: inv for inv in b2b_invoices}
    collection_totals_by_invoice = defaultdict(float)
    collection_dates_by_invoice = defaultdict(list)
    for payment in b2b_payment_records:
        amount = _num(payment["amount"])
        channels["b2b"]["gross_sales"] += amount
        channels["b2b"]["cash_collected"] += amount
        if payment["date"]:
            daily[payment["date"]]["gross_sales"] += amount
            daily[payment["date"]]["cash_collected"] += amount

        invoice = b2b_invoices_by_id.get(payment.get("invoice_id"))
        if not invoice:
            continue
        collection_totals_by_invoice[invoice.id] += amount
        collection_dates_by_invoice[invoice.id].append(payment["datetime"])
        invoice_total = _num(invoice.total)
        allocation_ratio = (amount / invoice_total) if invoice_total > 0 else 0.0
        for item in invoice.items:
            product_name = item.product.name if item.product else "-"
            allocated_qty = _num(item.qty) * allocation_ratio
            allocated_total = _num(item.total) * allocation_ratio
            add_product(
                item.product_id,
                product_name,
                allocated_qty,
                allocated_total,
                1,
            )
            add_sold_item_record(
                source="B2B Collection",
                reference=payment.get("reference") or invoice.invoice_number,
                datetime_value=payment["datetime"],
                counterparty=invoice.client.name if invoice.client else payment.get("client", "-"),
                user_name=payment.get("user_name") or "-",
                product=item.product,
                product_name=product_name,
                qty=allocated_qty,
                unit_price=item.unit_price,
                line_total=allocated_total,
                payment_method=payment.get("payment_method") or "-",
                status=payment.get("status") or "-",
            )

    b2b_records = []
    for inv in sorted(
        b2b_invoices,
        key=lambda x: max(collection_dates_by_invoice.get(x.id, [""])) or "",
        reverse=True,
    ):
        total = _num(inv.total)
        amount_paid = _num(inv.amount_paid)
        collected_in_period = collection_totals_by_invoice[inv.id]
        outstanding = max(total - amount_paid, 0.0)
        channels["b2b"]["outstanding"] += outstanding
        channels["b2b"]["count"] += 1

        items_data = []
        for item in inv.items:
            product_name = item.product.name if item.product else "-"
            item_qty = _num(item.qty)
            item_total = _num(item.total)
            items_data.append(
                {
                    "name": product_name,
                    "qty": item_qty,
                    "unit_price": _num(item.unit_price),
                    "total": item_total,
                }
            )

        b2b_records.append(
            {
                "invoice_number": inv.invoice_number,
                "client": inv.client.name if inv.client else "-",
                "datetime": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "-",
                "collection_datetime": max(collection_dates_by_invoice.get(inv.id, [""])) or "-",
                "user_name": inv.user.name if inv.user else "-",
                "invoice_type": inv.invoice_type,
                "status": inv.status or "-",
                "items": items_data,
                "total": total,
                "amount_paid": amount_paid,
                "collected_in_period": round(collected_in_period, 2),
                "balance_due": outstanding,
            }
        )

    refund_records = []
    retail_cash_refunds = 0.0
    b2b_cash_refunds = 0.0
    total_refunds = 0.0

    for refund in sorted(
        retail_refunds,
        key=lambda x: x.created_at or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    ):
        refund_total = _num(refund.total)
        day_key = refund.created_at.strftime("%Y-%m-%d") if refund.created_at else ""
        daily[day_key]["refunds"] += refund_total
        total_refunds += refund_total
        if (refund.refund_method or "").lower() == "cash":
            retail_cash_refunds += refund_total

        for item in refund.items:
            product_name = item.product.name if item.product else "-"
            add_product(item.product_id, product_name, item.qty, item.total, -1)
            add_sold_item_record(
                source="Retail Refund",
                reference=refund.refund_number,
                datetime_value=refund.created_at.strftime("%Y-%m-%d %H:%M") if refund.created_at else "-",
                counterparty=refund.customer.name if refund.customer else "-",
                user_name=refund.user.name if refund.user else "-",
                product=item.product,
                product_name=product_name,
                qty=-_num(item.qty),
                unit_price=item.unit_price,
                line_total=-_num(item.total),
                payment_method=refund.refund_method or "-",
                status="refunded",
                line_type="refund",
            )

        refund_records.append(
            {
                "refund_number": refund.refund_number,
                "source": "Retail",
                "counterparty": refund.customer.name if refund.customer else "-",
                "datetime": refund.created_at.strftime("%Y-%m-%d %H:%M") if refund.created_at else "-",
                "processed_by": refund.user.name if refund.user else "-",
                "reason": refund.reason or "-",
                "refund_method": refund.refund_method or "-",
                "total": refund_total,
            }
        )

    for refund in sorted(
        b2b_refunds,
        key=lambda x: x.created_at or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    ):
        refund_total = _num(refund.total)
        day_key = refund.created_at.strftime("%Y-%m-%d") if refund.created_at else ""
        daily[day_key]["refunds"] += refund_total
        total_refunds += refund_total
        b2b_cash_refunds += refund_total

        for item in refund.items:
            product_name = item.product.name if item.product else "-"
            add_product(item.product_id, product_name, item.qty, item.total, -1)
            add_sold_item_record(
                source="B2B Refund",
                reference=refund.refund_number,
                datetime_value=refund.created_at.strftime("%Y-%m-%d %H:%M") if refund.created_at else "-",
                counterparty=refund.client.name if refund.client else "-",
                user_name=refund.user.name if refund.user else "-",
                product=item.product,
                product_name=product_name,
                qty=-_num(item.qty),
                unit_price=item.unit_price,
                line_total=-_num(item.total),
                payment_method="cash",
                status="refunded",
                line_type="refund",
            )

        refund_records.append(
            {
                "refund_number": refund.refund_number,
                "source": "B2B",
                "counterparty": refund.client.name if refund.client else "-",
                "datetime": refund.created_at.strftime("%Y-%m-%d %H:%M") if refund.created_at else "-",
                "processed_by": refund.user.name if refund.user else "-",
                "reason": refund.notes or "-",
                "refund_method": "cash",
                "total": refund_total,
            }
        )

    gross_sales = channels["pos"]["gross_sales"] + channels["b2b"]["gross_sales"]
    cash_collected = (
        channels["pos"]["cash_collected"]
        + channels["b2b"]["cash_collected"]
        - retail_cash_refunds
        - b2b_cash_refunds
    )
    outstanding = channels["pos"]["outstanding"] + channels["b2b"]["outstanding"]
    net_sales = gross_sales - total_refunds

    daily_rows = []
    for day_key, bucket in sorted(daily.items()):
        daily_rows.append(
            {
                "date": day_key,
                "gross_sales": round(bucket["gross_sales"], 2),
                "refunds": round(bucket["refunds"], 2),
                "net_sales": round(bucket["gross_sales"] - bucket["refunds"], 2),
                "cash_collected": round(bucket["cash_collected"], 2),
            }
        )

    top_products = sorted(
        [
            values
            for values in product_sales.values()
            if _num(values["qty"]) > 0 or _num(values["revenue"]) > 0
        ],
        key=lambda values: (-_num(values["revenue"]), values["name"].lower()),
    )[:10]

    return {
        "gross_sales": round(gross_sales, 2),
        "refunds": round(total_refunds, 2),
        "net_sales": round(net_sales, 2),
        "cash_collected": round(cash_collected, 2),
        "outstanding": round(outstanding, 2),
        "channels": {
            "pos": {
                "gross_sales": round(channels["pos"]["gross_sales"], 2),
                "cash_collected": round(channels["pos"]["cash_collected"] - retail_cash_refunds, 2),
                "outstanding": round(channels["pos"]["outstanding"], 2),
                "count": channels["pos"]["count"],
            },
            "b2b": {
                "gross_sales": round(channels["b2b"]["gross_sales"], 2),
                "cash_collected": round(channels["b2b"]["cash_collected"] - b2b_cash_refunds, 2),
                "outstanding": round(channels["b2b"]["outstanding"], 2),
                "count": channels["b2b"]["count"],
            },
        },
        "refund_breakdown": {
            "retail": round(sum(_num(refund.total) for refund in retail_refunds), 2),
            "b2b": round(sum(_num(refund.total) for refund in b2b_refunds), 2),
        },
        "daily": daily_rows,
        "top_products": [
            {
                "name": values["name"],
                "qty": round(values["qty"], 2),
                "revenue": round(values["revenue"], 2),
            }
            for values in top_products
        ],
        "pos_records": _paginate_rows(pos_records, skip, limit, include_all=include_all),
        "b2b_records": _paginate_rows(b2b_records, skip, limit, include_all=include_all),
        "b2b_issued_invoice_records": _paginate_rows(b2b_issued_invoice_records, skip, limit, include_all=include_all),
        "b2b_payment_records": _paginate_rows(b2b_payment_records, skip, limit, include_all=include_all),
        "refund_records": _paginate_rows(refund_records, skip, limit, include_all=include_all),
        "sold_item_records": _paginate_rows(sold_item_records, skip, limit, include_all=include_all),
        "pos_count": len(pos_invoices),
        "b2b_count": len(b2b_invoices),
        "b2b_issued_invoice_count": len(b2b_issued_invoice_records),
        "b2b_payment_count": len(b2b_payment_records),
        "refund_count": len(refund_records),
        "sold_item_count": len(sold_item_records),
        "date_from": d_from.strftime("%Y-%m-%d"),
        "date_to": d_to.strftime("%Y-%m-%d"),
    }

# ── SALES ──────────────────────────────────────────────
@router.get("/api/sales")
async def sales_report(date_from: Optional[str] = None, date_to: Optional[str] = None, skip: int = 0, limit: int = Query(default=100, le=500), db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_sales"))):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    skip, limit = _resolve_pagination(skip, limit)
    return await _build_sales_report(db, d_from=d_from, d_to=d_to, skip=skip, limit=limit)
    result = await db.execute(
        select(Invoice)
        .where(Invoice.created_at >= d_from, Invoice.created_at <= d_to, Invoice.status == "paid")
        .options(selectinload(Invoice.items), selectinload(Invoice.user), selectinload(Invoice.customer))
    )
    pos_invoices = result.scalars().all()
    result = await db.execute(
        select(B2BInvoice)
        .where(B2BInvoice.created_at >= d_from, B2BInvoice.created_at <= d_to)
        .options(selectinload(B2BInvoice.items).selectinload(B2BInvoiceItem.product),
                 selectinload(B2BInvoice.client), selectinload(B2BInvoice.user))
    )
    b2b_invoices = result.scalars().all()
    result = await db.execute(
        select(RetailRefund)
        .where(RetailRefund.created_at >= d_from, RetailRefund.created_at <= d_to)
        .options(selectinload(RetailRefund.customer), selectinload(RetailRefund.user))
    )
    refunds = result.scalars().all()

    pos_total    = sum(float(i.total) for i in pos_invoices)
    b2b_total    = sum(float(i.amount_paid) for i in b2b_invoices)
    refund_total = sum(float(r.total) for r in refunds)
    pos_total    = max(0, pos_total - refund_total)

    daily = {}
    for i in pos_invoices:
        d = i.created_at.strftime("%Y-%m-%d")
        daily.setdefault(d, {"pos": 0, "b2b": 0, "refunds": 0})
        daily[d]["pos"] += float(i.total)
    for i in b2b_invoices:
        d = i.created_at.strftime("%Y-%m-%d")
        daily.setdefault(d, {"pos": 0, "b2b": 0, "refunds": 0})
        daily[d]["b2b"] += float(i.amount_paid)
    for r in refunds:
        d = r.created_at.strftime("%Y-%m-%d")
        daily.setdefault(d, {"pos": 0, "b2b": 0, "refunds": 0})
        daily[d]["refunds"] += float(r.total)
    daily_list = [{"date": k, "pos": round(max(0, v["pos"] - v["refunds"]), 2), "b2b": round(v["b2b"], 2), "refunds": round(v["refunds"], 2), "total": round(max(0, v["pos"] - v["refunds"]) + v["b2b"], 2)} for k, v in sorted(daily.items())]

    product_sales = {}
    for inv in pos_invoices:
        for item in inv.items:
            product_sales.setdefault(item.name, {"qty": 0, "revenue": 0})
            product_sales[item.name]["qty"]     += float(item.qty)
            product_sales[item.name]["revenue"] += float(item.total)
    top = sorted(product_sales.items(), key=lambda x: x[1]["revenue"], reverse=True)[:10]

    # Detailed POS records
    pos_records = []
    for inv in sorted(pos_invoices, key=lambda x: x.created_at, reverse=True):
        items = inv.items
        pos_records.append({
            "invoice_number": inv.invoice_number or f"POS-{inv.id}",
            "datetime": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "—",
            "user_name": inv.user.name if inv.user else "—",
            "payment": inv.payment_method or "—",
            "items": [{"name": it.name, "qty": float(it.qty), "unit_price": float(it.unit_price), "total": float(it.total)} for it in items],
            "total": float(inv.total),
        })

    # Detailed B2B records
    b2b_records = []
    for inv in sorted(b2b_invoices, key=lambda x: x.created_at, reverse=True):
        items_data = []
        for it in inv.items:
            items_data.append({"name": it.product.name if it.product else "—", "qty": float(it.qty), "unit_price": float(it.unit_price), "total": float(it.total)})
        b2b_records.append({
            "invoice_number": inv.invoice_number,
            "client": inv.client.name if inv.client else "—",
            "datetime": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "—",
            "user_name": inv.user.name if inv.user else "—",
            "invoice_type": inv.invoice_type,
            "status": inv.status,
            "items": items_data,
            "total": float(inv.total),
            "amount_paid": float(inv.amount_paid),
            "balance_due": float(inv.total) - float(inv.amount_paid),
        })

    # Detailed refund records
    refund_records = []
    for r in sorted(refunds, key=lambda x: x.created_at, reverse=True):
        refund_records.append({
            "refund_number":  r.refund_number,
            "customer":       r.customer.name if r.customer else "—",
            "datetime":       r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "—",
            "processed_by":   r.user.name if r.user else "—",
            "reason":         r.reason or "—",
            "refund_method":  r.refund_method,
            "total":          float(r.total),
        })

    pos_records = pos_records[skip : skip + limit]
    b2b_records = b2b_records[skip : skip + limit]
    refund_records = refund_records[skip : skip + limit]

    return {"pos_total": round(pos_total, 2), "b2b_total": round(b2b_total, 2),
            "refund_total": round(refund_total, 2),
            "grand_total": round(pos_total + b2b_total, 2),
            "pos_count": len(pos_invoices), "b2b_count": len(b2b_invoices), "refund_count": len(refunds),
            "daily": daily_list, "top_products": [{"name": k, "qty": round(v["qty"], 2), "revenue": round(v["revenue"], 2)} for k, v in top],
            "pos_records": pos_records, "b2b_records": b2b_records, "refund_records": refund_records,
            "date_from": d_from.strftime("%Y-%m-%d"), "date_to": d_to.strftime("%Y-%m-%d")}

@router.get("/export/sales", dependencies=[Depends(require_permission("action_export_excel")), Depends(require_permission("tab_reports_sales"))])
async def export_sales(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    data = await _build_sales_report(db, d_from=d_from, d_to=d_to, include_all=True)
    wb = build_report_workbook([
        {
            "sheet_name": "Summary",
            "report_title": "Sales Report Summary",
            "headers": ["Metric", "Value"],
            "rows": [
                ["Gross Sales", data["gross_sales"]],
                ["Refunds", data["refunds"]],
                ["Net Sales", data["net_sales"]],
                ["Cash Collected", data["cash_collected"]],
                ["Outstanding", data["outstanding"]],
                ["POS Gross Sales", data["channels"]["pos"]["gross_sales"]],
                ["POS Cash Collected", data["channels"]["pos"]["cash_collected"]],
                ["POS Outstanding", data["channels"]["pos"]["outstanding"]],
                ["B2B Gross Sales", data["channels"]["b2b"]["gross_sales"]],
                ["B2B Cash Collected", data["channels"]["b2b"]["cash_collected"]],
                ["B2B Outstanding", data["channels"]["b2b"]["outstanding"]],
                ["B2B Client Payments", data["b2b_payment_count"]],
                ["Retail Refunds", data["refund_breakdown"]["retail"]],
                ["B2B Refunds", data["refund_breakdown"]["b2b"]],
            ],
            "metadata": [
                ("Date Range", f"{data['date_from']} to {data['date_to']}"),
                ("POS Invoices", data["pos_count"]),
                ("B2B Issued Invoices", data["b2b_issued_invoice_count"]),
                ("B2B Collected Invoices", data["b2b_count"]),
                ("B2B Payment Records", data["b2b_payment_count"]),
                ("Refund Records", data["refund_count"]),
            ],
            "column_formats": {"Value": "money"},
            "total_row_indices": {1, 2, 3, 4, 5},
            "tab_color": "1F4E78",
        },
        {
            "sheet_name": "Daily",
            "report_title": "Sales Daily Breakdown",
            "headers": ["Date", "Gross Sales", "Refunds", "Net Sales", "Cash Collected"],
            "rows": [[row["date"], row["gross_sales"], row["refunds"], row["net_sales"], row["cash_collected"]] for row in data["daily"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}")],
            "column_formats": {"Date": "date", "Gross Sales": "money", "Refunds": "money", "Net Sales": "money", "Cash Collected": "money"},
            "tab_color": "2F6F4F",
        },
        {
            "sheet_name": "POS Invoices",
            "report_title": "POS Invoice Detail",
            "headers": ["Invoice #", "Date / Time", "Customer", "User", "Payment", "Status", "Invoice Total", "Cash Collected", "Outstanding"],
            "rows": [[row["invoice_number"], row["datetime"], row["customer"], row["user_name"], row["payment"], row["status"], row["total"], row["cash_collected"], row["outstanding"]] for row in data["pos_records"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Records", len(data["pos_records"]))],
            "column_formats": {"Date / Time": "datetime", "Invoice Total": "money", "Cash Collected": "money", "Outstanding": "money"},
            "tab_color": "4F81BD",
        },
        {
            "sheet_name": "B2B Invoices",
            "report_title": "B2B Issued Invoice Detail",
            "headers": ["Invoice #", "Client", "Issued At", "User", "Type", "Status", "Total Invoiced", "Lifetime Paid", "Outstanding"],
            "rows": [[row["invoice_number"], row["client"], row["datetime"], row["user_name"], row["invoice_type"], row["status"], row["total"], row["amount_paid"], row["balance_due"]] for row in data["b2b_issued_invoice_records"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Records", len(data["b2b_issued_invoice_records"]))],
            "column_formats": {"Issued At": "datetime", "Total Invoiced": "money", "Lifetime Paid": "money", "Outstanding": "money"},
            "tab_color": "C55A11",
        },
        {
            "sheet_name": "B2B Collections",
            "report_title": "B2B Client Payment Detail",
            "headers": ["Reference", "Client", "Date / Time", "User", "Cash", "Full Payment", "Consignment", "Total Amount", "Notes"],
            "rows": [[row["reference"], row["client"], row["datetime"], row["user_name"], row["cash_amount"], row["full_payment_amount"], row["consignment_amount"], row["amount"], row["notes"]] for row in data["b2b_payment_records"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Records", len(data["b2b_payment_records"]))],
            "column_formats": {"Date / Time": "datetime", "Cash": "money", "Full Payment": "money", "Consignment": "money", "Total Amount": "money"},
            "wrap_columns": {"Notes"},
            "tab_color": "2F6F4F",
        },
        {
            "sheet_name": "Refunds",
            "report_title": "Refund Detail",
            "headers": ["Refund #", "Source", "Counterparty", "Date / Time", "Processed By", "Method", "Reason", "Amount"],
            "rows": [[row["refund_number"], row["source"], row["counterparty"], row["datetime"], row["processed_by"], row["refund_method"], row["reason"], row["total"]] for row in data["refund_records"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Records", len(data["refund_records"]))],
            "column_formats": {"Date / Time": "datetime", "Amount": "money"},
            "wrap_columns": {"Reason"},
            "tab_color": "C00000",
        },
        {
            "sheet_name": "Items Sold",
            "report_title": "Items Sold Detail",
            "headers": ["Date / Time", "Source", "Reference", "Counterparty", "User", "SKU", "Product", "Category", "Qty", "Unit Price", "Line Total", "Payment", "Status"],
            "rows": [[row["datetime"], row["source"], row["reference"], row["counterparty"], row["user_name"], row["sku"], row["product"], row["category"], row["qty"], row["unit_price"], row["line_total"], row["payment_method"], row["status"]] for row in data["sold_item_records"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Records", len(data["sold_item_records"]))],
            "column_formats": {"Date / Time": "datetime", "Qty": "qty", "Unit Price": "money", "Line Total": "money"},
            "tab_color": "70AD47",
        },
        {
            "sheet_name": "Top Products",
            "report_title": "Top Products",
            "headers": ["Product", "Qty Sold", "Gross Sales"],
            "rows": [[row["name"], row["qty"], row["revenue"]] for row in data["top_products"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}")],
            "column_formats": {"Qty Sold": "qty", "Gross Sales": "money"},
            "tab_color": "70AD47",
        },
    ])
    buf = workbook_to_buffer(wb)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sales_report_{date.today()}.xlsx"},
    )


@router.get("/api/b2b-statement")
async def b2b_statement(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    skip: int = 0,
    limit: int = Query(default=100, le=500),
    db: AsyncSession = Depends(get_async_session),
    _=Depends(require_permission("tab_reports_b2b")),
):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    skip, limit = _resolve_pagination(skip, limit)
    return await _build_b2b_statement(db, d_from=d_from, d_to=d_to, skip=skip, limit=limit)


async def _build_b2b_statement(db, *, d_from, d_to, skip=0, limit=100, include_all=False):
    res = await db.execute(select(B2BClient).where(B2BClient.is_active == True).order_by(B2BClient.name))
    clients = res.scalars().all()
    result = []
    for client in clients:
        agg_res = await db.execute(
            select(
                func.count(B2BInvoice.id),
                func.sum(B2BInvoice.total),
                func.sum(B2BInvoice.amount_paid),
            ).where(
                B2BInvoice.client_id == client.id,
                B2BInvoice.created_at >= d_from,
                B2BInvoice.created_at <= d_to,
            )
        )
        invoice_count, total_invoiced, total_paid = agg_res.one()
        if not invoice_count:
            continue
        total_invoiced = _num(total_invoiced)
        total_paid = _num(total_paid)
        result.append(
            {
                "id": client.id,
                "name": client.name,
                "phone": client.phone or "-",
                "payment_terms": client.payment_terms or "-",
                "total_invoiced": round(total_invoiced, 2),
                "total_paid": round(total_paid, 2),
                "outstanding": round(total_invoiced - total_paid, 2),
                "invoice_count": int(invoice_count or 0),
            }
        )
    if include_all:
        return result
    return result[skip : skip + limit]


@router.get("/export/b2b-statement", dependencies=[Depends(require_permission("action_export_excel")), Depends(require_permission("tab_reports_b2b"))])
async def export_b2b(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    data = await _build_b2b_statement(db, d_from=d_from, d_to=d_to, include_all=True)
    rows = [[d["name"],d["phone"],d["payment_terms"],d["total_invoiced"],d["total_paid"],d["outstanding"],d["invoice_count"]] for d in data]
    buf = to_xlsx(
        ["Client","Phone","Payment Terms","Total Invoiced","Total Paid","Outstanding","Invoices"],
        rows,
        "B2B Statement",
        report_title="B2B Statement",
        metadata=[("Date Range", f"{d_from.strftime('%Y-%m-%d')} to {d_to.strftime('%Y-%m-%d')}"), ("Rows Exported", len(rows))],
        column_formats={"Total Invoiced": "money", "Total Paid": "money", "Outstanding": "money", "Invoices": "int"},
    )
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=b2b_statement_{date.today()}.xlsx"})


# ── INVENTORY ──────────────────────────────────────────
async def _build_inventory_report(
    db: AsyncSession,
    *,
    mode: str,
    d_from: Optional[datetime] = None,
    d_to: Optional[datetime] = None,
    skip: int = 0,
    limit: int = 100,
    include_all: bool = False,
):
    if mode == "movement":
        move_res = await db.execute(
            select(StockMove)
            .where(StockMove.created_at >= d_from, StockMove.created_at <= d_to)
            .options(selectinload(StockMove.product))
            .order_by(StockMove.created_at.desc(), StockMove.id.desc())
        )
        moves = move_res.scalars().all()
        grouped = {}
        detail_rows = []
        for move in moves:
            product = move.product
            if product is None:
                continue
            if product.id not in grouped:
                grouped[product.id] = {
                    "sku": product.sku,
                    "name": product.name,
                    "category": product.category or "—",
                    "unit": product.unit,
                    "stock_in": 0.0,
                    "stock_out": 0.0,
                    "receipts": 0.0,
                    "sales_usage": 0.0,
                    "spoilage": 0.0,
                    "transfers_in": 0.0,
                    "transfers_out": 0.0,
                    "adjustments_net": 0.0,
                    "net_movement": 0.0,
                }
            row = grouped[product.id]
            qty = abs(_num(move.qty))
            is_in = (move.type or "").lower() == "in"
            signed_qty = qty if is_in else -qty
            row["net_movement"] += signed_qty
            if is_in:
                row["stock_in"] += qty
            else:
                row["stock_out"] += qty
            ref_type = (move.ref_type or "").lower()
            if ref_type == "receipt":
                row["receipts"] += qty
            elif ref_type in {"invoice", "b2b", "consignment"}:
                row["sales_usage"] += qty
            elif ref_type == "spoilage":
                row["spoilage"] += qty
            elif ref_type == "transfer":
                if is_in:
                    row["transfers_in"] += qty
                else:
                    row["transfers_out"] += qty
            else:
                row["adjustments_net"] += signed_qty
            detail_rows.append({
                "date": move.created_at.strftime("%Y-%m-%d %H:%M") if move.created_at else "—",
                "sku": product.sku,
                "product": product.name,
                "transaction_type": ref_type or (move.type or "move"),
                "direction": "in" if is_in else "out",
                "qty": qty,
                "unit": product.unit,
                "reference": f"{move.ref_type or 'move'}:{move.ref_id or ''}".strip(":"),
                "note": move.note or "",
            })
        rows = sorted(grouped.values(), key=lambda x: x["name"])
        for row in rows:
            for key in ("stock_in", "stock_out", "receipts", "sales_usage", "spoilage", "transfers_in", "transfers_out", "adjustments_net", "net_movement"):
                row[key] = round(row[key], 2)
        return {
            "mode": "movement",
            "date_from": d_from.strftime("%Y-%m-%d"),
            "date_to": d_to.strftime("%Y-%m-%d"),
            "products": _paginate_rows(rows, skip, limit, include_all=include_all),
            "detail_rows": detail_rows if include_all else _paginate_rows(detail_rows, skip, limit, include_all=False),
            "total_products": len(rows),
            "summary": {
                "stock_in": round(sum(r["stock_in"] for r in rows), 2),
                "stock_out": round(sum(r["stock_out"] for r in rows), 2),
                "receipts": round(sum(r["receipts"] for r in rows), 2),
                "spoilage": round(sum(r["spoilage"] for r in rows), 2),
            },
        }

    prod_res = await db.execute(select(Product).where(Product.is_active == True).order_by(Product.name))
    products = prod_res.scalars().all()
    rows = []
    dead_stock_cutoff = datetime.now(timezone.utc) - timedelta(days=90)
    dead_stock_count = 0
    for product in products:
        stock_tracked = is_stock_tracked_product(product)
        threshold = _num(product.reorder_level if product.reorder_level is not None else product.min_stock if product.min_stock is not None else 5)
        last_move_res = await db.execute(select(func.max(StockMove.created_at)).where(StockMove.product_id == product.id))
        last_move_at = last_move_res.scalar()
        is_dead_stock = stock_tracked and _num(product.stock) > 0 and (last_move_at is None or last_move_at < dead_stock_cutoff)
        low_stock = stock_tracked and _num(product.stock) <= threshold
        if is_dead_stock:
            dead_stock_count += 1
        rows.append({
            "sku": product.sku,
            "name": product.name,
            "category": product.category or "—",
            "stock": _num(product.stock),
            "unit": product.unit,
            "price": _num(product.price),
            "value": round(_num(product.stock) * _num(product.cost), 2) if stock_tracked else 0.0,
            "threshold": round(threshold, 2),
            "reorder_qty": round(_num(product.reorder_qty), 2),
            "low_stock": low_stock,
            "dead_stock": is_dead_stock,
            "last_move_at": last_move_at.strftime("%Y-%m-%d") if last_move_at else "—",
        })
    return {
        "mode": "snapshot",
        "date_from": None,
        "date_to": None,
        "products": _paginate_rows(rows, skip, limit, include_all=include_all),
        "total_value": round(sum(r["value"] for r in rows), 2),
        "low_count": sum(1 for r in rows if r["low_stock"]),
        "dead_stock_count": dead_stock_count,
        "total_products": len(rows),
    }


@router.get("/api/inventory")
async def inventory_report(mode: str = "snapshot", date_from: Optional[str] = None, date_to: Optional[str] = None, skip: int = 0, limit: int = Query(default=100, le=500), db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_inventory"))):
    skip, limit = _resolve_pagination(skip, limit)
    if mode == "movement":
        d_from, d_to = parse_dates(date_from, date_to)
        if (d_to - d_from).days > 366:
            raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
        return await _build_inventory_report(db, mode="movement", d_from=d_from, d_to=d_to, skip=skip, limit=limit)
    return await _build_inventory_report(db, mode="snapshot", skip=skip, limit=limit)
    prod_res = await db.execute(select(Product).where(Product.is_active == True).order_by(Product.name))
    products = prod_res.scalars().all()
    rows = []
    for p in products:
        in_res  = await db.execute(select(func.sum(StockMove.qty)).where(StockMove.product_id==p.id, StockMove.type=="in"))
        out_res = await db.execute(select(func.sum(StockMove.qty)).where(StockMove.product_id==p.id, StockMove.type=="out"))
        total_in  = float(in_res.scalar() or 0)
        total_out = abs(float(out_res.scalar() or 0))
        rows.append({"sku":p.sku,"name":p.name,"stock":float(p.stock),"unit":p.unit,"price":float(p.price),
            "value":round(float(p.stock)*float(p.price),2),"total_in":round(total_in,2),"total_out":round(total_out,2),"low_stock":float(p.stock)<=5})
    total_value   = round(sum(r["value"] for r in rows), 2)
    low_count     = sum(1 for r in rows if r["low_stock"])
    total_products = len(rows)
    rows = rows[skip : skip + limit]
    return {"products":rows,"total_value":total_value,"low_count":low_count,"total_products":total_products}

@router.get("/export/inventory", dependencies=[Depends(require_permission("action_export_excel")), Depends(require_permission("tab_reports_inventory"))])
async def export_inventory(mode: str = "snapshot", date_from: Optional[str] = None, date_to: Optional[str] = None, db: AsyncSession = Depends(get_async_session)):
    if mode == "movement":
        d_from, d_to = parse_dates(date_from, date_to)
        data = await _build_inventory_report(db, mode="movement", d_from=d_from, d_to=d_to, include_all=True)
        rows = [[p["sku"], p["name"], p["category"], p["unit"], p["stock_in"], p["stock_out"], p["receipts"], p["sales_usage"], p["spoilage"], p["transfers_in"], p["transfers_out"], p["adjustments_net"], p["net_movement"]] for p in data["products"]]
        buf = to_xlsx(
            ["SKU","Product","Category","Unit","Stock In","Stock Out","Receipts","Sales/Usage","Spoilage","Transfers In","Transfers Out","Adjustments Net","Net Movement"],
            rows,
            "Inventory Movement",
            report_title="Inventory Movement Report",
            metadata=[
                ("Date Range", f"{data['date_from']} to {data['date_to']}"),
                ("Products", data["total_products"]),
            ],
            column_formats={"Stock In": "qty", "Stock Out": "qty", "Receipts": "qty", "Sales/Usage": "qty", "Spoilage": "qty", "Transfers In": "qty", "Transfers Out": "qty", "Adjustments Net": "qty", "Net Movement": "qty"},
        )
    else:
        data = await _build_inventory_report(db, mode="snapshot", include_all=True)
        rows = [[p["sku"], p["name"], p["category"], p["stock"], p["unit"], p["price"], p["value"], p["threshold"], p["reorder_qty"], p["last_move_at"], "YES" if p["low_stock"] else "", "YES" if p["dead_stock"] else ""] for p in data["products"]]
        buf = to_xlsx(
            ["SKU","Product","Category","Stock","Unit","Price (EGP)","Stock Value","Threshold","Reorder Qty","Last Move","Low Stock","Dead Stock"],
            rows,
            "Inventory Snapshot",
            report_title="Inventory Snapshot Report",
            metadata=[
                ("Products", data["total_products"]),
                ("Low Stock Items", data["low_count"]),
                ("Dead Stock Items", data["dead_stock_count"]),
                ("Total Stock Value", f"{data['total_value']:.2f}"),
            ],
            column_formats={"Stock": "qty", "Price (EGP)": "money", "Stock Value": "money", "Threshold": "qty", "Reorder Qty": "qty", "Last Move": "date"},
        )
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=inventory_{date.today()}.xlsx"})


# ── FARM INTAKE ────────────────────────────────────────
async def _build_farm_intake_report(
    db: AsyncSession,
    *,
    d_from: datetime,
    d_to: datetime,
    skip: int = 0,
    limit: int = 100,
    include_all: bool = False,
):
    delivery_res = await db.execute(
        select(FarmDelivery)
        .where(FarmDelivery.delivery_date >= d_from.date(), FarmDelivery.delivery_date <= d_to.date())
        .options(
            selectinload(FarmDelivery.farm),
            selectinload(FarmDelivery.items).selectinload(FarmDeliveryItem.product),
            selectinload(FarmDelivery.user),
        )
        .order_by(FarmDelivery.delivery_date.desc(), FarmDelivery.id.desc())
    )
    deliveries = delivery_res.scalars().all()
    farm_summary = {}
    detail_rows = []

    def get_summary(farm_id: int, farm_name: str) -> dict:
        return farm_summary.setdefault(
            farm_id,
            {
                "farm_id": farm_id,
                "farm": farm_name,
                "delivery_count": 0,
                "line_count": 0,
                "total_qty": 0.0,
                "products": defaultdict(float),
                "salary_cost": 0.0,
                "labor_cost": 0.0,
                "expense_count": 0,
            },
        )

    for delivery in deliveries:
        farm_name = delivery.farm.name if delivery.farm and delivery.farm.name else f"Farm {delivery.farm_id}"
        summary = get_summary(delivery.farm_id, farm_name)
        summary["delivery_count"] += 1
        for item in delivery.items:
            product = item.product
            sku = product.sku if product else "—"
            name = product.name if product else "—"
            qty = _num(item.qty)
            unit = item.unit or (product.unit if product else "")
            summary["line_count"] += 1
            summary["total_qty"] += qty
            summary["products"][f"{sku}|{name}|{unit}"] += qty
            detail_rows.append({
                "farm": farm_name,
                "date": str(delivery.delivery_date),
                "delivery_number": delivery.delivery_number,
                "sku": sku,
                "product": name,
                "qty": round(qty, 2),
                "unit": unit,
                "received_by": delivery.received_by or "—",
                "user_name": delivery.user.name if delivery.user else "—",
                "notes": item.notes or delivery.notes or "",
            })
    salary_res = await db.execute(
        select(Expense)
        .join(ExpenseCategory, Expense.category_id == ExpenseCategory.id)
        .options(selectinload(Expense.farm))
        .where(
            ExpenseCategory.name == SALARY_CATEGORY_NAME,
            Expense.farm_id.is_not(None),
            Expense.expense_date >= d_from.date(),
            Expense.expense_date <= d_to.date(),
        )
    )
    for expense in salary_res.scalars().all():
        farm_name = expense.farm.name if expense.farm and expense.farm.name else f"Farm {expense.farm_id}"
        summary = get_summary(expense.farm_id, farm_name)
        amount = _num(expense.amount)
        # Salary & Wages is exposed as labor cost here; it is not added to harvest quantity totals.
        summary["salary_cost"] += amount
        summary["labor_cost"] += amount
        summary["expense_count"] += 1

    summary_rows = []
    for _farm_id, summary in sorted(farm_summary.items(), key=lambda item: item[1]["farm"]):
        top_product = ""
        if summary["products"]:
            top_key, top_qty = max(summary["products"].items(), key=lambda x: x[1])
            _, top_name, top_unit = top_key.split("|")
            top_product = f"{top_name} ({round(top_qty, 2)} {top_unit})"
        summary_rows.append({
            "farm": summary["farm"],
            "delivery_count": summary["delivery_count"],
            "line_count": summary["line_count"],
            "total_qty": round(summary["total_qty"], 2),
            "salary_cost": round(summary["salary_cost"], 2),
            "labor_cost": round(summary["labor_cost"], 2),
            "expense_count": summary["expense_count"],
            "top_product": top_product or "—",
        })
    return {
        "date_from": d_from.strftime("%Y-%m-%d"),
        "date_to": d_to.strftime("%Y-%m-%d"),
        "summary": summary_rows,
        "detail": _paginate_rows(detail_rows, skip, limit, include_all=include_all),
        "totals": {
            "delivery_count": len(deliveries),
            "line_count": len(detail_rows),
            "total_qty": round(sum(r["qty"] for r in detail_rows), 2),
            "salary_cost": round(sum(r["salary_cost"] for r in summary_rows), 2),
            "labor_cost": round(sum(r["labor_cost"] for r in summary_rows), 2),
            "salary_expense_count": sum(r["expense_count"] for r in summary_rows),
            "farm_count": len(summary_rows),
        },
    }


@router.get("/api/farm-intake")
async def farm_intake_report(date_from: Optional[str] = None, date_to: Optional[str] = None, skip: int = 0, limit: int = Query(default=100, le=500), db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_farm"))):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    skip, limit = _resolve_pagination(skip, limit)
    return await _build_farm_intake_report(db, d_from=d_from, d_to=d_to, skip=skip, limit=limit)
    farm_res = await db.execute(select(Farm).where(Farm.is_active == 1))
    farms = farm_res.scalars().all()
    # Auto-fix unnamed farms
    default_names = ["Organic Farm", "Regenerative Farm"]
    for i, farm in enumerate(farms):
        if not farm.name or str(farm.name).strip().lower() in ("none", ""):
            farm.name = default_names[i] if i < len(default_names) else f"Farm {farm.id}"
    try: await db.commit()
    except Exception: await db.rollback()
    result = []
    delivery_rows = []
    for farm in farms:
        del_res = await db.execute(
            select(FarmDelivery)
            .where(FarmDelivery.farm_id==farm.id, FarmDelivery.delivery_date>=d_from.date(), FarmDelivery.delivery_date<=d_to.date())
            .options(selectinload(FarmDelivery.items).selectinload(FarmDeliveryItem.product),
                     selectinload(FarmDelivery.user))
        )
        deliveries = del_res.scalars().all()
        product_totals = {}
        for d in deliveries:
            delivery_rows.append({
                "delivery_number": d.delivery_number,
                "farm": farm.name or f"Farm {farm.id}",
                "delivery_date": str(d.delivery_date),
                "received_by": d.received_by or "—",
                "user_name": d.user.name if d.user else "—",
                "total_items": len(d.items),
                "total_qty": round(sum(float(item.qty) for item in d.items), 2),
                "notes": d.notes or "",
            })
            for item in d.items:
                name = item.product.name if item.product else "—"
                unit = item.unit or ""
                key  = f"{name}|{unit}"
                product_totals[key] = product_totals.get(key, 0) + float(item.qty)
        products = [{"name":k.split("|")[0],"unit":k.split("|")[1],"total_qty":round(v,2)} for k,v in sorted(product_totals.items(), key=lambda x: x[1], reverse=True)]
        result.append({"name": farm.name or f"Farm {farm.id}", "delivery_count":len(deliveries), "products":products, "total_qty":round(sum(p["total_qty"] for p in products),2)})
    delivery_rows.sort(key=lambda row: (row["delivery_date"], row["delivery_number"]), reverse=True)
    delivery_rows = delivery_rows[skip : skip + limit]
    return {"farms": result, "deliveries": delivery_rows}

@router.get("/export/farm-intake", dependencies=[Depends(require_permission("action_export_excel")), Depends(require_permission("tab_reports_farm"))])
async def export_farm(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    data = await _build_farm_intake_report(db, d_from=d_from, d_to=d_to, include_all=True)
    wb = build_report_workbook([
        {
            "sheet_name": "Farm Intake Summary",
            "report_title": "Farm Intake Summary",
            "headers": ["Farm", "Deliveries", "Line Items", "Total Qty", "Salary & Wages", "Labor Cost", "Top Product"],
            "rows": [[row["farm"], row["delivery_count"], row["line_count"], row["total_qty"], row["salary_cost"], row["labor_cost"], row["top_product"]] for row in data["summary"]],
            "metadata": [
                ("Date Range", f"{data['date_from']} to {data['date_to']}"),
                ("Farms", data["totals"]["farm_count"]),
                ("Deliveries", data["totals"]["delivery_count"]),
                ("Line Items", data["totals"]["line_count"]),
                ("Total Qty", data["totals"]["total_qty"]),
                ("Salary & Wages", data["totals"]["salary_cost"]),
            ],
            "column_formats": {"Deliveries": "int", "Line Items": "int", "Total Qty": "qty", "Salary & Wages": "money", "Labor Cost": "money"},
            "tab_color": "70AD47",
        },
        {
            "sheet_name": "Farm Intake Detail",
            "report_title": "Farm Intake Detail",
            "headers": ["Farm", "Date", "Delivery #", "SKU", "Product", "Qty", "Unit", "Received By", "Performed By", "Notes"],
            "rows": [[row["farm"], row["date"], row["delivery_number"], row["sku"], row["product"], row["qty"], row["unit"], row["received_by"], row["user_name"], row["notes"]] for row in data["detail"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Rows Exported", len(data["detail"]))],
            "column_formats": {"Date": "date", "Qty": "qty"},
            "wrap_columns": {"Notes"},
            "tab_color": "548235",
        },
    ])
    buf = workbook_to_buffer(wb)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=farm_intake_{date.today()}.xlsx"})


# ── SPOILAGE ───────────────────────────────────────────
async def _build_spoilage_report(db, *, d_from, d_to, skip=0, limit=100, include_all=False):
    sp_res = await db.execute(
        select(SpoilageRecord)
        .where(SpoilageRecord.spoilage_date>=d_from.date(), SpoilageRecord.spoilage_date<=d_to.date())
        .order_by(SpoilageRecord.spoilage_date.desc())
        .options(selectinload(SpoilageRecord.product), selectinload(SpoilageRecord.farm), selectinload(SpoilageRecord.user))
    )
    records = sp_res.scalars().all()
    by_product, by_reason, rows = {}, {}, []
    for r in records:
        name = r.product.name if r.product else "—"; unit = r.product.unit if r.product else ""; reason = r.reason or "—"
        by_product[name]  = by_product.get(name, 0)  + float(r.qty)
        by_reason[reason] = by_reason.get(reason, 0) + float(r.qty)
        rows.append({"ref":r.ref_number,"product":name,"qty":float(r.qty),"unit":unit,"reason":reason,"farm":r.farm.name if r.farm else "—","date":str(r.spoilage_date),"user_name":r.user.name if r.user else "—","notes":r.notes or ""})
    total_qty   = round(sum(float(r.qty) for r in records), 2)
    total_count = len(records)
    if not include_all:
        rows = rows[skip : skip + limit]
    return {"records":rows,"total_qty":total_qty,"total_count":total_count,
            "by_product":[{"name":k,"qty":round(v,2)} for k,v in sorted(by_product.items(),key=lambda x:x[1],reverse=True)[:8]],
            "by_reason": [{"reason":k,"qty":round(v,2)} for k,v in sorted(by_reason.items(), key=lambda x:x[1],reverse=True)]}


@router.get("/api/spoilage")
async def spoilage_report(date_from: Optional[str] = None, date_to: Optional[str] = None, skip: int = 0, limit: int = Query(default=100, le=500), db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_spoilage"))):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    skip, limit = _resolve_pagination(skip, limit)
    return await _build_spoilage_report(db, d_from=d_from, d_to=d_to, skip=skip, limit=limit)

@router.get("/export/spoilage", dependencies=[Depends(require_permission("action_export_excel")), Depends(require_permission("tab_reports_spoilage"))])
async def export_spoilage(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    data = await _build_spoilage_report(db, d_from=d_from, d_to=d_to, include_all=True)
    rows = [[r["ref"],r["product"],r["qty"],r["unit"],r["reason"],r["farm"],r["date"],r["user_name"],r["notes"]] for r in data["records"]]
    buf = to_xlsx(
        ["Ref #","Product","Qty","Unit","Reason","Farm","Date","Performed By","Notes"],
        rows,
        "Spoilage",
        report_title="Spoilage Report",
        metadata=[("Date Range", f"{d_from.strftime('%Y-%m-%d')} to {d_to.strftime('%Y-%m-%d')}"), ("Records", data["total_count"]), ("Total Qty", data["total_qty"])],
        column_formats={"Qty": "qty", "Date": "date"},
        wrap_columns={"Notes", "Reason"},
    )
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=spoilage_{date.today()}.xlsx"})


# ── PRODUCTION ─────────────────────────────────────────
@router.get("/api/production")
async def production_report(date_from: Optional[str] = None, date_to: Optional[str] = None, skip: int = 0, limit: int = Query(default=100, le=500), db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_production"))):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    skip, limit = _resolve_pagination(skip, limit)
    return await _build_production_report(db, d_from=d_from, d_to=d_to, skip=skip, limit=limit)


async def _build_production_report(db, *, d_from, d_to, skip=0, limit=100, include_all=False):
    batch_res = await db.execute(
        select(ProductionBatch)
        .where(ProductionBatch.created_at>=d_from, ProductionBatch.created_at<=d_to)
        .order_by(ProductionBatch.created_at.desc())
        .options(selectinload(ProductionBatch.inputs).selectinload(BatchInput.product),
                 selectinload(ProductionBatch.outputs).selectinload(BatchOutput.product),
                 selectinload(ProductionBatch.recipe), selectinload(ProductionBatch.user))
    )
    batches = batch_res.scalars().all()
    rows, losses, total_proc, total_pkg = [], [], 0, 0
    for b in batches:
        is_pkg = b.batch_number.startswith("PKG")
        inputs_str  = ", ".join(f"{float(i.qty):.0f}{i.product.unit if i.product else ''} {i.product.name if i.product else '—'}" for i in b.inputs)
        outputs_str = ", ".join(f"{float(o.qty):.0f}{o.product.unit if o.product else ''} {o.product.name if o.product else '—'}" for o in b.outputs)
        rows.append({"batch_number":b.batch_number,"type":"Packaging" if is_pkg else "Processing",
            "recipe":b.recipe.name if b.recipe else "Custom","waste_pct":float(b.waste_pct),
            "notes":b.notes or "","date":b.created_at.strftime("%Y-%m-%d") if b.created_at else "—",
            "inputs_str":inputs_str,"outputs_str":outputs_str,"user_name":b.user.name if b.user else "—"})
        if is_pkg: total_pkg  += 1
        else:      total_proc += 1; losses.append(float(b.waste_pct))
    # ── Drying batches ──────────────────────────────────
    drying_res = await db.execute(
        select(DryingBatch)
        .where(DryingBatch.started_at >= d_from, DryingBatch.started_at <= d_to)
        .order_by(DryingBatch.started_at.desc())
        .options(
            selectinload(DryingBatch.stages)
                .selectinload(DryingBatchStage.inputs)
                .selectinload(DryingBatchStageInput.product),
            selectinload(DryingBatch.stages)
                .selectinload(DryingBatchStage.outputs)
                .selectinload(DryingBatchStageOutput.product),
            selectinload(DryingBatch.started_by),
        )
    )
    drying_batches = drying_res.scalars().all()
    total_drying = len(drying_batches)
    for db_b in drying_batches:
        stages = db_b.stages or []
        stage1 = stages[0] if stages else None
        last_closed = next((s for s in reversed(stages) if s.total_output_qty is not None), None)
        inputs_str  = ", ".join(
            f"{float(i.qty):.0f}{i.product.unit if i.product else ''} {i.product.name if i.product else '—'}"
            for i in (stage1.inputs if stage1 else [])
        )
        outputs_str = ", ".join(
            f"{float(o.qty):.0f}{o.product.unit if o.product else ''} {o.product.name if o.product else '—'}"
            for o in (last_closed.outputs if last_closed else [])
        )
        yield_pct = float(last_closed.cumulative_yield_pct) if (last_closed and last_closed.cumulative_yield_pct is not None) else None
        waste_pct = (100.0 - yield_pct) if yield_pct is not None else 0.0
        rows.append({
            "batch_number": db_b.batch_number,
            "type": "Drying",
            "recipe": "—",
            "waste_pct": waste_pct,
            "notes": db_b.notes or "",
            "date": db_b.started_at.strftime("%Y-%m-%d") if db_b.started_at else "—",
            "inputs_str": inputs_str,
            "outputs_str": outputs_str,
            "user_name": db_b.started_by.name if db_b.started_by else "—",
        })
        if yield_pct is not None:
            losses.append(waste_pct)

    # Re-sort all rows by date desc
    rows.sort(key=lambda r: r["date"], reverse=True)
    total_batches = len(rows)
    if not include_all:
        rows = rows[skip : skip + limit]
    return {"batches":rows,"total_processing":total_proc,"total_packaging":total_pkg,"total_drying":total_drying,
            "avg_loss_pct":round(sum(losses)/len(losses),2) if losses else 0,"total_batches":total_batches}

@router.get("/export/production", dependencies=[Depends(require_permission("action_export_excel")), Depends(require_permission("tab_reports_production"))])
async def export_production(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    data = await _build_production_report(db, d_from=d_from, d_to=d_to, include_all=True)
    rows = [[b["batch_number"],b["type"],b["recipe"],b["inputs_str"],b["outputs_str"],b["waste_pct"],b["date"],b["user_name"],b["notes"]] for b in data["batches"]]
    buf = to_xlsx(
        ["Batch #","Type","Recipe","Inputs","Outputs","Loss %","Date","Performed By","Notes"],
        rows,
        "Production",
        report_title="Production Report",
        metadata=[
            ("Date Range", f"{d_from.strftime('%Y-%m-%d')} to {d_to.strftime('%Y-%m-%d')}"),
            ("Total Batches", data["total_batches"]),
            ("Processing Batches", data["total_processing"]),
            ("Packaging Batches", data["total_packaging"]),
            ("Drying Batches", data.get("total_drying", 0)),
            ("Average Loss %", f"{data['avg_loss_pct']:.2f}%"),
        ],
        column_formats={"Loss %": "percent_value", "Date": "date"},
        wrap_columns={"Inputs", "Outputs", "Notes"},
    )
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=production_{date.today()}.xlsx"})


# ── P&L ────────────────────────────────────────────────
@router.get("/api/pl")
async def pl_report(date_from: Optional[str] = None, date_to: Optional[str] = None, db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_pl"))):
    """
    Profit & Loss report.

    Uses the SAME revenue definition as the Dashboard and Sales Report so that
    all three surfaces always agree:

        Revenue = paid POS invoices (Invoice.total, status="paid")
                + paid B2B invoices (B2BInvoice.total, status="paid")
                - retail refunds    (RetailRefund.total)
                - B2B refunds       (B2BRefund.total)

    Expenses come from the Expenses module (Expense.expense_date).
    Date filtering uses APP_TIMEZONE-aware UTC bounds — same as the Dashboard.
    """
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    return await _build_pl_report(db, d_from=d_from, d_to=d_to)


async def _build_pl_report(db, *, d_from, d_to):

    # ── Revenue ──────────────────────────────────────────
    pos_result = await db.execute(
        select(func.coalesce(func.sum(Invoice.total), 0))
        .where(Invoice.created_at >= d_from, Invoice.created_at <= d_to, Invoice.status == "paid")
    )
    pos_sales = float(pos_result.scalar())

    b2b_result = await db.execute(
        select(func.coalesce(func.sum(B2BInvoice.total), 0))
        .where(B2BInvoice.created_at >= d_from, B2BInvoice.created_at <= d_to, B2BInvoice.status == "paid")
    )
    b2b_sales = float(b2b_result.scalar())

    retail_refund_result = await db.execute(
        select(func.coalesce(func.sum(RetailRefund.total), 0))
        .where(RetailRefund.created_at >= d_from, RetailRefund.created_at <= d_to)
    )
    retail_refunds = float(retail_refund_result.scalar())

    b2b_refund_result = await db.execute(
        select(func.coalesce(func.sum(B2BRefund.total), 0))
        .where(B2BRefund.created_at >= d_from, B2BRefund.created_at <= d_to)
    )
    b2b_refunds = float(b2b_refund_result.scalar())

    total_refunds  = retail_refunds + b2b_refunds
    total_revenue  = round(pos_sales + b2b_sales - total_refunds, 2)

    # ── Per-line entry drill-downs ──────────────────────
    # Fetch individual invoice/refund rows so the frontend can expand each
    # P&L line and show what's behind it.
    from app.core.time_utils import app_tz
    tz = app_tz()

    def _local_date(dt) -> str:
        if dt is None:
            return ""
        try:
            return dt.astimezone(tz).strftime("%Y-%m-%d")
        except Exception:
            return dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else ""

    # POS invoice entries
    pos_entries: list[dict] = []
    if abs(pos_sales) >= 0.01:
        pos_rows = await db.execute(
            select(Invoice.invoice_number, Invoice.total, Invoice.created_at)
            .where(Invoice.created_at >= d_from, Invoice.created_at <= d_to, Invoice.status == "paid")
            .order_by(Invoice.created_at)
        )
        pos_entries = [
            {
                "date":        _local_date(r.created_at),
                "ref_type":    "pos",
                "description": f"Invoice {r.invoice_number or '—'}",
                "amount":      round(float(r.total or 0), 2),
            }
            for r in pos_rows.all()
        ]

    # B2B invoice entries (with client name when available)
    b2b_entries: list[dict] = []
    if abs(b2b_sales) >= 0.01:
        b2b_rows = await db.execute(
            select(
                B2BInvoice.invoice_number,
                B2BInvoice.total,
                B2BInvoice.created_at,
                B2BClient.name.label("client_name"),
            )
            .join(B2BClient, B2BClient.id == B2BInvoice.client_id, isouter=True)
            .where(
                B2BInvoice.created_at >= d_from,
                B2BInvoice.created_at <= d_to,
                B2BInvoice.status == "paid",
            )
            .order_by(B2BInvoice.created_at)
        )
        for r in b2b_rows.all():
            client = r.client_name or ""
            desc = f"Invoice {r.invoice_number or '—'}" + (f" — {client}" if client else "")
            b2b_entries.append({
                "date":        _local_date(r.created_at),
                "ref_type":    "b2b",
                "description": desc,
                "amount":      round(float(r.total or 0), 2),
            })

    # Refund entries (retail + B2B, shown as negative)
    refund_entries: list[dict] = []
    if abs(total_refunds) >= 0.01:
        rr_rows = await db.execute(
            select(RetailRefund.refund_number, RetailRefund.total, RetailRefund.created_at)
            .where(RetailRefund.created_at >= d_from, RetailRefund.created_at <= d_to)
            .order_by(RetailRefund.created_at)
        )
        for r in rr_rows.all():
            refund_entries.append({
                "date":        _local_date(r.created_at),
                "ref_type":    "pos",
                "description": f"POS refund {r.refund_number or '—'}",
                "amount":      -round(float(r.total or 0), 2),
            })
        br_rows = await db.execute(
            select(B2BRefund.refund_number, B2BRefund.total, B2BRefund.created_at)
            .where(B2BRefund.created_at >= d_from, B2BRefund.created_at <= d_to)
            .order_by(B2BRefund.created_at)
        )
        for r in br_rows.all():
            refund_entries.append({
                "date":        _local_date(r.created_at),
                "ref_type":    "b2b",
                "description": f"B2B refund {r.refund_number or '—'}",
                "amount":      -round(float(r.total or 0), 2),
            })
        refund_entries.sort(key=lambda e: e["date"])

    # Build revenue_lines in the same shape the export and frontend expect:
    # [{code, name, amount, entries: [{date, ref_type, description, amount}]}]
    revenue_lines = []
    if abs(pos_sales) >= 0.01:
        revenue_lines.append({
            "code": "POS",
            "name": "Paid POS Sales",
            "amount": round(pos_sales, 2),
            "entries": pos_entries,
        })
    if abs(b2b_sales) >= 0.01:
        revenue_lines.append({
            "code": "B2B",
            "name": "Paid B2B Sales",
            "amount": round(b2b_sales, 2),
            "entries": b2b_entries,
        })
    if abs(total_refunds) >= 0.01:
        revenue_lines.append({
            "code": "REF",
            "name": "Refunds (deducted)",
            "amount": round(-total_refunds, 2),   # negative — reduces revenue
            "entries": refund_entries,
        })

    # ── Expenses ─────────────────────────────────────────
    # Grouped by expense category, filtered by expense_date (local date, no tz shift).
    local_from  = d_from.astimezone(tz).date()
    local_to    = d_to.astimezone(tz).date()

    expense_rows = await db.execute(
        select(
            ExpenseCategory.id.label("cat_id"),
            ExpenseCategory.account_code.label("code"),
            ExpenseCategory.name.label("name"),
            func.coalesce(func.sum(Expense.amount), 0).label("amount"),
        )
        .select_from(Expense)
        .join(ExpenseCategory, ExpenseCategory.id == Expense.category_id, isouter=True)
        .where(Expense.expense_date >= local_from, Expense.expense_date <= local_to)
        .group_by(ExpenseCategory.id, ExpenseCategory.account_code, ExpenseCategory.name)
        .order_by(ExpenseCategory.name)
    )

    # Individual expense entries per category, for drill-down
    detail_rows = await db.execute(
        select(
            Expense.category_id,
            Expense.ref_number,
            Expense.expense_date,
            Expense.amount,
            Expense.vendor,
            Expense.description,
        )
        .where(Expense.expense_date >= local_from, Expense.expense_date <= local_to)
        .order_by(Expense.category_id, Expense.expense_date, Expense.id)
    )
    expenses_by_cat: dict[Optional[int], list[dict]] = {}
    for r in detail_rows.all():
        desc_parts = [f"{r.ref_number}"] if r.ref_number else []
        if r.vendor:      desc_parts.append(str(r.vendor))
        if r.description: desc_parts.append(str(r.description))
        expenses_by_cat.setdefault(r.category_id, []).append({
            "date":        r.expense_date.strftime("%Y-%m-%d") if r.expense_date else "",
            "ref_type":    "manual",
            "description": " — ".join(desc_parts) if desc_parts else "Expense",
            "amount":      round(float(r.amount or 0), 2),
        })

    expense_lines = []
    for row in expense_rows.mappings().all():
        amt = round(float(row["amount"]), 2)
        if abs(amt) < 0.01:
            continue
        expense_lines.append({
            "code":    row["code"] or "OPC",
            "name":    row["name"] or "Operational Cost",
            "amount":  amt,
            "entries": expenses_by_cat.get(row["cat_id"], []),
        })

    total_expense = round(sum(e["amount"] for e in expense_lines), 2)
    net_profit    = round(total_revenue - total_expense, 2)

    return {
        "revenue_lines":  revenue_lines,
        "expense_lines":  expense_lines,
        "total_revenue":  total_revenue,
        "total_expense":  total_expense,
        "net_profit":     net_profit,
        "date_from":      d_from.astimezone(tz).strftime("%Y-%m-%d"),
        "date_to":        d_to.astimezone(tz).strftime("%Y-%m-%d"),
        "used_balance_fallback": False,
        "warning":        None,
    }


# ── Utilities (Water / Gas / Electricity / Fuel ...) ───────────────────
@router.get("/api/utilities")
async def utilities_report(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: AsyncSession = Depends(get_async_session),
    _=Depends(require_permission("tab_reports_utilities")),
):
    """
    Consumption-tracked expense categories report.

    Auto-detects any expense category that has a `unit_name` configured (e.g.
    "m³", "kWh", "litre") and surfaces it as a utility line. Returns:
      • totals by category (cost, consumption, EGP/unit)
      • last-12-month trend per category
      • per-farm breakdown
      • carbon kg CO₂e if the category has an emission factor mapped
    """
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    return await _build_utilities_report(db, d_from=d_from, d_to=d_to)


async def _build_utilities_report(db, *, d_from, d_to):
    from app.core.time_utils import app_tz
    from app.models.carbon import CarbonEmissionFactor, CarbonLog

    # Utilities tracking went live on 15 May 2026 — earlier dates have no
    # meaningful data, so we floor every query (range, trend, per-farm) at
    # that date to avoid showing misleading zero-or-partial months.
    UTILITIES_DATA_START = date(2026, 5, 15)

    tz = app_tz()
    local_from = d_from.astimezone(tz).date()
    local_to   = d_to.astimezone(tz).date()
    # Floor the requested window — never query data before tracking started
    if local_from < UTILITIES_DATA_START:
        local_from = UTILITIES_DATA_START
    if local_to < UTILITIES_DATA_START:
        # Entire range is before tracking started — return empty payload
        return {
            "date_from": UTILITIES_DATA_START.isoformat(),
            "date_to":   UTILITIES_DATA_START.isoformat(),
            "utilities": [],
            "trend":     [],
            "by_farm":   [],
            "totals":    {"cost": 0.0, "categories": 0, "carbon_kg_co2e": 0.0},
            "warning":   f"Utilities tracking started on {UTILITIES_DATA_START.isoformat()}. "
                         f"Please pick a date in that range or later.",
        }

    # 1) Pick up every active category that tracks consumption
    cat_rows = await db.execute(
        select(ExpenseCategory)
        .where(
            ExpenseCategory.is_active == "1",
            ExpenseCategory.unit_name.isnot(None),
            ExpenseCategory.unit_name != "",
        )
        .order_by(ExpenseCategory.name)
    )
    categories = cat_rows.scalars().all()

    if not categories:
        return {
            "date_from": local_from.isoformat(),
            "date_to":   local_to.isoformat(),
            "utilities": [],
            "trend":     [],
            "by_farm":   [],
            "totals":    {"cost": 0.0, "categories": 0, "carbon_kg_co2e": 0.0},
            "warning":   "No expense categories have a unit name configured yet. "
                         "Set a unit (e.g. m³, kWh, litre) on the category to track consumption.",
        }

    cat_ids   = [c.id for c in categories]
    cat_by_id = {c.id: c for c in categories}

    # 2) Per-category totals over the selected window
    totals_rows = await db.execute(
        select(
            Expense.category_id,
            func.coalesce(func.sum(Expense.amount), 0).label("cost"),
            func.coalesce(func.sum(Expense.consumption), 0).label("consumption"),
            func.count(Expense.id).label("entries"),
        )
        .where(
            Expense.category_id.in_(cat_ids),
            Expense.expense_date >= local_from,
            Expense.expense_date <= local_to,
        )
        .group_by(Expense.category_id)
    )
    totals_by_cat = {
        r.category_id: {"cost": float(r.cost or 0), "consumption": float(r.consumption or 0), "entries": int(r.entries or 0)}
        for r in totals_rows.all()
    }

    # 3) Carbon kg CO₂e per category — sum CarbonLog rows for "expense" refs in window,
    #    joined by ref_id back to expenses with this category.
    carbon_rows = await db.execute(
        select(
            Expense.category_id,
            func.coalesce(func.sum(CarbonLog.kg_co2e), 0).label("kg"),
        )
        .select_from(CarbonLog)
        .join(Expense, Expense.id == CarbonLog.ref_id)
        .where(
            CarbonLog.ref_type == "expense",
            Expense.category_id.in_(cat_ids),
            CarbonLog.log_date >= local_from,
            CarbonLog.log_date <= local_to,
        )
        .group_by(Expense.category_id)
    )
    carbon_by_cat = {r.category_id: float(r.kg or 0) for r in carbon_rows.all()}

    # 4) Build the per-utility list
    utilities = []
    total_cost = 0.0
    total_carbon = 0.0
    for c in categories:
        t = totals_by_cat.get(c.id, {"cost": 0.0, "consumption": 0.0, "entries": 0})
        if t["cost"] == 0 and t["consumption"] == 0:
            continue  # skip utilities with no activity in this window
        cost_per_unit = (t["cost"] / t["consumption"]) if t["consumption"] > 0 else None
        kg = carbon_by_cat.get(c.id, 0.0)
        utilities.append({
            "id":               c.id,
            "name":             c.name,
            "account_code":     c.account_code or "",
            "unit_name":        c.unit_name,
            "cost":             round(t["cost"], 2),
            "consumption":      round(t["consumption"], 4),
            "entries":          t["entries"],
            "cost_per_unit":    round(cost_per_unit, 4) if cost_per_unit is not None else None,
            "default_unit_price": float(c.unit_price) if c.unit_price is not None else None,
            "carbon_factor_key": c.carbon_factor_key or "",
            "carbon_kg_co2e":   round(kg, 4),
        })
        total_cost   += t["cost"]
        total_carbon += kg

    # 5) Last 12 months trend (rolling, anchored to today) for every active utility.
    # Months entirely before UTILITIES_DATA_START are skipped; the month
    # containing it is clipped so it only covers the in-range portion.
    from calendar import monthrange
    today = date.today()
    anchor = today.replace(day=1)
    months: list[tuple[date, date, str]] = []
    for i in range(11, -1, -1):
        y, m = anchor.year, anchor.month - i
        while m <= 0:
            m += 12
            y -= 1
        m_start = date(y, m, 1)
        m_end   = date(y, m, monthrange(y, m)[1])
        if m_end < UTILITIES_DATA_START:
            continue                                    # whole month pre-tracking
        if m_start < UTILITIES_DATA_START:
            m_start = UTILITIES_DATA_START              # clip the partial first month
        months.append((m_start, m_end, m_start.strftime("%b %y")))

    # one query covering all 12 months, group by (year, month, category)
    if cat_ids and months:
        earliest = months[0][0]
        latest   = months[-1][1]
        year_col  = func.extract("year",  Expense.expense_date)
        month_col = func.extract("month", Expense.expense_date)
        trend_rows = await db.execute(
            select(
                Expense.category_id,
                year_col.label("y"),
                month_col.label("m"),
                func.coalesce(func.sum(Expense.amount), 0).label("cost"),
                func.coalesce(func.sum(Expense.consumption), 0).label("consumption"),
            )
            .where(
                Expense.category_id.in_(cat_ids),
                Expense.expense_date >= earliest,
                Expense.expense_date <= latest,
            )
            .group_by(Expense.category_id, year_col, month_col)
        )
        trend_idx: dict[tuple[int, int, int], dict] = {}
        for r in trend_rows.all():
            try:
                key = (int(r.category_id), int(r.y), int(r.m))
            except (TypeError, ValueError):
                continue
            trend_idx[key] = {
                "cost": float(r.cost or 0),
                "consumption": float(r.consumption or 0),
            }
    else:
        trend_idx = {}

    trend = []
    for m_start, m_end, label in months:
        entry = {"month": m_start.strftime("%Y-%m"), "label": label, "items": []}
        for c in categories:
            cell = trend_idx.get((c.id, m_start.year, m_start.month), {"cost": 0.0, "consumption": 0.0})
            entry["items"].append({
                "id":          c.id,
                "name":        c.name,
                "unit_name":   c.unit_name,
                "cost":        round(cell["cost"], 2),
                "consumption": round(cell["consumption"], 4),
            })
        trend.append(entry)

    # 6) Per-farm breakdown over the selected window
    farm_rows = await db.execute(
        select(
            Expense.category_id,
            Expense.farm_id,
            Farm.name.label("farm_name"),
            func.coalesce(func.sum(Expense.amount), 0).label("cost"),
            func.coalesce(func.sum(Expense.consumption), 0).label("consumption"),
        )
        .select_from(Expense)
        .join(Farm, Farm.id == Expense.farm_id, isouter=True)
        .where(
            Expense.category_id.in_(cat_ids),
            Expense.expense_date >= local_from,
            Expense.expense_date <= local_to,
        )
        .group_by(Expense.category_id, Expense.farm_id, Farm.name)
        .order_by(Farm.name)
    )
    by_farm: list[dict] = []
    for r in farm_rows.all():
        c = cat_by_id.get(r.category_id)
        if c is None:
            continue
        by_farm.append({
            "category_id":   c.id,
            "category":      c.name,
            "unit_name":     c.unit_name,
            "farm_id":       r.farm_id,
            "farm_name":     r.farm_name or "Unassigned",
            "cost":          round(float(r.cost or 0), 2),
            "consumption":   round(float(r.consumption or 0), 4),
        })

    return {
        "date_from": local_from.isoformat(),
        "date_to":   local_to.isoformat(),
        "utilities": utilities,
        "trend":     trend,
        "by_farm":   by_farm,
        "totals": {
            "cost":           round(total_cost, 2),
            "categories":     len(utilities),
            "carbon_kg_co2e": round(total_carbon, 4),
        },
        "warning": None,
    }


@router.get("/export/utilities", dependencies=[Depends(require_permission("action_export_excel")), Depends(require_permission("tab_reports_utilities"))])
async def export_utilities(
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    db: AsyncSession = Depends(get_async_session),
):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    data = await _build_utilities_report(db, d_from=d_from, d_to=d_to)
    openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter = _excel_dependencies()
    wb = openpyxl.Workbook()

    green_fill = PatternFill("solid", fgColor="2a7a2a")
    head_font  = Font(bold=True, color="FFFFFF", size=11)
    thin       = Side(style="thin", color="CCCCCC")
    bord       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt        = PatternFill("solid", fgColor="F5FAF5")

    # Sheet 1 — Summary
    ws = wb.active
    ws.title = "Utilities Summary"
    ws.append(["Utilities Consumption Report"])
    ws["A1"].font = Font(bold=True, size=14, color="2a7a2a")
    ws.append([f"Period: {data['date_from']} to {data['date_to']}"])
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.append([])

    headers = ["Utility", "Unit", "Entries", "Consumption", "Cost (EGP)", "Cost / Unit", "Default Unit Price", "kg CO₂e"]
    ws.append(headers)
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=c_idx)
        cell.fill = green_fill
        cell.font = head_font
        cell.border = bord
        cell.alignment = Alignment(horizontal="center")

    for i, u in enumerate(data["utilities"], start=5):
        row = [
            u["name"],
            u["unit_name"] or "—",
            u["entries"],
            u["consumption"],
            u["cost"],
            u["cost_per_unit"] if u["cost_per_unit"] is not None else "—",
            u["default_unit_price"] if u["default_unit_price"] is not None else "—",
            u["carbon_kg_co2e"],
        ]
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c_idx, value=val)
            cell.border = bord
            if i % 2 == 0:
                cell.fill = alt
            if c_idx >= 3 and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00" if c_idx in (4, 5, 6, 7, 8) else "#,##0"
                cell.alignment = Alignment(horizontal="right")

    for c_idx, width in enumerate([26, 8, 10, 14, 14, 14, 18, 14], start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    # Sheet 2 — Monthly trend (one column per utility, cost-based)
    ws2 = wb.create_sheet("Monthly Cost Trend")
    ws2.append(["Month"] + [u["name"] for u in data["utilities"]])
    for c_idx in range(1, len(data["utilities"]) + 2):
        cell = ws2.cell(row=1, column=c_idx)
        cell.fill = green_fill
        cell.font = head_font
        cell.border = bord
        cell.alignment = Alignment(horizontal="center")
    for i, m in enumerate(data["trend"], start=2):
        ws2.cell(row=i, column=1, value=m["label"]).border = bord
        # build a lookup so column order matches the summary sheet exactly
        cost_by_id = {it["id"]: it["cost"] for it in m["items"]}
        for c_idx, u in enumerate(data["utilities"], start=2):
            cell = ws2.cell(row=i, column=c_idx, value=cost_by_id.get(u["id"], 0))
            cell.number_format = "#,##0.00"
            cell.border = bord
            cell.alignment = Alignment(horizontal="right")
            if i % 2 == 0:
                cell.fill = alt
    ws2.column_dimensions["A"].width = 12
    for c_idx in range(2, len(data["utilities"]) + 2):
        ws2.column_dimensions[get_column_letter(c_idx)].width = 18

    # Sheet 3 — Per-farm breakdown
    ws3 = wb.create_sheet("Per-Farm Breakdown")
    headers3 = ["Utility", "Farm", "Consumption", "Unit", "Cost (EGP)"]
    ws3.append(headers3)
    for c_idx in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=c_idx)
        cell.fill = green_fill
        cell.font = head_font
        cell.border = bord
        cell.alignment = Alignment(horizontal="center")
    for i, r in enumerate(data["by_farm"], start=2):
        row = [r["category"], r["farm_name"], r["consumption"], r["unit_name"] or "—", r["cost"]]
        for c_idx, val in enumerate(row, start=1):
            cell = ws3.cell(row=i, column=c_idx, value=val)
            cell.border = bord
            if i % 2 == 0:
                cell.fill = alt
            if c_idx in (3, 5):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
    for c_idx, width in enumerate([24, 22, 14, 8, 14], start=1):
        ws3.column_dimensions[get_column_letter(c_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"utilities_{data['date_from']}_to_{data['date_to']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/export/pl", dependencies=[Depends(require_permission("action_export_excel")), Depends(require_permission("tab_reports_pl"))])
async def export_pl(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    data = await _build_pl_report(db, d_from=d_from, d_to=d_to)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        green_fill  = PatternFill("solid", fgColor="2a7a2a")
        red_fill    = PatternFill("solid", fgColor="8a1a2a")
        blue_fill   = PatternFill("solid", fgColor="1a3a7a")
        white_font  = Font(bold=True, color="FFFFFF", size=11)
        thin  = Side(style="thin", color="CCCCCC")
        bord  = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt   = PatternFill("solid", fgColor="F5FAF5")
        alt2  = PatternFill("solid", fgColor="FFF5F5")
        total_font  = Font(bold=True, size=11)
        total_green = PatternFill("solid", fgColor="D0F0D0")
        total_red   = PatternFill("solid", fgColor="F0D0D0")
        section_font = Font(bold=True, size=12, color="FFFFFF")

        def add_cell(ws, row, col, value, fill=None, font=None, align="left"):
            c = ws.cell(row=row, column=col, value=value)
            c.border = bord
            c.alignment = Alignment(horizontal=align, vertical="center")
            if fill: c.fill = fill
            if font: c.font = font
            return c

        def auto_width(ws):
            for ci, col in enumerate(ws.columns, 1):
                mx = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(ci)].width = min(mx + 4, 50)

        # ── Sheet 1: P&L Summary ──
        ws1 = wb.active
        ws1.title = "P&L Summary"

        # Title row
        ws1.merge_cells("A1:D1")
        tc = ws1["A1"]
        tc.value = f"Profit & Loss Statement  |  {data['date_from']}  →  {data['date_to']}"
        tc.font = Font(bold=True, size=13, color="FFFFFF")
        tc.fill = green_fill
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 28

        # Headers
        ri = 2
        for ci, h in enumerate(["Category", "Code", "Account", "Amount (EGP)"], 1):
            add_cell(ws1, ri, ci, h, fill=PatternFill("solid", fgColor="4a4a4a"), font=Font(bold=True, color="FFFFFF", size=10), align="center")
        ri += 1

        # Revenue section header
        ws1.merge_cells(f"A{ri}:D{ri}")
        c = ws1.cell(row=ri, column=1, value="REVENUE")
        c.font = section_font; c.fill = green_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ri += 1

        for line in data["revenue_lines"]:
            fill = alt if ri % 2 == 0 else None
            add_cell(ws1, ri, 1, "Revenue", fill=fill)
            add_cell(ws1, ri, 2, line["code"], fill=fill)
            add_cell(ws1, ri, 3, line["name"], fill=fill)
            add_cell(ws1, ri, 4, line["amount"], fill=fill, align="right")
            ri += 1

        add_cell(ws1, ri, 1, "", fill=total_green)
        add_cell(ws1, ri, 2, "", fill=total_green)
        add_cell(ws1, ri, 3, "TOTAL REVENUE", fill=total_green, font=total_font)
        add_cell(ws1, ri, 4, data["total_revenue"], fill=total_green, font=Font(bold=True, size=12), align="right")
        ri += 2

        # Expenses section header
        ws1.merge_cells(f"A{ri}:D{ri}")
        c = ws1.cell(row=ri, column=1, value="EXPENSES")
        c.font = section_font; c.fill = red_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ri += 1

        for line in data["expense_lines"]:
            fill = alt2 if ri % 2 == 0 else None
            add_cell(ws1, ri, 1, "Expense", fill=fill)
            add_cell(ws1, ri, 2, line["code"], fill=fill)
            add_cell(ws1, ri, 3, line["name"], fill=fill)
            add_cell(ws1, ri, 4, line["amount"], fill=fill, align="right")
            ri += 1

        add_cell(ws1, ri, 1, "", fill=total_red)
        add_cell(ws1, ri, 2, "", fill=total_red)
        add_cell(ws1, ri, 3, "TOTAL EXPENSES", fill=total_red, font=total_font)
        add_cell(ws1, ri, 4, data["total_expense"], fill=total_red, font=Font(bold=True, size=12), align="right")
        ri += 2

        # Net profit/loss
        is_profit = data["net_profit"] >= 0
        net_fill = PatternFill("solid", fgColor="B0E0B0") if is_profit else PatternFill("solid", fgColor="E0B0B0")
        net_font = Font(bold=True, size=13, color="1a5a1a" if is_profit else "8a0000")
        ws1.merge_cells(f"A{ri}:C{ri}")
        c = ws1.cell(row=ri, column=1, value="NET PROFIT" if is_profit else "NET LOSS")
        c.font = net_font; c.fill = net_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bord
        ws1.cell(row=ri, column=2).border = bord
        ws1.cell(row=ri, column=3).border = bord
        add_cell(ws1, ri, 4, abs(data["net_profit"]), fill=net_fill, font=net_font, align="right")
        ws1.row_dimensions[ri].height = 24

        auto_width(ws1)

        # ── Sheet 2: Revenue Entries ──
        ws2 = wb.create_sheet("Revenue Entries")
        for ci, h in enumerate(["Account Code", "Account Name", "Date", "Type", "Description", "Amount (EGP)"], 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill = green_fill; c.font = white_font; c.border = bord
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 22
        ri = 2
        for line in data["revenue_lines"]:
            for entry in line["entries"]:
                fill = alt if ri % 2 == 0 else None
                for ci, val in enumerate([line["code"], line["name"], entry["date"], entry["ref_type"], entry["description"], entry["amount"]], 1):
                    c = ws2.cell(row=ri, column=ci, value=val)
                    c.border = bord
                    c.alignment = Alignment(vertical="center")
                    if fill: c.fill = fill
                ri += 1
        auto_width(ws2)

        # ── Sheet 3: Expense Entries ──
        ws3 = wb.create_sheet("Expense Entries")
        for ci, h in enumerate(["Account Code", "Account Name", "Date", "Type", "Description", "Amount (EGP)"], 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.fill = red_fill; c.font = white_font; c.border = bord
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 22
        ri = 2
        for line in data["expense_lines"]:
            for entry in line["entries"]:
                fill = alt2 if ri % 2 == 0 else None
                for ci, val in enumerate([line["code"], line["name"], entry["date"], entry["ref_type"], entry["description"], entry["amount"]], 1):
                    c = ws3.cell(row=ri, column=ci, value=val)
                    c.border = bord
                    c.alignment = Alignment(vertical="center")
                    if fill: c.fill = fill
                ri += 1
        auto_width(ws3)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=pl_report_{date.today()}.xlsx"})
    except ImportError:
        raise Exception("Run: pip install openpyxl --break-system-packages")


# ── TRANSACTIONS ────────────────────────────────────────
async def _build_transactions_report(
    db: AsyncSession,
    *,
    d_from: datetime,
    d_to: datetime,
    source: Optional[str] = None,
):
    from app.models.refund import RetailRefundItem

    b2b_payment_records = await _load_b2b_client_payment_records(db, d_from=d_from, d_to=d_to)
    rows = []

    if not source or source == "pos":
        pos_res = await db.execute(
            select(Invoice)
            .where(Invoice.created_at >= d_from, Invoice.created_at <= d_to)
            .options(selectinload(Invoice.items).selectinload(InvoiceItem.product), selectinload(Invoice.user), selectinload(Invoice.customer))
        )
        for inv in pos_res.scalars().all():
            for item in inv.items:
                product = getattr(item, "product", None)
                rows.append({
                    "date": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "—",
                    "_sort_date": _transaction_sort_value(inv.created_at),
                    "reference": inv.invoice_number,
                    "transaction_type": "POS Sale",
                    "source": "POS",
                    "counterparty_type": "Customer",
                    "counterparty_name": inv.customer.name if inv.customer else "Walk-in",
                    "sku": item.sku or "—",
                    "product": item.name or "—",
                    "product_category": _product_category(product),
                    "qty": _num(item.qty),
                    "unit_price": _num(item.unit_price),
                    "money_effect": _num(item.total),
                    "stock_effect": -_num(item.qty),
                    "direction": "out",
                    "payment_method": inv.payment_method or "cash",
                    "status": inv.status or "—",
                    "user_name": inv.user.name if inv.user else "—",
                    "notes": inv.notes or "",
                })

    if not source or source == "b2b":
        b2b_res = await db.execute(
            select(B2BInvoice)
            .where(B2BInvoice.created_at >= d_from, B2BInvoice.created_at <= d_to)
            .options(selectinload(B2BInvoice.items).selectinload(B2BInvoiceItem.product), selectinload(B2BInvoice.client), selectinload(B2BInvoice.user))
        )
        for inv in b2b_res.scalars().all():
            for item in inv.items:
                product = getattr(item, "product", None)
                rows.append({
                    "date": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "—",
                    "_sort_date": _transaction_sort_value(inv.created_at),
                    "reference": inv.invoice_number,
                    "transaction_type": "B2B Invoice",
                    "source": "B2B",
                    "counterparty_type": "Client",
                    "counterparty_name": inv.client.name if inv.client else "—",
                    "sku": product.sku if product else "—",
                    "product": product.name if product else "—",
                    "product_category": _product_category(product),
                    "qty": _num(item.qty),
                    "unit_price": _num(item.unit_price),
                    "money_effect": _num(item.total),
                    "stock_effect": -_num(item.qty),
                    "direction": "out",
                    "payment_method": inv.payment_method or inv.invoice_type,
                    "status": inv.status or "—",
                    "user_name": inv.user.name if inv.user else "—",
                    "notes": inv.notes or "",
                })
        for payment in b2b_payment_records:
            rows.append({
                "date": payment["datetime"],
                "_sort_date": _transaction_sort_value(payment["datetime"]),
                "reference": payment["reference"],
                "transaction_type": "B2B Client Payment",
                "source": "B2B Collection",
                "counterparty_type": "Client",
                "counterparty_name": payment["client"],
                "sku": "—",
                "product": "Consignment client payment",
                "product_category": "—",
                "qty": 0.0,
                "unit_price": payment["amount"],
                "money_effect": payment["amount"],
                "stock_effect": 0.0,
                "direction": "in",
                "payment_method": payment["payment_method"],
                "status": payment["status"],
                "user_name": payment["user_name"],
                "notes": payment["notes"],
            })

    if not source or source == "refund":
        refund_res = await db.execute(
            select(RetailRefund)
            .where(RetailRefund.created_at >= d_from, RetailRefund.created_at <= d_to)
            .options(selectinload(RetailRefund.customer), selectinload(RetailRefund.user), selectinload(RetailRefund.items).selectinload(RetailRefundItem.product))
        )
        for refund in refund_res.scalars().all():
            for item in refund.items:
                product = getattr(item, "product", None)
                rows.append({
                    "date": refund.created_at.strftime("%Y-%m-%d %H:%M") if refund.created_at else "—",
                    "_sort_date": _transaction_sort_value(refund.created_at),
                    "reference": refund.refund_number,
                    "transaction_type": "Retail Refund",
                    "source": "Refund",
                    "counterparty_type": "Customer",
                    "counterparty_name": refund.customer.name if refund.customer else "—",
                    "sku": product.sku if product else "—",
                    "product": product.name if product else "—",
                    "product_category": _product_category(product),
                    "qty": _num(item.qty),
                    "unit_price": _num(item.unit_price),
                    "money_effect": -_num(item.total),
                    "stock_effect": _num(item.qty),
                    "direction": "in",
                    "payment_method": refund.refund_method or "—",
                    "status": "refunded",
                    "user_name": refund.user.name if refund.user else "—",
                    "notes": refund.reason or "",
                })

    if not source or source == "receive":
        rec_res = await db.execute(
            select(ProductReceipt)
            .where(ProductReceipt.receive_date >= d_from.date(), ProductReceipt.receive_date <= d_to.date())
            .options(selectinload(ProductReceipt.product), selectinload(ProductReceipt.user), selectinload(ProductReceipt.expense))
        )
        for rec in rec_res.scalars().all():
            product = getattr(rec, "product", None)
            rows.append({
                "date": rec.receive_date.isoformat() if rec.receive_date else "—",
                "_sort_date": _transaction_sort_value(rec.receive_date),
                "reference": rec.ref_number,
                "transaction_type": "Stock Receipt",
                "source": "Receive",
                "counterparty_type": "Supplier",
                "counterparty_name": rec.supplier_ref or "—",
                "sku": product.sku if product else "—",
                "product": product.name if product else "—",
                "product_category": _product_category(product),
                "qty": _num(rec.qty),
                "unit_price": _num(rec.unit_cost),
                "money_effect": -_num(rec.total_cost),
                "stock_effect": _num(rec.qty),
                "direction": "in",
                "payment_method": rec.expense.payment_method if rec.expense and rec.expense.payment_method else "cash",
                "status": "received",
                "user_name": rec.user.name if rec.user else "—",
                "notes": rec.notes or "",
            })

    if not source or source == "expense":
        skip_receipt_linked_expenses = not source
        receipt_expense_ids = set()
        if skip_receipt_linked_expenses:
            receipt_expense_ids_res = await db.execute(select(ProductReceipt.expense_id).where(ProductReceipt.expense_id.is_not(None)))
            receipt_expense_ids = {row[0] for row in receipt_expense_ids_res.all() if row[0] is not None}
        exp_res = await db.execute(
            select(Expense)
            .where(Expense.expense_date >= d_from.date(), Expense.expense_date <= d_to.date())
            .options(selectinload(Expense.category), selectinload(Expense.user))
        )
        for exp in exp_res.scalars().all():
            if skip_receipt_linked_expenses and exp.id in receipt_expense_ids:
                continue
            rows.append({
                "date": exp.expense_date.isoformat() if exp.expense_date else "—",
                "_sort_date": _transaction_sort_value(exp.expense_date),
                "reference": exp.ref_number,
                "transaction_type": "Expense",
                "source": "Expense",
                "counterparty_type": "Vendor",
                "counterparty_name": exp.vendor or "—",
                "sku": "—",
                "product": exp.category.name if exp.category else "Expense",
                "product_category": "—",
                "qty": 0.0,
                "unit_price": _num(exp.amount),
                "money_effect": -_num(exp.amount),
                "stock_effect": 0.0,
                "direction": "out",
                "payment_method": exp.payment_method or "cash",
                "status": "posted",
                "user_name": exp.user.name if exp.user else "—",
                "notes": exp.description or "",
            })

    if not source or source == "production":
        move_res = await db.execute(
            select(StockMove)
            .where(
                StockMove.created_at >= d_from,
                StockMove.created_at <= d_to,
                StockMove.ref_type.in_(PRODUCTION_STOCK_MOVE_REF_TYPES),
            )
            .options(selectinload(StockMove.product), selectinload(StockMove.user))
        )
        for move in move_res.scalars().all():
            product = getattr(move, "product", None)
            tx_type, source_name, status = _production_move_label(move)
            stock_effect = _num(move.qty)
            rows.append({
                "date": move.created_at.strftime("%Y-%m-%d %H:%M") if move.created_at else "â€”",
                "_sort_date": _transaction_sort_value(move.created_at),
                "reference": _production_move_reference(move),
                "transaction_type": tx_type,
                "source": source_name,
                "counterparty_type": "Internal",
                "counterparty_name": "Production",
                "sku": product.sku if product else "â€”",
                "product": product.name if product else "â€”",
                "product_category": _product_category(product),
                "qty": abs(stock_effect),
                "unit_price": 0.0,
                "money_effect": 0.0,
                "stock_effect": stock_effect,
                "direction": "in" if stock_effect >= 0 else "out",
                "payment_method": "stock",
                "status": status,
                "user_name": move.user.name if move.user else "â€”",
                "notes": move.note or "",
            })

    rows.sort(key=lambda x: x["_sort_date"], reverse=True)
    for row in rows:
        row.pop("_sort_date", None)
    return {
        "rows": rows,
        "total_rows": len(rows),
        "money_in": round(sum(r["money_effect"] for r in rows if r["money_effect"] > 0), 2),
        "money_out": round(abs(sum(r["money_effect"] for r in rows if r["money_effect"] < 0)), 2),
        "net_money": round(sum(r["money_effect"] for r in rows), 2),
        "stock_in": round(sum(r["stock_effect"] for r in rows if r["stock_effect"] > 0), 2),
        "stock_out": round(abs(sum(r["stock_effect"] for r in rows if r["stock_effect"] < 0)), 2),
    }


@router.get("/api/transactions")
async def transactions_report(
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    source:    Optional[str] = None,
    db: AsyncSession = Depends(get_async_session),
    _=Depends(require_permission("tab_reports_transactions")),
):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    return await _build_transactions_report(db, d_from=d_from, d_to=d_to, source=source)

@router.get("/export/transactions", dependencies=[Depends(require_permission("action_export_excel")), Depends(require_permission("tab_reports_transactions"))])
async def export_transactions(date_from: str = None, date_to: str = None, source: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    data = await _build_transactions_report(db, d_from=d_from, d_to=d_to, source=source)
    wb = build_report_workbook([
        {
            "sheet_name": "Summary",
            "report_title": "Transactions Summary",
            "headers": ["Metric", "Value"],
            "rows": [
                ["Money In", data["money_in"]],
                ["Money Out", data["money_out"]],
                ["Net Money", data["net_money"]],
            ],
            "metadata": [("Date Range", f"{d_from.strftime('%Y-%m-%d')} to {d_to.strftime('%Y-%m-%d')}"), ("Source Filter", source or "All"), ("Rows", data["total_rows"])],
            "column_formats": {"Value": "money"},
            "tab_color": "1F4E78",
        },
        {
            "sheet_name": "Transactions",
            "report_title": "Transaction Detail",
            "headers": ["Date","Reference","Transaction Type","Source","Counterparty Type","Counterparty","Performed By","SKU","Product","Product Category","Qty","Unit Price","Money Effect","Stock Effect","Direction","Payment Method","Status","Notes"],
            "rows": [[r["date"], r["reference"], r["transaction_type"], r["source"], r["counterparty_type"], r["counterparty_name"], r["user_name"], r["sku"], r["product"], r["product_category"], r["qty"], r["unit_price"], r["money_effect"], r["stock_effect"], r["direction"], r["payment_method"], r["status"], r["notes"]] for r in data["rows"]],
            "metadata": [("Date Range", f"{d_from.strftime('%Y-%m-%d')} to {d_to.strftime('%Y-%m-%d')}"), ("Source Filter", source or "All"), ("Rows Exported", data["total_rows"])],
            "column_formats": {"Date": "datetime", "Qty": "qty", "Unit Price": "money", "Money Effect": "money", "Stock Effect": "qty"},
            "wrap_columns": {"Notes"},
            "tab_color": "2F6F4F",
        },
    ])
    buf = workbook_to_buffer(wb)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=transactions_{date.today()}.xlsx"})


def _hr_periods_for_range(start_date: date, end_date: date) -> list[str]:
    periods = []
    year = start_date.year
    month = start_date.month
    while (year, month) <= (end_date.year, end_date.month):
        periods.append(f"{year:04d}-{month:02d}")
        month += 1
        if month > 12:
            year += 1
            month = 1
    return periods


def _valid_hr_period(period: Optional[str]) -> Optional[str]:
    if not period:
        return None
    if not re.fullmatch(r"\d{4}-\d{2}", period):
        raise HTTPException(status_code=400, detail="Invalid period. Use YYYY-MM.")
    month = int(period[5:7])
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="Invalid period. Use YYYY-MM.")
    return period


def _attendance_rate(present_days: int, attendance_records: int) -> float:
    if attendance_records <= 0:
        return 0.0
    return round((present_days / attendance_records) * 100, 2)


def _hr_group_bucket(**extra):
    bucket = {
        "employees": 0,
        "base_salary": 0.0,
        "present_days": 0,
        "absent_days": 0,
        "late_days": 0,
        "leave_days": 0,
        "net_salary": 0.0,
    }
    bucket.update(extra)
    return bucket


def _hr_farm_bucket_key(employee: Employee, animal_employee_ids: set[int] | None = None) -> tuple[str, Optional[int]]:
    animal_employee_ids = animal_employee_ids or set()
    if getattr(employee, "works_with_animals", False) or employee.id in animal_employee_ids:
        return ("animals", None)
    if employee.farm_id is not None:
        return ("farm", employee.farm_id)
    return ("unassigned", None)


def _hr_farm_bucket_label(employee: Employee, animal_employee_ids: set[int] | None = None) -> str:
    animal_employee_ids = animal_employee_ids or set()
    if getattr(employee, "works_with_animals", False) or employee.id in animal_employee_ids:
        return "Animals"
    farm = getattr(employee, "farm", None)
    return farm.name if farm else "Unassigned"


async def _build_hr_report(
    db: AsyncSession,
    *,
    d_from: datetime,
    d_to: datetime,
    period: Optional[str] = None,
    department: Optional[str] = None,
    farm_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    include_all: bool = False,
):
    normalized_period = _valid_hr_period(period)
    periods = [normalized_period] if normalized_period else _hr_periods_for_range(d_from.date(), d_to.date())
    department_filter = (department or "").strip()

    employee_stmt = select(Employee).options(selectinload(Employee.farm))
    if department_filter:
        employee_stmt = employee_stmt.where(func.lower(Employee.department).like(f"%{department_filter.lower()}%"))
    if farm_id is not None:
        employee_stmt = employee_stmt.where(Employee.farm_id == farm_id)
    employee_stmt = employee_stmt.order_by(Employee.name)
    employees_res = await db.execute(employee_stmt)
    employee_map = {employee.id: employee for employee in employees_res.scalars().all()}
    filtered_employee_ids = set(employee_map)
    has_employee_filter = bool(department_filter) or farm_id is not None

    attendance_records = []
    if not has_employee_filter or filtered_employee_ids:
        attendance_stmt = select(Attendance).where(
            Attendance.date >= d_from.date(),
            Attendance.date <= d_to.date(),
        )
        if has_employee_filter:
            attendance_stmt = attendance_stmt.where(Attendance.employee_id.in_(filtered_employee_ids))
        attendance_res = await db.execute(attendance_stmt)
        attendance_records = attendance_res.scalars().all()

    payroll_deduction_columns_available = await _schema_has_columns(db, {
        "payroll": {
            "loan_deductions",
            "day_deduction_days",
            "day_deductions",
            "manual_deductions",
        },
    })
    payroll_payment_columns_available = await _schema_has_columns(db, {
        "payroll": {"paid_amount", "days_off_credited"},
    })
    hr_ledger_tables_available = await _schema_has_columns(db, {
        "employee_loans": {"id", "employee_id", "loan_date", "amount", "description", "status"},
        "employee_loan_repayments": {"id", "loan_id", "amount"},
        "employee_payroll_deductions": {
            "id",
            "employee_id",
            "payroll_id",
            "period",
            "type",
            "days",
            "daily_rate",
            "amount",
            "note",
            "created_at",
        },
    })

    payroll_records = []
    if periods and (not has_employee_filter or filtered_employee_ids):
        payroll_columns = [
            Payroll.id,
            Payroll.employee_id,
            Payroll.period,
            Payroll.base_salary,
            Payroll.bonuses,
            Payroll.deductions,
            Payroll.net_salary,
            Payroll.paid,
            Payroll.days_worked,
            Payroll.working_days,
        ]
        if payroll_deduction_columns_available:
            payroll_columns.extend([
                Payroll.loan_deductions,
                Payroll.day_deduction_days,
                Payroll.day_deductions,
                Payroll.manual_deductions,
            ])
        else:
            payroll_columns.extend([
                literal(0).label("loan_deductions"),
                literal(0).label("day_deduction_days"),
                literal(0).label("day_deductions"),
                literal(0).label("manual_deductions"),
            ])
        if payroll_payment_columns_available:
            payroll_columns.extend([
                Payroll.paid_amount,
                Payroll.days_off_credited,
            ])
        # When these columns don't exist yet, they're simply omitted from the
        # row; readers use getattr(..., default) so paid_amount falls back to
        # None (treated as a full cash payment) and days_off_credited to 0.
        payroll_stmt = select(*payroll_columns).where(Payroll.period.in_(periods))
        if has_employee_filter:
            payroll_stmt = payroll_stmt.where(Payroll.employee_id.in_(filtered_employee_ids))
        payroll_res = await db.execute(payroll_stmt)
        payroll_records = [SimpleNamespace(**dict(row._mapping)) for row in payroll_res.all()]

    active_ids = {employee.id for employee in employee_map.values() if employee.is_active}
    attendance_employee_ids = {record.employee_id for record in attendance_records}
    payroll_employee_ids = {record.employee_id for record in payroll_records}
    included_ids = active_ids | attendance_employee_ids | payroll_employee_ids
    included_ids &= filtered_employee_ids

    # ── Days-off credit (all-time snapshot, mirrors the Days Off tab) ──
    # Credit = monthly allowance accrued from hire (carried over)
    #          + days off earned via partial payroll payment
    #          - days taken (attendance marked "Day Off", stored as 'absent').
    days_off_credited_all_by_emp = defaultdict(float)
    days_off_taken_all_by_emp = defaultdict(int)
    if included_ids:
        if payroll_payment_columns_available:
            credited_res = await db.execute(
                select(Payroll.employee_id, func.coalesce(func.sum(Payroll.days_off_credited), 0))
                .where(Payroll.employee_id.in_(included_ids))
                .group_by(Payroll.employee_id)
            )
            for emp_id, total in credited_res.all():
                days_off_credited_all_by_emp[emp_id] = _num(total)
        taken_res = await db.execute(
            select(Attendance.employee_id, func.count(Attendance.id))
            .where(Attendance.employee_id.in_(included_ids), Attendance.status == "absent")
            .group_by(Attendance.employee_id)
        )
        for emp_id, cnt in taken_res.all():
            days_off_taken_all_by_emp[emp_id] = int(cnt or 0)

    loan_balance_by_employee = defaultdict(float)
    loan_history = []
    deduction_history = []
    if included_ids and hr_ledger_tables_available:
        loans_res = await db.execute(
            select(EmployeeLoan)
            .where(EmployeeLoan.employee_id.in_(included_ids))
            .order_by(EmployeeLoan.loan_date.desc(), EmployeeLoan.id.desc())
        )
        loans = loans_res.scalars().all()
        loan_ids = [loan.id for loan in loans]
        repayment_by_loan = defaultdict(float)
        if loan_ids:
            repayments_res = await db.execute(
                select(
                    EmployeeLoanRepayment.loan_id,
                    func.coalesce(func.sum(EmployeeLoanRepayment.amount), 0),
                )
                .where(EmployeeLoanRepayment.loan_id.in_(loan_ids))
                .group_by(EmployeeLoanRepayment.loan_id)
            )
            for loan_id, total in repayments_res.all():
                repayment_by_loan[loan_id] = _num(total)
        for loan in loans:
            repaid_amount = round(repayment_by_loan[loan.id], 2)
            balance = 0.0 if loan.status == "cancelled" else round(max(_num(loan.amount) - repaid_amount, 0), 2)
            if loan.status != "cancelled":
                loan_balance_by_employee[loan.employee_id] += balance
            employee = employee_map.get(loan.employee_id)
            loan_history.append({
                "loan_id": loan.id,
                "employee_id": loan.employee_id,
                "employee": employee.name if employee else f"Employee #{loan.employee_id}",
                "loan_date": loan.loan_date.isoformat() if loan.loan_date else "—",
                "amount": round(_num(loan.amount), 2),
                "repaid_amount": repaid_amount,
                "balance": balance,
                "status": loan.status,
                "description": loan.description or "",
            })

        deduction_stmt = (
            select(EmployeePayrollDeduction)
            .options(selectinload(EmployeePayrollDeduction.payroll))
            .where(EmployeePayrollDeduction.employee_id.in_(included_ids))
            .order_by(EmployeePayrollDeduction.created_at.desc(), EmployeePayrollDeduction.id.desc())
        )
        if periods:
            deduction_stmt = deduction_stmt.where(EmployeePayrollDeduction.period.in_(periods))
        deductions_res = await db.execute(deduction_stmt)
        for deduction in deductions_res.scalars().all():
            employee = employee_map.get(deduction.employee_id)
            payroll = getattr(deduction, "payroll", None)
            deduction_history.append({
                "deduction_id": deduction.id,
                "employee_id": deduction.employee_id,
                "employee": employee.name if employee else f"Employee #{deduction.employee_id}",
                "period": deduction.period or (payroll.period if payroll else "—"),
                "type": deduction.type,
                "days": _num(deduction.days) if deduction.days is not None else None,
                "daily_rate": _num(deduction.daily_rate) if deduction.daily_rate is not None else None,
                "amount": round(_num(deduction.amount), 2),
                "payroll_id": deduction.payroll_id,
                "note": deduction.note or "",
                "created_at": str(deduction.created_at) if deduction.created_at else "",
            })

    attendance_by_employee = defaultdict(lambda: {"present": 0, "absent": 0, "late": 0, "leave": 0, "records": 0})
    for record in attendance_records:
        if record.employee_id not in included_ids:
            continue
        status = (record.status or "").lower()
        stats = attendance_by_employee[record.employee_id]
        stats["records"] += 1
        if status in {"present", "absent", "late", "leave"}:
            stats[status] += 1

    payroll_by_employee = defaultdict(list)
    for record in payroll_records:
        if record.employee_id in included_ids:
            payroll_by_employee[record.employee_id].append(record)

    animal_employee_ids: set[int] = set()
    payroll_ids = [record.id for record in payroll_records if record.employee_id in included_ids]
    if payroll_ids:
        animal_expense_columns_available = await _schema_has_columns(db, {
            "expenses": {"payroll_id", "is_animal_expense"},
        })
        if animal_expense_columns_available:
            animal_expense_res = await db.execute(
                select(Payroll.employee_id)
                .join(Expense, Expense.payroll_id == Payroll.id)
                .where(
                    Payroll.id.in_(payroll_ids),
                    Expense.is_animal_expense == True,
                )
            )
            animal_employee_ids = {employee_id for (employee_id,) in animal_expense_res.all()}

    employee_rows = []
    departments = {}
    farms = {}
    active_count = 0
    inactive_count = 0
    total_base_salary = 0.0

    for employee in sorted((employee_map[employee_id] for employee_id in included_ids), key=lambda item: (item.name or "", item.id or 0)):
        is_active = bool(employee.is_active)
        if is_active:
            active_count += 1
        else:
            inactive_count += 1

        base_salary = _num(employee.base_salary)
        total_base_salary += base_salary
        att = attendance_by_employee[employee.id]
        payrolls = sorted(payroll_by_employee.get(employee.id, []), key=lambda item: item.period or "", reverse=True)
        payroll_period = payrolls[0].period if payrolls else "—"
        days_worked = sum(int(payroll.days_worked or 0) for payroll in payrolls)
        working_days = sum(int(payroll.working_days or 0) for payroll in payrolls)
        bonuses = round(sum(_num(payroll.bonuses) for payroll in payrolls), 2)
        deductions = round(sum(_num(payroll.deductions) for payroll in payrolls), 2)
        loan_deductions = round(sum(_num(getattr(payroll, "loan_deductions", 0)) for payroll in payrolls), 2)
        day_deduction_days = round(sum(_num(getattr(payroll, "day_deduction_days", 0)) for payroll in payrolls), 2)
        day_deductions = round(sum(_num(getattr(payroll, "day_deductions", 0)) for payroll in payrolls), 2)
        manual_deductions = round(sum(_num(getattr(payroll, "manual_deductions", 0)) for payroll in payrolls), 2)
        net_salary = round(sum(_num(payroll.net_salary) for payroll in payrolls), 2)
        paid = bool(payrolls) and all(bool(payroll.paid) for payroll in payrolls)
        # Cash actually paid out (partial payments pay less than net; the rest
        # becomes days off). Falls back to net for fully-paid older records.
        def _cash_paid(p):
            if not bool(p.paid):
                return 0.0
            pa = getattr(p, "paid_amount", None)
            return _num(pa) if pa is not None else _num(p.net_salary)
        paid_cash = round(sum(_cash_paid(payroll) for payroll in payrolls), 2)
        days_off_credited = round(sum(_num(getattr(payroll, "days_off_credited", 0)) for payroll in payrolls), 2)
        # All-time days-off credit balance (same definition as the Days Off tab)
        per_month_allowance = max(0, int(getattr(employee, "vacation_days_per_month", 0) or 0))
        hire = employee.hire_date or (employee.created_at.date() if getattr(employee, "created_at", None) else None)
        months_accrued = max(0, (date.today().year - hire.year) * 12 + (date.today().month - hire.month) + 1) if (hire and per_month_allowance > 0) else 0
        accrued_off = per_month_allowance * months_accrued
        days_off_credit_balance = round(accrued_off + days_off_credited_all_by_emp.get(employee.id, 0.0) - days_off_taken_all_by_emp.get(employee.id, 0), 2)
        farm_name = _hr_farm_bucket_label(employee, animal_employee_ids)
        department_name = employee.department or "Unassigned"

        row = {
            "employee_id": employee.id,
            "employee": employee.name or "—",
            "phone": employee.phone or "—",
            "position": employee.position or "—",
            "department": department_name,
            "farm_name": farm_name,
            "hire_date": employee.hire_date.isoformat() if employee.hire_date else "—",
            "base_salary": round(base_salary, 2),
            "present_days": att["present"],
            "absent_days": att["absent"],
            "late_days": att["late"],
            "leave_days": att["leave"],
            "attendance_records": att["records"],
            "attendance_rate": _attendance_rate(att["present"], att["records"]),
            "payroll_period": payroll_period,
            "days_worked": days_worked,
            "working_days": working_days,
            "bonuses": bonuses,
            "outstanding_loan_balance": round(loan_balance_by_employee[employee.id], 2),
            "loan_deductions": loan_deductions,
            "day_deduction_days": day_deduction_days,
            "day_deductions": day_deductions,
            "manual_deductions": manual_deductions,
            "total_deductions": deductions,
            "deductions": deductions,
            "net_salary": net_salary,
            "paid": paid,
            "paid_cash": paid_cash,
            "days_off_credited": days_off_credited,
            "vacation_days_per_month": per_month_allowance,
            "days_off_credit_balance": days_off_credit_balance,
        }
        employee_rows.append(row)

        dept_bucket = departments.setdefault(department_name, _hr_group_bucket(department=department_name))
        dept_bucket["employees"] += 1
        dept_bucket["base_salary"] += base_salary
        dept_bucket["present_days"] += att["present"]
        dept_bucket["absent_days"] += att["absent"]
        dept_bucket["late_days"] += att["late"]
        dept_bucket["leave_days"] += att["leave"]
        dept_bucket["net_salary"] += net_salary

        farm_bucket = farms.setdefault(
            _hr_farm_bucket_key(employee, animal_employee_ids),
            _hr_group_bucket(
                farm_id=employee.farm_id,
                farm_name=farm_name,
            ),
        )
        farm_bucket["employees"] += 1
        farm_bucket["base_salary"] += base_salary
        farm_bucket["present_days"] += att["present"]
        farm_bucket["absent_days"] += att["absent"]
        farm_bucket["late_days"] += att["late"]
        farm_bucket["leave_days"] += att["leave"]
        farm_bucket["net_salary"] += net_salary

    by_department = sorted(
        (
            {
                **bucket,
                "base_salary": round(bucket["base_salary"], 2),
                "net_salary": round(bucket["net_salary"], 2),
            }
            for bucket in departments.values()
        ),
        key=lambda item: item["department"],
    )
    by_farm = sorted(
        (
            {
                **bucket,
                "base_salary": round(bucket["base_salary"], 2),
                "net_salary": round(bucket["net_salary"], 2),
            }
            for bucket in farms.values()
        ),
        key=lambda item: (item["farm_name"], item["farm_id"] or 0),
    )

    total_attendance_records = sum(row["attendance_records"] for row in employee_rows)
    present_days = sum(row["present_days"] for row in employee_rows)
    absent_days = sum(row["absent_days"] for row in employee_rows)
    late_days = sum(row["late_days"] for row in employee_rows)
    leave_days = sum(row["leave_days"] for row in employee_rows)
    gross_salary = round(sum(_num(payroll.base_salary) for payroll in payroll_records if payroll.employee_id in included_ids), 2)
    bonuses_total = round(sum(_num(payroll.bonuses) for payroll in payroll_records if payroll.employee_id in included_ids), 2)
    loan_deductions_total = round(sum(_num(getattr(payroll, "loan_deductions", 0)) for payroll in payroll_records if payroll.employee_id in included_ids), 2)
    day_deduction_days_total = round(sum(_num(getattr(payroll, "day_deduction_days", 0)) for payroll in payroll_records if payroll.employee_id in included_ids), 2)
    day_deductions_total = round(sum(_num(getattr(payroll, "day_deductions", 0)) for payroll in payroll_records if payroll.employee_id in included_ids), 2)
    manual_deductions_total = round(sum(_num(getattr(payroll, "manual_deductions", 0)) for payroll in payroll_records if payroll.employee_id in included_ids), 2)
    deductions_total = round(sum(_num(payroll.deductions) for payroll in payroll_records if payroll.employee_id in included_ids), 2)
    net_salary_total = round(sum(_num(payroll.net_salary) for payroll in payroll_records if payroll.employee_id in included_ids), 2)
    paid_salary = round(sum(_num(payroll.net_salary) for payroll in payroll_records if payroll.employee_id in included_ids and payroll.paid), 2)
    paid_cash_total = round(sum(
        (_num(getattr(payroll, "paid_amount", None)) if getattr(payroll, "paid_amount", None) is not None else _num(payroll.net_salary))
        for payroll in payroll_records if payroll.employee_id in included_ids and payroll.paid
    ), 2)
    days_off_credited_total = round(sum(_num(getattr(payroll, "days_off_credited", 0)) for payroll in payroll_records if payroll.employee_id in included_ids), 2)
    days_off_credit_balance_total = round(sum(row["days_off_credit_balance"] for row in employee_rows), 2)
    total_outstanding_loans = round(sum(loan_balance_by_employee.values()), 2)

    return {
        "date_from": d_from.strftime("%Y-%m-%d"),
        "date_to": d_to.strftime("%Y-%m-%d"),
        "period": normalized_period,
        "summary": {
            "active_employees": active_count,
            "inactive_employees": inactive_count,
            "total_base_salary": round(total_base_salary, 2),
            "attendance_records": total_attendance_records,
            "present_days": present_days,
            "absent_days": absent_days,
            "late_days": late_days,
            "leave_days": leave_days,
            "attendance_rate": _attendance_rate(present_days, total_attendance_records),
            "payroll_records": len([payroll for payroll in payroll_records if payroll.employee_id in included_ids]),
            "gross_salary": gross_salary,
            "bonuses": bonuses_total,
            "total_outstanding_loans": total_outstanding_loans,
            "total_loan_deductions": loan_deductions_total,
            "total_day_deduction_days": day_deduction_days_total,
            "total_day_deductions": day_deductions_total,
            "total_manual_deductions": manual_deductions_total,
            "deductions": deductions_total,
            "net_salary": net_salary_total,
            "paid_salary": paid_salary,
            "unpaid_salary": round(net_salary_total - paid_salary, 2),
            "paid_cash": paid_cash_total,
            "salary_settled_as_days_off": round(paid_salary - paid_cash_total, 2),
            "days_off_credited": days_off_credited_total,
            "days_off_credit_balance": days_off_credit_balance_total,
        },
        "by_department": by_department,
        "by_farm": by_farm,
        "employees": _paginate_rows(employee_rows, skip, limit, include_all=include_all),
        "loans": loan_history,
        "deduction_history": deduction_history,
        "total_rows": len(employee_rows),
    }


@router.get("/api/hr")
async def hr_report(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    period: Optional[str] = None,
    department: Optional[str] = None,
    farm_id: Optional[int] = None,
    skip: int = 0,
    limit: int = Query(default=100, le=500),
    db: AsyncSession = Depends(get_async_session),
    _=Depends(require_permission("tab_reports_hr")),
):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    skip, limit = _resolve_pagination(skip, limit)
    return await _build_hr_report(
        db,
        d_from=d_from,
        d_to=d_to,
        period=period,
        department=department,
        farm_id=farm_id,
        skip=skip,
        limit=limit,
    )


@router.get("/export/hr", dependencies=[Depends(require_permission("action_export_excel"))])
async def export_hr(
    date_from: str = None,
    date_to: str = None,
    period: str = None,
    department: str = None,
    farm_id: Optional[int] = None,
    db: AsyncSession = Depends(get_async_session),
    _=Depends(require_permission("tab_reports_hr")),
):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    data = await _build_hr_report(
        db,
        d_from=d_from,
        d_to=d_to,
        period=period,
        department=department,
        farm_id=farm_id,
        include_all=True,
    )
    summary = data["summary"]
    wb = build_report_workbook([
        {
            "sheet_name": "HR Summary",
            "report_title": "HR Report Summary",
            "headers": ["Metric", "Value"],
            "rows": [
                ["Active Employees", summary["active_employees"]],
                ["Inactive Employees", summary["inactive_employees"]],
                ["Total Base Salary", summary["total_base_salary"]],
                ["Attendance Records", summary["attendance_records"]],
                ["Present Days", summary["present_days"]],
                ["Absent Days", summary["absent_days"]],
                ["Late Days", summary["late_days"]],
                ["Leave Days", summary["leave_days"]],
                ["Attendance Rate", summary["attendance_rate"]],
                ["Payroll Records", summary["payroll_records"]],
                ["Gross Salary", summary["gross_salary"]],
                ["Bonuses", summary["bonuses"]],
                ["Outstanding Loan Balance", summary["total_outstanding_loans"]],
                ["Loan Deductions", summary["total_loan_deductions"]],
                ["Day Deduction Days", summary["total_day_deduction_days"]],
                ["Day Deductions", summary["total_day_deductions"]],
                ["Manual Deductions", summary["total_manual_deductions"]],
                ["Deductions", summary["deductions"]],
                ["Net Salary", summary["net_salary"]],
                ["Paid Salary", summary["paid_salary"]],
                ["Unpaid Salary", summary["unpaid_salary"]],
                ["Paid in Cash", summary.get("paid_cash", summary["paid_salary"])],
                ["Salary Settled as Days Off", summary.get("salary_settled_as_days_off", 0)],
                ["Days Off Credited (from payroll)", summary.get("days_off_credited", 0)],
                ["Days Off Credit Balance", summary.get("days_off_credit_balance", 0)],
            ],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Payroll Period", data["period"] or "Range months")],
            "tab_color": "1F4E78",
        },
        {
            "sheet_name": "By Department",
            "report_title": "HR By Department",
            "headers": ["Department", "Employees", "Base Salary", "Present Days", "Absent Days", "Late Days", "Leave Days", "Net Salary"],
            "rows": [[row["department"], row["employees"], row["base_salary"], row["present_days"], row["absent_days"], row["late_days"], row["leave_days"], row["net_salary"]] for row in data["by_department"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Rows", len(data["by_department"]))],
            "column_formats": {"Employees": "int", "Base Salary": "money", "Present Days": "int", "Absent Days": "int", "Late Days": "int", "Leave Days": "int", "Net Salary": "money"},
            "tab_color": "2F6F4F",
        },
        {
            "sheet_name": "By Farm",
            "report_title": "HR By Farm",
            "headers": ["Farm ID", "Farm", "Employees", "Base Salary", "Present Days", "Absent Days", "Late Days", "Leave Days", "Net Salary"],
            "rows": [[row["farm_id"], row["farm_name"], row["employees"], row["base_salary"], row["present_days"], row["absent_days"], row["late_days"], row["leave_days"], row["net_salary"]] for row in data["by_farm"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Rows", len(data["by_farm"]))],
            "column_formats": {"Employees": "int", "Base Salary": "money", "Present Days": "int", "Absent Days": "int", "Late Days": "int", "Leave Days": "int", "Net Salary": "money"},
            "tab_color": "5B7F95",
        },
        {
            "sheet_name": "Employees",
            "report_title": "HR Employee Detail",
            "headers": ["Employee ID", "Employee", "Phone", "Position", "Department", "Farm", "Hire Date", "Base Salary", "Present Days", "Absent Days", "Late Days", "Leave Days", "Attendance Records", "Attendance Rate", "Payroll Period", "Days Worked", "Working Days", "Bonuses", "Outstanding Loan Balance", "Loan Deductions", "Day Deduction Days", "Day Deductions", "Manual Deductions", "Total Deductions", "Net Salary", "Paid", "Paid in Cash", "Days Off Credited", "Days Off / Month", "Days Off Credit Balance"],
            "rows": [[row["employee_id"], row["employee"], row["phone"], row["position"], row["department"], row["farm_name"], row["hire_date"], row["base_salary"], row["present_days"], row["absent_days"], row["late_days"], row["leave_days"], row["attendance_records"], row["attendance_rate"], row["payroll_period"], row["days_worked"], row["working_days"], row["bonuses"], row["outstanding_loan_balance"], row["loan_deductions"], row["day_deduction_days"], row["day_deductions"], row["manual_deductions"], row["total_deductions"], row["net_salary"], "Yes" if row["paid"] else "No", row.get("paid_cash", 0), row.get("days_off_credited", 0), row.get("vacation_days_per_month", 0), row.get("days_off_credit_balance", 0)] for row in data["employees"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Payroll Period", data["period"] or "Range months"), ("Rows", data["total_rows"])],
            "column_formats": {"Employee ID": "int", "Hire Date": "date", "Base Salary": "money", "Present Days": "int", "Absent Days": "int", "Late Days": "int", "Leave Days": "int", "Attendance Records": "int", "Attendance Rate": "percent_value", "Days Worked": "int", "Working Days": "int", "Bonuses": "money", "Outstanding Loan Balance": "money", "Loan Deductions": "money", "Day Deductions": "money", "Manual Deductions": "money", "Total Deductions": "money", "Net Salary": "money", "Paid in Cash": "money", "Days Off / Month": "int"},
            "tab_color": "7C3AED",
        },
        {
            "sheet_name": "Loan History",
            "report_title": "HR Loan History",
            "headers": ["Loan ID", "Employee ID", "Employee", "Loan Date", "Amount", "Repaid", "Balance", "Status", "Description"],
            "rows": [[row["loan_id"], row["employee_id"], row["employee"], row["loan_date"], row["amount"], row["repaid_amount"], row["balance"], row["status"], row["description"]] for row in data["loans"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Rows", len(data["loans"]))],
            "column_formats": {"Loan ID": "int", "Employee ID": "int", "Loan Date": "date", "Amount": "money", "Repaid": "money", "Balance": "money"},
            "tab_color": "B45309",
        },
        {
            "sheet_name": "Deductions",
            "report_title": "HR Deduction History",
            "headers": ["Deduction ID", "Employee ID", "Employee", "Period", "Type", "Days", "Daily Rate", "Amount", "Payroll ID", "Note", "Created At"],
            "rows": [[row["deduction_id"], row["employee_id"], row["employee"], row["period"], row["type"], row["days"], row["daily_rate"], row["amount"], row["payroll_id"], row["note"], row["created_at"]] for row in data["deduction_history"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Rows", len(data["deduction_history"]))],
            "column_formats": {"Deduction ID": "int", "Employee ID": "int", "Daily Rate": "money", "Amount": "money", "Payroll ID": "int"},
            "tab_color": "BE123C",
        },
    ])
    buf = workbook_to_buffer(wb)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=hr_report_{date.today()}.xlsx"})


def _animal_age(bd) -> Optional[str]:
    """Compact age from a birth date: '2y 3m', '5 mo', '12 days'.
    Returns None when the date is missing or in the future. Accepts a
    date object or an ISO date string."""
    if not bd:
        return None
    if isinstance(bd, str):
        try:
            bd = date.fromisoformat(bd[:10])
        except ValueError:
            return None
    today = date.today()
    if bd > today:
        return None
    years = today.year - bd.year
    months = today.month - bd.month
    days = today.day - bd.day
    if days < 0:
        months -= 1
    if months < 0:
        years -= 1
        months += 12
    if years >= 1:
        return f"{years}y {months}m" if months else f"{years}y"
    if months >= 1:
        return f"{months} mo"
    total_days = (today - bd).days
    return f"{total_days} day{'s' if total_days != 1 else ''}"


async def _build_animals_report(db, *, d_from, d_to, include_all=False):
    """Per-group animal activity over a date range: headcount, intakes
    (purchase/birth/transfer), deaths, purchase cost, feeding cost, cost/head."""
    from collections import defaultdict

    groups = (
        await db.execute(select(AnimalGroup).options(selectinload(AnimalGroup.farm)))
    ).scalars().all()

    intakes = (
        await db.execute(
            select(AnimalIntakeLog)
            .options(selectinload(AnimalIntakeLog.group))
            .where(AnimalIntakeLog.intake_date >= d_from, AnimalIntakeLog.intake_date <= d_to)
            .order_by(AnimalIntakeLog.intake_date.desc(), AnimalIntakeLog.id.desc())
        )
    ).scalars().all()

    deaths = (
        await db.execute(
            select(MortalityLog)
            .options(selectinload(MortalityLog.group))
            .where(MortalityLog.death_date >= d_from, MortalityLog.death_date <= d_to)
            .order_by(MortalityLog.death_date.desc(), MortalityLog.id.desc())
        )
    ).scalars().all()

    feedings = (
        await db.execute(
            select(FeedingLog)
            .options(selectinload(FeedingLog.product))
            .where(FeedingLog.feed_date >= d_from, FeedingLog.feed_date <= d_to)
        )
    ).scalars().all()

    per = defaultdict(lambda: {
        "received": 0, "born": 0, "transferred": 0, "died": 0,
        "purchase_cost": 0.0, "feeding_cost": 0.0, "feed_qty": 0.0,
    })
    for i in intakes:
        p = per[i.animal_group_id]
        c = int(i.count or 0)
        t = (getattr(i, "intake_type", None) or "purchase").lower()
        if t == "birth":
            p["born"] += c
        elif t == "transfer":
            p["transferred"] += c
        else:
            p["received"] += c  # purchase + other
        p["purchase_cost"] += float(i.total_cost or 0)
    for d in deaths:
        per[d.animal_group_id]["died"] += int(d.count or 0)
    for f in feedings:
        unit_cost = float(f.product.cost) if (f.product and f.product.cost is not None) else 0.0
        per[f.animal_group_id]["feeding_cost"] += float(f.qty or 0) * unit_cost
        per[f.animal_group_id]["feed_qty"] += float(f.qty or 0)

    group_rows = []
    for g in groups:
        p = per.get(g.id, {})
        received    = int(p.get("received", 0))
        born        = int(p.get("born", 0))
        transferred = int(p.get("transferred", 0))
        died        = int(p.get("died", 0))
        purchase_cost = round(float(p.get("purchase_cost", 0.0)), 2)
        feeding_cost  = round(float(p.get("feeding_cost", 0.0)), 2)
        head = int(g.headcount or 0)
        total_cost = round(purchase_cost + feeding_cost, 2)
        cost_per_head = round(total_cost / head, 2) if head > 0 else 0.0
        had_activity = received or born or transferred or died or purchase_cost or feeding_cost
        if not include_all and not (g.status == "active" or had_activity):
            continue
        group_rows.append({
            "group_id": g.id,
            "name": g.name,
            "animal_type": g.animal_type,
            "farm_name": g.farm.name if g.farm else None,
            "status": g.status,
            "headcount": head,
            "male_count": int(g.male_count) if g.male_count is not None else None,
            "female_count": int(g.female_count) if g.female_count is not None else None,
            "birth_date": g.birth_date.isoformat() if g.birth_date else None,
            "age": _animal_age(g.birth_date),
            "received": received,
            "born": born,
            "transferred": transferred,
            "received_total": received + born + transferred,
            "died": died,
            "net_change": received + born + transferred - died,
            "purchase_cost": purchase_cost,
            "feeding_cost": feeding_cost,
            "total_cost": total_cost,
            "cost_per_head": cost_per_head,
        })
    group_rows.sort(key=lambda r: r["headcount"], reverse=True)

    summary = {
        "active_groups":   sum(1 for g in groups if g.status == "active"),
        "total_headcount": sum(int(g.headcount or 0) for g in groups if g.status == "active"),
        "male_count":   sum(int(g.male_count or 0) for g in groups if g.status == "active"),
        "female_count": sum(int(g.female_count or 0) for g in groups if g.status == "active"),
        "received":     sum(r["received"] for r in group_rows),
        "born":         sum(r["born"] for r in group_rows),
        "transferred":  sum(r["transferred"] for r in group_rows),
        "died":         sum(r["died"] for r in group_rows),
        "purchase_cost": round(sum(r["purchase_cost"] for r in group_rows), 2),
        "feeding_cost":  round(sum(r["feeding_cost"] for r in group_rows), 2),
        "total_cost":    round(sum(r["total_cost"] for r in group_rows), 2),
    }
    summary["net_change"] = summary["received"] + summary["born"] + summary["transferred"] - summary["died"]

    intake_rows = [{
        "date":   i.intake_date.isoformat() if i.intake_date else None,
        "group":  i.group.name if i.group else None,
        "type":   getattr(i, "intake_type", None) or "purchase",
        "count":  int(i.count or 0),
        "male_count":   int(i.male_count) if getattr(i, "male_count", None) is not None else None,
        "female_count": int(i.female_count) if getattr(i, "female_count", None) is not None else None,
        "source": i.source,
        "cost":   float(i.total_cost) if i.total_cost is not None else None,
    } for i in intakes]

    mortality_rows = [{
        "date":  d.death_date.isoformat() if d.death_date else None,
        "group": d.group.name if d.group else None,
        "count": int(d.count or 0),
        "cause": d.cause or "unknown",
        "note":  d.note,
    } for d in deaths]

    return {
        "date_from": d_from.isoformat(),
        "date_to":   d_to.isoformat(),
        "summary":   summary,
        "groups":    group_rows,
        "intakes":   intake_rows,
        "mortality": mortality_rows,
    }


@router.get("/api/animals")
async def animals_report(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: AsyncSession = Depends(get_async_session),
    _=Depends(require_permission("tab_reports_animals")),
):
    d_from, d_to = _plain_date_range(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    return await _build_animals_report(db, d_from=d_from, d_to=d_to)


@router.get("/export/animals", dependencies=[Depends(require_permission("action_export_excel"))])
async def export_animals(
    date_from: str = None,
    date_to: str = None,
    db: AsyncSession = Depends(get_async_session),
    _=Depends(require_permission("tab_reports_animals")),
):
    d_from, d_to = _plain_date_range(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    data = await _build_animals_report(db, d_from=d_from, d_to=d_to, include_all=True)
    s = data["summary"]
    wb = build_report_workbook([
        {
            "sheet_name": "Animals Summary",
            "report_title": "Animals Report Summary",
            "headers": ["Metric", "Value"],
            "rows": [
                ["Active Groups", s["active_groups"]],
                ["Total Headcount", s["total_headcount"]],
                ["Males (active groups)", s["male_count"]],
                ["Females (active groups)", s["female_count"]],
                ["Received (purchased)", s["received"]],
                ["Born", s["born"]],
                ["Transferred In", s["transferred"]],
                ["Died", s["died"]],
                ["Net Change", s["net_change"]],
                ["Purchase Cost", s["purchase_cost"]],
                ["Feeding Cost", s["feeding_cost"]],
                ["Total Animal Cost", s["total_cost"]],
            ],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}")],
            "tab_color": "1F4E78",
        },
        {
            "sheet_name": "By Group",
            "report_title": "Animals by Group",
            "headers": ["Group ID", "Group", "Type", "Farm", "Status", "Headcount", "Males", "Females", "Birth Date", "Age", "Received", "Born", "Transferred", "Died", "Net Change", "Purchase Cost", "Feeding Cost", "Total Cost", "Cost / Head"],
            "rows": [[r["group_id"], r["name"], r["animal_type"], r["farm_name"], r["status"], r["headcount"], r["male_count"], r["female_count"], r["birth_date"], r["age"], r["received"], r["born"], r["transferred"], r["died"], r["net_change"], r["purchase_cost"], r["feeding_cost"], r["total_cost"], r["cost_per_head"]] for r in data["groups"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Rows", len(data["groups"]))],
            "column_formats": {"Group ID": "int", "Headcount": "int", "Males": "int", "Females": "int", "Birth Date": "date", "Received": "int", "Born": "int", "Transferred": "int", "Died": "int", "Net Change": "int", "Purchase Cost": "money", "Feeding Cost": "money", "Total Cost": "money", "Cost / Head": "money"},
            "tab_color": "2F6F4F",
        },
        {
            "sheet_name": "Intake Log",
            "report_title": "Animal Intake (Receive) Log",
            "headers": ["Date", "Group", "Type", "Count", "Males", "Females", "Source", "Cost"],
            "rows": [[r["date"], r["group"], r["type"], r["count"], r["male_count"], r["female_count"], r["source"], r["cost"]] for r in data["intakes"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Rows", len(data["intakes"]))],
            "column_formats": {"Count": "int", "Males": "int", "Females": "int", "Cost": "money"},
            "tab_color": "7C3AED",
        },
        {
            "sheet_name": "Mortality Log",
            "report_title": "Animal Mortality Log",
            "headers": ["Date", "Group", "Count", "Cause", "Note"],
            "rows": [[r["date"], r["group"], r["count"], r["cause"], r["note"]] for r in data["mortality"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Rows", len(data["mortality"]))],
            "column_formats": {"Count": "int"},
            "tab_color": "BE123C",
        },
    ])
    buf = workbook_to_buffer(wb)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=animals_report_{date.today()}.xlsx"})


# ── UI ─────────────────────────────────────────────────
@router.get("/", response_class=HTMLResponse)
def reports_ui(current_user: User = Depends(require_permission("page_reports"))):
    return """<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reports — AZed Farm</title>
<script src="/static/theme-init.js"></script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<script src="/static/theme.js"></script>
<style>
:root{
    --bg:#060810;--card:#0f1424;--card2:#151c30;
    --border:rgba(255,255,255,0.06);--border2:rgba(255,255,255,0.11);
    --green:#00ff9d;--blue:#4d9fff;--orange:#fb923c;--teal:#2dd4bf;
    --danger:#ff4d6d;--warn:#ffb547;--lime:#84cc16;--purple:#a855f7;
    --text:#f0f4ff;--sub:#8899bb;--muted:#445066;
    --sans:'Outfit',sans-serif;--mono:'JetBrains Mono',monospace;--r:12px;
}
body.light{
    --bg:#f4f5ef;--surface:#f1f3eb;--card:#eceee6;--card2:#e4e6de;
    --border:rgba(0,0,0,0.08);--border2:rgba(0,0,0,0.14);
    --green:#0f8a43;
    --text:#1a1e14;--sub:#4a5040;--muted:#7b816f;
}
body.light nav{background:rgba(244,245,239,.92);}
body.light .nav-link:hover{background:rgba(0,0,0,.05);}
body.light tr:hover td{background:rgba(0,0,0,.03);}
.topbar-right{display:flex;align-items:center;gap:12px;}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0;}
body{font-family:var(--sans);background:var(--bg);color:var(--text);min-height:100vh;font-size:14px;}
nav{position:sticky;top:0;z-index:100;display:flex;align-items:center;gap:8px;padding:0 24px;height:58px;background:rgba(10,13,24,.92);backdrop-filter:blur(20px);border-bottom:1px solid var(--border);}
.logo{font-size:17px;font-weight:900;background:linear-gradient(135deg,var(--green),var(--blue));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;text-decoration:none;display:flex;align-items:center;gap:8px;margin-right:10px;}
.nav-link{padding:7px 12px;border-radius:8px;color:var(--sub);font-size:12px;font-weight:600;text-decoration:none;transition:all .2s;}
.nav-link:hover{background:rgba(255,255,255,.05);color:var(--text);}
.nav-link.active{background:rgba(132,204,22,.1);color:var(--lime);}
.nav-spacer{flex:1;}
.content{max-width:1300px;margin:0 auto;padding:28px 24px;display:flex;flex-direction:column;gap:18px;}
.page-title{font-size:24px;font-weight:800;letter-spacing:-.5px;}
.page-sub{color:var(--muted);font-size:13px;margin-top:3px;}
.tabs{display:flex;gap:4px;background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:4px;flex-wrap:wrap;}
.tab{padding:7px 13px;border-radius:9px;font-size:12px;font-weight:700;cursor:pointer;border:none;background:transparent;color:var(--muted);transition:all .2s;font-family:var(--sans);white-space:nowrap;}
.tab.active{background:var(--card2);color:var(--text);}
.section{display:none;flex-direction:column;gap:16px;}
.section.active{display:flex;}
/* FILTER BAR */
.filter-bar{display:flex;align-items:center;gap:8px;flex-wrap:wrap;background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:12px 16px;}
.filter-bar label{font-size:11px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--muted);white-space:nowrap;}
.filter-bar input[type=date],.filter-bar input[type=month],.filter-bar input[type=text],.filter-bar input[type=number],.filter-bar select{background:var(--card2);border:1px solid var(--border2);border-radius:8px;padding:7px 11px;color:var(--text);font-family:var(--sans);font-size:13px;outline:none;}
.filter-bar input[type=date]:focus,.filter-bar input[type=month]:focus,.filter-bar input[type=text]:focus,.filter-bar input[type=number]:focus,.filter-bar select:focus{border-color:var(--lime);}
.filter-sep{width:1px;height:24px;background:var(--border2);margin:0 4px;}
.btn{display:inline-flex;align-items:center;gap:6px;padding:8px 14px;border-radius:var(--r);font-family:var(--sans);font-size:12px;font-weight:700;cursor:pointer;border:none;transition:all .2s;white-space:nowrap;}
.btn-lime {background:linear-gradient(135deg,var(--lime),var(--green));color:#0a1a00;}
.btn-lime:hover{filter:brightness(1.1);}
.btn-excel{background:linear-gradient(135deg,#217346,#1e6b3f);color:white;}
.btn-excel:hover{filter:brightness(1.1);}
.btn-print{background:var(--card2);border:1px solid var(--border2);color:var(--sub);}
.btn-print:hover{border-color:var(--blue);color:var(--blue);}
/* STATS */
.stats-row{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;}
.stat-card{background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:14px 16px;position:relative;overflow:hidden;}
.stat-card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;}
.sc-green::before {background:linear-gradient(90deg,var(--green),transparent);}
.sc-blue::before  {background:linear-gradient(90deg,var(--blue),transparent);}
.sc-orange::before{background:linear-gradient(90deg,var(--orange),transparent);}
.sc-danger::before{background:linear-gradient(90deg,var(--danger),transparent);}
.sc-teal::before  {background:linear-gradient(90deg,var(--teal),transparent);}
.stat-label{font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:var(--muted);margin-bottom:6px;}
.stat-value{font-family:var(--mono);font-size:22px;font-weight:700;}
.sv-green {color:var(--green);}
.sv-blue  {color:var(--blue);}
.sv-orange{color:var(--orange);}
.sv-danger{color:var(--danger);}
.sv-teal  {color:var(--teal);}
/* TABLE */
.table-wrap{background:var(--card);border:1px solid var(--border);border-radius:var(--r);overflow:hidden;}
.table-title{padding:12px 16px;font-size:11px;font-weight:700;letter-spacing:.5px;text-transform:uppercase;border-bottom:1px solid var(--border);color:var(--muted);display:flex;justify-content:space-between;align-items:center;}
table{width:100%;border-collapse:collapse;}
thead{background:var(--card2);}
th{text-align:left;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--muted);padding:10px 14px;}
td{padding:10px 14px;border-top:1px solid var(--border);color:var(--sub);font-size:13px;}
tr:hover td{background:rgba(255,255,255,.02);}
td.name{color:var(--text);font-weight:600;}
td.mono{font-family:var(--mono);}
/* CHARTS */
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:14px;}
.chart-card{background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:16px;}
.chart-title{font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:var(--muted);margin-bottom:12px;}
.bar-row{display:flex;align-items:center;gap:10px;margin-bottom:8px;}
.bar-label{font-size:12px;color:var(--sub);width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex-shrink:0;}
.bar-track{flex:1;background:var(--card2);border-radius:4px;height:8px;overflow:hidden;}
.bar-fill{height:100%;border-radius:4px;transition:width .5s ease;}
.bar-val{font-family:var(--mono);font-size:11px;width:70px;text-align:right;flex-shrink:0;}
/* BADGES */
.badge{display:inline-flex;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:700;}
.badge-low{background:rgba(255,77,109,.1);color:var(--danger);}
.badge-ok {background:rgba(0,255,157,.1);color:var(--green);}
/* P&L */
.pl-section{background:var(--card);border:1px solid var(--border);border-radius:var(--r);overflow:hidden;margin-bottom:8px;}
.pl-header{background:var(--card2);padding:10px 16px;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--muted);}
.pl-row{display:flex;justify-content:space-between;padding:9px 16px;border-top:1px solid var(--border);font-size:13px;}
.pl-row:hover{background:rgba(255,255,255,.02);}
.pl-total{font-weight:700;font-size:14px;background:rgba(0,0,0,.15);}
.pl-net{font-size:16px;font-weight:800;padding:12px 16px;}
/* TOAST */
.toast{position:fixed;bottom:22px;left:50%;transform:translateX(-50%) translateY(16px);background:var(--card2);border:1px solid var(--border2);border-radius:var(--r);padding:12px 20px;font-size:13px;font-weight:600;color:var(--text);opacity:0;pointer-events:none;transition:opacity .25s,transform .25s;z-index:999;}
.toast.show{opacity:1;transform:translateX(-50%) translateY(0);}
::-webkit-scrollbar{width:4px;}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:4px;}
@media(max-width:700px){.two-col{grid-template-columns:1fr;}}

/* ══════════════ PRINT STYLES ══════════════ */
.print-header{display:none;}
@media print{
    nav,.filter-bar,.tabs,.no-print{display:none!important;}
    body{background:white;color:#111;font-family:Arial,sans-serif;padding:0;}
    .content{padding:10px;max-width:100%;}
    .section{display:flex!important;}
    .section:not(.active){display:none!important;}
    .print-header{display:flex;align-items:center;justify-content:space-between;padding-bottom:14px;margin-bottom:20px;border-bottom:3px solid #2a7a2a;}
    .stat-card{border:1px solid #ddd!important;background:white!important;}
    .stat-label{color:#555!important;}
    .stat-value{color:#111!important;}
    .table-wrap{border:1px solid #ddd!important;background:white!important;}
    table thead{background:#2a7a2a!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    th{color:white!important;}
    td{color:#333!important;border-top:1px solid #eee!important;}
    td.name{color:#111!important;}
    .table-title{color:#555!important;background:white!important;}
    .chart-card{border:1px solid #ddd!important;background:white!important;}
    .chart-title{color:#555!important;}
    .bar-label,.bar-val{color:#333!important;}
    .bar-track{background:#eee!important;}
    .bar-fill{-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    .pl-section{border:1px solid #ddd!important;background:white!important;}
    .pl-header{background:#f0f0f0!important;color:#555!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    .pl-row{color:#333!important;}
    .pl-total{background:#f5f5f5!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    .badge-low{background:#fee!important;color:#c00!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    .badge-ok {background:#efe!important;color:#2a7a2a!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    .two-col{grid-template-columns:1fr 1fr;}
}
</style>
    <script src="/static/auth-guard.js"></script>
</head>
<body>
""" + render_app_header(current_user, "page_reports") + """

<div class="content">
    <div class="no-print">
        <div class="page-title">📊 Reports</div>
        <div class="page-sub">Filter by date period · Export to Excel · Print with logo</div>
    </div>

    <div class="tabs no-print">
        <button class="tab active" onclick="switchTab('sales')">📈 Sales</button>
        <button class="tab"        onclick="switchTab('transactions')">🧾 Transactions</button>
        <button class="tab"        onclick="switchTab('b2b')">🤝 B2B Statement</button>
        <button class="tab"        onclick="switchTab('inventory')">📦 Inventory</button>
        <button class="tab"        onclick="switchTab('farm')">🌾 Farm Intake</button>
        <button class="tab"        onclick="switchTab('spoilage')">🗑 Spoilage</button>
        <button class="tab"        onclick="switchTab('production')">⚙️ Production</button>
        <button class="tab"        onclick="switchTab('hr')">👥 HR</button>
        <button class="tab"        onclick="switchTab('utilities')">💧 Utilities</button>
        <button class="tab"        onclick="switchTab('pl')">💰 P&amp;L</button>
        <button class="tab"        onclick="switchTab('animals')">🐄 Animals</button>
    </div>

    <!-- ──────────── SALES ──────────── -->
    <div id="section-sales" class="section active">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Sales Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-sales-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="sales-from">
            <label>To</label>  <input type="date" id="sales-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadSales()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('sales')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-green"><div class="stat-label">Net Revenue</div><div class="stat-value sv-green" id="s-total">—</div></div>
            <div class="stat-card sc-blue" ><div class="stat-label">POS Revenue</div> <div class="stat-value sv-blue"  id="s-pos">—</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">B2B Revenue</div><div class="stat-value sv-orange" id="s-b2b">—</div></div>
            <div class="stat-card" style="border-color:rgba(255,77,109,.3);background:rgba(255,77,109,.04);position:relative;overflow:hidden;">
                <div style="position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,#ff4d6d,transparent)"></div>
                <div class="stat-label" style="color:#ff4d6d">↩ Refunds</div>
                <div class="stat-value" style="color:#ff4d6d;font-family:var(--mono)" id="s-refunds">—</div>
                <div style="font-size:11px;color:rgba(255,77,109,.6);margin-top:4px" id="s-refund-count">— refunds</div>
            </div>
        </div>
        <div class="two-col">
            <div class="table-wrap">
                <div class="table-title">Daily Breakdown</div>
                <table><thead><tr><th>Date</th><th>POS</th><th>B2B</th><th style="color:#ff4d6d">Refunds</th><th>Net Total</th></tr></thead>
                <tbody id="sales-daily"></tbody></table>
            </div>
            <div class="chart-card">
                <div class="chart-title">Top Products by Revenue</div>
                <div id="sales-top"></div>
            </div>
        </div>
        <div id="sales-records"></div>
    </div>

    <!-- ──────────── B2B ──────────── -->
    <!-- ── TRANSACTIONS ── -->
    <div id="section-transactions" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:20px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:13px;color:#555;margin-top:2px">Transactions Report</div>
                    <div style="font-size:12px;color:#555" id="ph-tx-dates"></div>
                </div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="tx-from">
            <label>To</label><input type="date" id="tx-to">
            <select id="tx-source">
                <option value="">All Sources</option>
                <option value="pos">POS Only</option>
                <option value="b2b">B2B Only</option>
                <option value="refund">Refunds Only</option>
                <option value="receive">Receive Only</option>
                <option value="production">Production / Processing Only</option>
                <option value="expense">Expenses Only</option>
            </select>
            <button class="btn btn-lime" onclick="loadTransactions()">Apply</button>
            <div class="filter-sep"></div>
            <button class="btn btn-excel no-print" onclick="exportSection('transactions')">Export Excel</button>
            <button class="btn-print no-print" onclick="printSection()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card lime" ><div class="stat-label">Total Lines</div><div class="stat-value lime"   id="tx-count">—</div></div>
            <div class="stat-card green"><div class="stat-label">Total Revenue</div><div class="stat-value green" id="tx-revenue">—</div></div>
            <div class="stat-card blue" ><div class="stat-label">Total Qty Sold</div><div class="stat-value blue"  id="tx-qty">—</div></div>
            <div class="stat-card warn" ><div class="stat-label">Total Discount</div><div class="stat-value warn"  id="tx-discount">—</div></div>
        </div>
        <div class="table-wrap">
            <div class="table-title">All Transactions</div>
            <div style="overflow-x:auto">
            <table>
                <thead><tr>
                    <th>Date</th><th>Invoice #</th><th>Source</th><th>Customer</th><th>By</th>
                    <th>SKU</th><th>Product</th><th>Product Category</th><th>QTY</th><th>Unit Price</th>
                    <th>Line Total</th><th>Discount</th><th>Disc %</th>
                    <th>Payment</th><th>Inv. Total</th><th>Status</th>
                </tr></thead>
                <tbody id="tx-body"></tbody>
            </table>
            </div>
        </div>
    </div>

    <div id="section-b2b" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">B2B Client Statement</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-b2b-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="b2b-from">
            <label>To</label>  <input type="date" id="b2b-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadB2B()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('b2b')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-blue"  ><div class="stat-label">Clients</div>      <div class="stat-value sv-blue"   id="b-clients">—</div></div>
            <div class="stat-card sc-green" ><div class="stat-label">Total Invoiced</div><div class="stat-value sv-green"  id="b-invoiced">—</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Outstanding</div>   <div class="stat-value sv-danger" id="b-outstanding">—</div></div>
        </div>
        <div class="table-wrap">
            <div class="table-title">Client Statements</div>
            <table><thead><tr><th>Client</th><th>Phone</th><th>Terms</th><th>Invoiced</th><th>Paid</th><th>Outstanding</th><th>Invoices</th></tr></thead>
            <tbody id="b2b-body"></tbody></table>
        </div>
    </div>

    <!-- ──────────── INVENTORY ──────────── -->
    <div id="section-inventory" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Inventory Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-inv-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <span style="font-size:12px;color:var(--muted)">Current stock snapshot</span>
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadInventory()">Refresh</button>
            <button class="btn btn-excel" onclick="exportSection('inventory')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-blue"  ><div class="stat-label">Products</div>   <div class="stat-value sv-blue"   id="inv-count">—</div></div>
            <div class="stat-card sc-green" ><div class="stat-label">Stock Value</div><div class="stat-value sv-green"  id="inv-value">—</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Low Stock</div>   <div class="stat-value sv-danger" id="inv-low">—</div></div>
        </div>
        <div class="table-wrap">
            <div class="table-title">Stock Levels</div>
            <table><thead><tr><th>SKU</th><th>Product</th><th>Stock</th><th>Unit</th><th>Price</th><th>Stock Value</th><th>Status</th></tr></thead>
            <tbody id="inv-body"></tbody></table>
        </div>
    </div>

    <!-- ──────────── FARM ──────────── -->
    <div id="section-farm" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Farm Intake Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-farm-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="farm-from">
            <label>To</label>  <input type="date" id="farm-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadFarm()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('farm')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div id="farm-content"></div>
    </div>

    <!-- ──────────── SPOILAGE ──────────── -->
    <div id="section-spoilage" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Spoilage Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-spl-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="spl-from">
            <label>To</label>  <input type="date" id="spl-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadSpoilage()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('spoilage')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-danger"><div class="stat-label">Total Records</div><div class="stat-value sv-danger" id="spl-count">—</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Total Qty Lost</div><div class="stat-value sv-orange" id="spl-qty">—</div></div>
        </div>
        <div class="two-col">
            <div class="chart-card"><div class="chart-title">By Product</div><div id="spl-by-product"></div></div>
            <div class="chart-card"><div class="chart-title">By Reason</div> <div id="spl-by-reason"></div></div>
        </div>
        <div class="table-wrap">
            <div class="table-title">All Records</div>
            <table><thead><tr><th>Ref #</th><th>Product</th><th>Qty</th><th>Reason</th><th>Farm</th><th>Date</th><th>By</th><th>Notes</th></tr></thead>
            <tbody id="spl-body"></tbody></table>
        </div>
    </div>

    <!-- ──────────── PRODUCTION ──────────── -->
    <div id="section-production" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Production Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-prod-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="prod-from">
            <label>To</label>  <input type="date" id="prod-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadProduction()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('production')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-orange"><div class="stat-label">Processing Batches</div><div class="stat-value sv-orange" id="prod-proc">—</div></div>
            <div class="stat-card sc-teal"  ><div class="stat-label">Packaging Runs</div>   <div class="stat-value sv-teal"   id="prod-pkg">—</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Drying Batches</div>   <div class="stat-value sv-orange" id="prod-drying">—</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Avg Loss %</div>        <div class="stat-value sv-danger" id="prod-loss">—</div></div>
        </div>
        <div class="table-wrap">
            <div class="table-title">All Batches</div>
            <table><thead><tr><th>Batch #</th><th>Type</th><th>Recipe</th><th>Inputs</th><th>Outputs</th><th>Loss %</th><th>Date</th><th>By</th></tr></thead>
            <tbody id="prod-body"></tbody></table>
        </div>
    </div>

    <!-- ──────────── HR ──────────── -->
    <div id="section-hr" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">HR Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-hr-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="hr-from">
            <label>To</label><input type="date" id="hr-to">
            <label>Period</label><input type="month" id="hr-period">
            <label>Department</label><input type="text" id="hr-department" placeholder="All departments">
            <label>Farm ID</label><input type="number" id="hr-farm-id" min="1" placeholder="All farms">
            <div class="filter-sep"></div>
            <button class="btn btn-lime" onclick="loadHR()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('hr')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-blue"><div class="stat-label">Active Employees</div><div class="stat-value sv-blue" id="hr-active">—</div></div>
            <div class="stat-card sc-green"><div class="stat-label">Attendance Rate</div><div class="stat-value sv-green" id="hr-rate">—</div></div>
            <div class="stat-card sc-teal"><div class="stat-label">Present Days</div><div class="stat-value sv-teal" id="hr-present">—</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Absent Days</div><div class="stat-value sv-danger" id="hr-absent">—</div></div>
            <div class="stat-card sc-green"><div class="stat-label">Net Salary</div><div class="stat-value sv-green" id="hr-net">—</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Unpaid Salary</div><div class="stat-value sv-orange" id="hr-unpaid">—</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Outstanding Loans</div><div class="stat-value sv-orange" id="hr-loans">—</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Total Deductions</div><div class="stat-value sv-danger" id="hr-deductions">—</div></div>
            <div class="stat-card sc-green"><div class="stat-label">Paid in Cash</div><div class="stat-value sv-green" id="hr-paid-cash">—</div></div>
            <div class="stat-card sc-teal"><div class="stat-label">Days Off Credit</div><div class="stat-value sv-teal" id="hr-daysoff">—</div></div>
        </div>
        <div class="two-col">
            <div class="table-wrap">
                <div class="table-title">Department Summary</div>
                <table><thead><tr><th>Department</th><th>Employees</th><th>Present</th><th>Absent</th><th>Late</th><th>Leave</th><th>Net Salary</th></tr></thead>
                <tbody id="hr-dept-body"></tbody></table>
            </div>
            <div class="table-wrap">
                <div class="table-title">Farm Summary</div>
                <table><thead><tr><th>Farm</th><th>Employees</th><th>Present</th><th>Absent</th><th>Late</th><th>Leave</th><th>Net Salary</th></tr></thead>
                <tbody id="hr-farm-body"></tbody></table>
            </div>
        </div>
        <div class="table-wrap">
            <div class="table-title">Employee Detail</div>
            <div style="overflow-x:auto">
            <table><thead><tr><th>Employee</th><th>Phone</th><th>Position</th><th>Base Salary</th><th>Attendance</th><th>Worked</th><th>Loan Bal.</th><th>Loan Ded.</th><th>Day Ded.</th><th>Manual Ded.</th><th>Net Salary</th><th>Paid Cash</th><th>Days Off Bal.</th><th>Paid</th></tr></thead>
            <tbody id="hr-emp-body"></tbody></table>
            </div>
        </div>
    </div>

    <!-- ──────────── UTILITIES ──────────── -->
    <div id="section-utilities" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Utilities Consumption Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-util-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="util-from">
            <label>To</label>  <input type="date" id="util-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadUtilities()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('utilities')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-blue"><div class="stat-label">Total Spend</div>      <div class="stat-value sv-blue"   id="util-total-cost">—</div></div>
            <div class="stat-card sc-green"><div class="stat-label">Utilities Tracked</div><div class="stat-value sv-green"  id="util-total-cats">—</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Carbon Footprint</div><div class="stat-value sv-orange" id="util-total-carbon">—</div></div>
        </div>
        <div id="util-warning" style="display:none;font-size:13px;color:var(--warn);margin-bottom:12px"></div>
        <div class="table-wrap">
            <div class="table-title">Utility Summary</div>
            <div style="overflow-x:auto">
            <table><thead><tr>
                <th>Utility</th><th>Unit</th><th>Entries</th><th>Consumption</th>
                <th>Cost (EGP)</th><th>Cost / Unit</th><th>Default Unit Price</th><th>kg CO₂e</th>
            </tr></thead>
            <tbody id="util-summary-body"></tbody></table>
            </div>
        </div>
        <div class="table-wrap">
            <div class="table-title">Monthly Trend — Last 12 Months (Cost in EGP)</div>
            <div style="overflow-x:auto">
            <table id="util-trend-table"><thead id="util-trend-head"></thead>
            <tbody id="util-trend-body"></tbody></table>
            </div>
        </div>
        <div class="table-wrap">
            <div class="table-title">Per-Farm Breakdown</div>
            <div style="overflow-x:auto">
            <table><thead><tr>
                <th>Utility</th><th>Farm</th><th>Consumption</th><th>Unit</th><th>Cost (EGP)</th>
            </tr></thead>
            <tbody id="util-farm-body"></tbody></table>
            </div>
        </div>
    </div>

    <!-- ──────────── P&L ──────────── -->
    <div id="section-animals" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Animals Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-animals-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="animals-from">
            <label>To</label>  <input type="date" id="animals-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadAnimals()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('animals')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div id="animals-content"></div>
    </div>

    <div id="section-pl" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Profit &amp; Loss Statement</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-pl-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="pl-from">
            <label>To</label>  <input type="date" id="pl-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadPL()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('pl')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div id="pl-content"></div>
    </div>

</div><!-- end .content -->
<div class="toast" id="toast"></div>

<script>
  let __currentUser = null;

  // Theme is handled by window.__appTheme (theme.js / theme-init.js).
  // toggleMode is kept as a named alias so window.__appNav.toggleTheme() can find it.
  function toggleMode(){
    if(window.__appTheme) window.__appTheme.toggle();
  }

async function initUser() {
    try {
        const r = await fetch("/auth/me");
        if (!r.ok) { _redirectToLogin(); return; }
        const u = await r.json();
        __currentUser = u;
        return u;
    } catch(e) { _redirectToLogin(); }
}

  function hasPermission(permission){
      const role = __currentUser ? (__currentUser.role || "") : "";
      let permsArray = [];
      if (__currentUser && __currentUser.permissions) {
          permsArray = typeof __currentUser.permissions === "string" 
              ? __currentUser.permissions.split(",").map(v => String(v).trim()).filter(Boolean) 
              : __currentUser.permissions;
      }
      const perms = new Set(permsArray);
      return role === "admin" || perms.has(permission);
  }
  function configureReportsPermissions(){
      ensureTabMetadata();
      let firstAllowed = null;
      document.querySelectorAll(".tabs .tab").forEach((btn) => {
          const tab = btn.dataset.tab;
          if(!tab) return;
          if(!isTabAllowed(tab)){
              btn.style.display = "none";
          } else if(!firstAllowed) {
              firstAllowed = tab;
          }
      });
      if(!isTabAllowed(currentTab) && firstAllowed){
          currentTab = firstAllowed;
      }
      if(!hasPermission("action_export_excel")){
          document.querySelectorAll(".btn-excel").forEach(btn => btn.style.display = "none");
      }
  }
  let currentTab = "sales";
let toastTimer = null;
initUser().then(u => {
    if(!u) return;
    configureReportsPermissions();
    switchTab(currentTab);
});

function switchTab(tab){
    ensureTabMetadata();
    if(!isTabAllowed(tab)){
        const fallback = REPORT_TAB_ORDER.find(isTabAllowed);
        if(!fallback){
            showToast("No report tabs available for this account.");
            return;
        }
        tab = fallback;
    }
    currentTab = tab;
    document.querySelectorAll(".tab").forEach(btn => btn.classList.toggle("active", btn.dataset.tab===tab));
    document.querySelectorAll(".section").forEach(s => s.classList.remove("active"));
    const section = document.getElementById("section-"+tab);
    if(!section) return;
    section.classList.add("active");
    const loaders = {sales:loadSales, transactions:loadTransactions, b2b:loadB2B, inventory:loadInventory, farm:loadFarm, spoilage:loadSpoilage, production:loadProduction, hr:loadHR, utilities:loadUtilities, pl:loadPL, animals:loadAnimals};
    if(loaders[tab]){
        loaders[tab]();
    } else {
        setSectionStatus(tab, "error", "This report tab is not wired correctly.");
    }
}

function today(){ return new Date().toISOString().split("T")[0]; }
function monthStart(){ let d=new Date(); d.setDate(1); return d.toISOString().split("T")[0]; }
function yearStart() { let d=new Date(); d.setMonth(0); d.setDate(1); return d.toISOString().split("T")[0]; }
function getRange(f, t){ return {from: document.getElementById(f).value, to: document.getElementById(t).value}; }
function setEl(id, v){ document.getElementById(id).value = v; }
function setPrintDates(id, from, to){ let el=document.getElementById(id); if(el) el.innerText=`Period: ${from}  →  ${to}`; }

function showToast(msg){
    let t=document.getElementById("toast"); t.innerText=msg; t.classList.add("show");
    clearTimeout(toastTimer); toastTimer=setTimeout(()=>t.classList.remove("show"),3000);
}

const REPORT_TAB_ORDER = ["sales","transactions","b2b","inventory","farm","spoilage","production","hr","utilities","pl","animals"];
const REPORT_TAB_PERMISSIONS = {
    sales: "tab_reports_sales",
    transactions: "tab_reports_transactions",
    b2b: "tab_reports_b2b",
    inventory: "tab_reports_inventory",
    farm: "tab_reports_farm",
    spoilage: "tab_reports_spoilage",
    production: "tab_reports_production",
    hr: "tab_reports_hr",
    utilities: "tab_reports_utilities",
    pl: "tab_reports_pl",
    animals: "tab_reports_animals",
};

function ensureTabMetadata(){
    document.querySelectorAll(".tabs .tab").forEach((btn, index) => {
        if(!btn.dataset.tab && REPORT_TAB_ORDER[index]){
            btn.dataset.tab = REPORT_TAB_ORDER[index];
        }
    });
}

function getTabPermission(tab){
    return REPORT_TAB_PERMISSIONS[tab] || null;
}

function isTabAllowed(tab){
    const permission = getTabPermission(tab);
    return !permission || hasPermission(permission);
}

function getSectionElement(tab){
    return document.getElementById(`section-${tab}`);
}

function ensureSectionStatus(tab){
    const section = getSectionElement(tab);
    if(!section) return null;
    let status = section.querySelector(".report-status");
    if(!status){
        status = document.createElement("div");
        status.className = "report-status";
        status.style.cssText = "display:none;margin:0 0 14px;padding:12px 14px;border-radius:10px;border:1px solid var(--border2);font-size:13px;line-height:1.5;";
        const firstChild = section.firstElementChild;
        if(firstChild && firstChild.classList.contains("print-header")){
            firstChild.insertAdjacentElement("afterend", status);
        } else {
            section.prepend(status);
        }
    }
    return status;
}

function setSectionStatus(tab, kind, message){
    const status = ensureSectionStatus(tab);
    if(!status) return;
    if(!message){
        status.style.display = "none";
        status.textContent = "";
        return;
    }
    const palettes = {
        info: "background:rgba(77,159,255,.08);border-color:rgba(77,159,255,.25);color:var(--blue);",
        error: "background:rgba(255,77,109,.08);border-color:rgba(255,77,109,.25);color:var(--danger);",
        empty: "background:rgba(255,181,71,.08);border-color:rgba(255,181,71,.25);color:var(--warn);",
    };
    status.style.cssText = `display:block;margin:0 0 14px;padding:12px 14px;border-radius:10px;border:1px solid var(--border2);font-size:13px;line-height:1.5;${palettes[kind] || palettes.info}`;
    status.textContent = message;
}

// ── XSS protection ──────────────────────────────────────────────
// Report rows are rendered into innerHTML from DB-derived strings
// (product/group/employee/customer names, notes, sources, reasons…).
// Escape every string value once, here at the data-ingress choke point,
// so no individual render site can forget to. Numbers, booleans, dates
// and structure pass through untouched; Excel exports are built server-
// side from raw data and are unaffected.
const _HTML_ESCAPES = {"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#39;"};
function escapeHtml(value){
    if(value == null) return "";
    return String(value).replace(/[&<>"']/g, c => _HTML_ESCAPES[c]);
}
function deepEscapeStrings(value){
    if(typeof value === "string") return escapeHtml(value);
    if(Array.isArray(value)) return value.map(deepEscapeStrings);
    if(value && typeof value === "object"){
        const out = {};
        for(const key in value){
            if(Object.prototype.hasOwnProperty.call(value, key)){
                out[key] = deepEscapeStrings(value[key]);
            }
        }
        return out;
    }
    return value;
}

async function fetchReportJson(url){
    const response = await fetch(url, { credentials: "same-origin" });
    const contentType = response.headers.get("content-type") || "";
    let payload = null;
    if(contentType.includes("application/json")){
        payload = await response.json().catch(() => null);
    } else {
        const text = await response.text().catch(() => "");
        if(!response.ok){
            throw new Error(text || `Request failed (${response.status})`);
        }
        throw new Error("Unexpected non-JSON response from reports endpoint.");
    }
    if(!response.ok){
        const detail = payload && (payload.detail || payload.message || payload.error);
        throw new Error(detail || `Request failed (${response.status})`);
    }
    return deepEscapeStrings(payload);
}

async function runReportLoader(tab, loader){
    if(!isTabAllowed(tab)){
        setSectionStatus(tab, "error", "You do not have permission to view this report.");
        return;
    }
    setSectionStatus(tab, "info", "Loading report...");
    try{
        await loader();
        setSectionStatus(tab, "", "");
    } catch(error){
        console.error(`Report load failed for ${tab}:`, error);
        setSectionStatus(tab, "error", error && error.message ? error.message : "Could not load this report.");
        showToast(`Could not load ${tab} report`);
    }
}

function getDownloadFilename(response, fallback){
    const disposition = response.headers.get("Content-Disposition") || "";
    const utf8Match = disposition.match(/filename\\*=UTF-8''([^;]+)/i);
    if(utf8Match) return decodeURIComponent(utf8Match[1]);
    const plainMatch = disposition.match(/filename="?([^\";]+)"?/i);
    return plainMatch ? plainMatch[1] : fallback;
}

async function loadAnimals(){
    let r = getRange("animals-from","animals-to");
    let data = await fetchReportJson(`/reports/api/animals?date_from=${r.from}&date_to=${r.to}`);
    setPrintDates("ph-animals-dates", r.from, r.to);
    const s = data.summary;
    const money = v => Number(v||0).toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2});
    const groupRows = data.groups.length
        ? data.groups.map(row=>`<tr>
            <td class="name">${row.name}</td>
            <td>${row.animal_type||""}</td>
            <td>${row.farm_name||"—"}</td>
            <td class="mono">${row.headcount}</td>
            <td class="mono">${(row.male_count!=null||row.female_count!=null)?`${row.male_count!=null?row.male_count:"–"} / ${row.female_count!=null?row.female_count:"–"}`:"—"}</td>
            <td>${row.age||"—"}</td>
            <td class="mono" style="color:var(--green)">${row.received}</td>
            <td class="mono" style="color:var(--green)">${row.born}</td>
            <td class="mono">${row.transferred}</td>
            <td class="mono" style="color:var(--orange)">${row.died}</td>
            <td class="mono">${row.net_change}</td>
            <td class="mono" style="color:var(--orange)">${money(row.purchase_cost)}</td>
            <td class="mono" style="color:var(--orange)">${money(row.feeding_cost)}</td>
            <td class="mono" style="color:var(--orange)">${money(row.total_cost)}</td>
            <td class="mono">${money(row.cost_per_head)}</td>
          </tr>`).join("")
        : `<tr><td colspan="15" style="text-align:center;color:var(--muted);padding:24px">No animal groups or activity in this period</td></tr>`;
    const intakeRows = data.intakes.length
        ? data.intakes.map(row=>`<tr>
            <td class="mono">${row.date||""}</td>
            <td class="name">${row.group||""}</td>
            <td>${row.type||""}</td>
            <td class="mono">${row.count}</td>
            <td class="mono">${(row.male_count!=null||row.female_count!=null)?`${row.male_count!=null?row.male_count:"–"} / ${row.female_count!=null?row.female_count:"–"}`:"—"}</td>
            <td>${row.source||"—"}</td>
            <td class="mono">${row.cost!=null?money(row.cost):"—"}</td>
          </tr>`).join("")
        : `<tr><td colspan="7" style="text-align:center;color:var(--muted);padding:24px">No intakes in this period</td></tr>`;
    const mortRows = data.mortality.length
        ? data.mortality.map(row=>`<tr>
            <td class="mono">${row.date||""}</td>
            <td class="name">${row.group||""}</td>
            <td class="mono">${row.count}</td>
            <td>${row.cause||""}</td>
            <td>${row.note||"—"}</td>
          </tr>`).join("")
        : `<tr><td colspan="5" style="text-align:center;color:var(--muted);padding:24px">No mortality in this period</td></tr>`;
    document.getElementById("animals-content").innerHTML = `
        <div class="stats-row">
            <div class="stat-card sc-blue"><div class="stat-label">Active Groups</div><div class="stat-value sv-blue">${s.active_groups}</div></div>
            <div class="stat-card sc-green"><div class="stat-label">Total Headcount</div><div class="stat-value sv-green">${s.total_headcount}</div></div>
            <div class="stat-card sc-green"><div class="stat-label">Received + Born + In</div><div class="stat-value sv-green">${s.received+s.born+s.transferred}</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Died</div><div class="stat-value sv-orange">${s.died}</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Total Cost</div><div class="stat-value sv-orange">${money(s.total_cost)}</div></div>
        </div>
        <div class="table-wrap" style="margin-bottom:18px">
            <div class="table-title">By Group</div>
            <table><thead><tr><th>Group</th><th>Type</th><th>Farm</th><th>Headcount</th><th>Sex (M/F)</th><th>Age</th><th>Received</th><th>Born</th><th>Transferred</th><th>Died</th><th>Net</th><th>Purchase Cost</th><th>Feeding Cost</th><th>Total Cost</th><th>Cost/Head</th></tr></thead><tbody>${groupRows}</tbody></table>
        </div>
        <div class="table-wrap" style="margin-bottom:18px">
            <div class="table-title">Intake (Receive) Log</div>
            <table><thead><tr><th>Date</th><th>Group</th><th>Type</th><th>Count</th><th>Sex (M/F)</th><th>Source</th><th>Cost</th></tr></thead><tbody>${intakeRows}</tbody></table>
        </div>
        <div class="table-wrap">
            <div class="table-title">Mortality Log</div>
            <table><thead><tr><th>Date</th><th>Group</th><th>Count</th><th>Cause</th><th>Note</th></tr></thead><tbody>${mortRows}</tbody></table>
        </div>`;
}

async function exportSection(tab){
    const build = {
        sales:      ()=>{ let r=getRange("sales-from","sales-to"); return `/reports/export/sales?date_from=${r.from}&date_to=${r.to}`; },
        b2b:        ()=>{ let r=getRange("b2b-from","b2b-to");     return `/reports/export/b2b-statement?date_from=${r.from}&date_to=${r.to}`; },
        inventory:  ()=>{ let mode=document.getElementById("inv-mode")?.value || "snapshot"; let from=document.getElementById("inv-from")?.value || ""; let to=document.getElementById("inv-to")?.value || ""; return `/reports/export/inventory?mode=${mode}${mode==="movement"?`&date_from=${from}&date_to=${to}`:""}`; },
        farm:       ()=>{ let r=getRange("farm-from","farm-to");   return `/reports/export/farm-intake?date_from=${r.from}&date_to=${r.to}`; },
        spoilage:   ()=>{ let r=getRange("spl-from","spl-to");     return `/reports/export/spoilage?date_from=${r.from}&date_to=${r.to}`; },
        production: ()=>{ let r=getRange("prod-from","prod-to");   return `/reports/export/production?date_from=${r.from}&date_to=${r.to}`; },
        hr:         ()=>{ let r=getRange("hr-from","hr-to"); let p=document.getElementById("hr-period").value; let d=document.getElementById("hr-department").value.trim(); let f=document.getElementById("hr-farm-id").value; return `/reports/export/hr?date_from=${r.from}&date_to=${r.to}${p?"&period="+encodeURIComponent(p):""}${d?"&department="+encodeURIComponent(d):""}${f?"&farm_id="+encodeURIComponent(f):""}`; },
        utilities:  ()=>{ let r=getRange("util-from","util-to"); return `/reports/export/utilities?date_from=${r.from}&date_to=${r.to}`; },
        pl:           ()=>{ let r=getRange("pl-from","pl-to");   return `/reports/export/pl?date_from=${r.from}&date_to=${r.to}`; },
        animals:      ()=>{ let r=getRange("animals-from","animals-to"); return `/reports/export/animals?date_from=${r.from}&date_to=${r.to}`; },
        transactions: ()=>{ let r=getRange("tx-from","tx-to"); let s=document.getElementById("tx-source").value; return `/reports/export/transactions?date_from=${r.from}&date_to=${r.to}${s?"&source="+s:""}`; },
    };
    const fallbackFilename = `${tab}_report.xlsx`;
    showToast("Preparing Excel...");
    try{
        const response = await fetch(build[tab](), { credentials: "same-origin" });
        if(!response.ok){
            const contentType = response.headers.get("content-type") || "";
            let message = `Excel export failed (${response.status}).`;
            if(contentType.includes("application/json")){
                const payload = await response.json().catch(()=>null);
                if(payload && payload.detail) message = payload.detail;
            } else if(response.status >= 500){
                message = "Excel export failed on the server. Please try again.";
            } else {
                const text = await response.text().catch(()=> "");
                if(text && !text.trim().startsWith("<!DOCTYPE") && !text.trim().startsWith("<html")){
                    message = text.trim().slice(0, 160);
                }
            }
            showToast(message);
            return;
        }
        const blob = await response.blob();
        const filename = getDownloadFilename(response, fallbackFilename);
        const blobUrl = URL.createObjectURL(blob);
        const link = document.createElement("a");
        link.href = blobUrl;
        link.download = filename;
        document.body.appendChild(link);
        link.click();
        link.remove();
        setTimeout(() => URL.revokeObjectURL(blobUrl), 1000);
        showToast("Excel downloaded");
    } catch(e){
        console.error("Export failed:", e);
        showToast("Excel export failed. Please check your connection and try again.");
    }
}

/* ── TRANSACTIONS ── */
async function loadTransactions(){
    let r      = getRange("tx-from","tx-to");
    let source = document.getElementById("tx-source").value;
    let url    = `/reports/api/transactions?date_from=${r.from}&date_to=${r.to}${source?"&source="+source:""}`;
    let data   = await fetchReportJson(url);
    const statsRow = document.querySelector("#section-transactions .stats-row");
    if (statsRow) {
        statsRow.innerHTML = `
            <div class="stat-card lime"><div class="stat-label">Rows</div><div class="stat-value lime" id="tx-count">${data.total_rows}</div></div>
            <div class="stat-card green"><div class="stat-label">Money In</div><div class="stat-value green" id="tx-money-in">${data.money_in.toFixed(2)}</div></div>
            <div class="stat-card warn"><div class="stat-label">Money Out</div><div class="stat-value warn" id="tx-money-out">${data.money_out.toFixed(2)}</div></div>
            <div class="stat-card blue"><div class="stat-label">Stock In</div><div class="stat-value blue" id="tx-stock-in">${data.stock_in.toFixed(2)}</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Stock Out</div><div class="stat-value sv-danger" id="tx-stock-out">${data.stock_out.toFixed(2)}</div></div>
        `;
    }
    const txTableHead = document.querySelector("#tx-body")?.closest("table")?.querySelector("thead tr");
    if (txTableHead) {
        txTableHead.innerHTML = "<th>Date</th><th>Reference</th><th>Type</th><th>Source</th><th>Counterparty Type</th><th>Counterparty</th><th>By</th><th>SKU</th><th>Product</th><th>Product Category</th><th>Qty</th><th>Unit Price</th><th>Money Effect</th><th>Stock Effect</th><th>Direction</th><th>Payment</th><th>Status</th>";
    }
    setPrintDates("ph-tx-dates", r.from, r.to);
    document.getElementById("tx-body").innerHTML = data.rows.length
        ? data.rows.map(r => `<tr>
            <td class="mono" style="font-size:11px;white-space:nowrap">${r.date}</td>
            <td class="mono" style="font-size:11px;color:var(--blue)">${r.reference}</td>
            <td>${r.transaction_type}</td>
            <td>${r.source}</td>
            <td>${r.counterparty_type}</td>
            <td class="name">${r.counterparty_name}</td>
            <td style="font-size:12px;color:var(--muted)">${r.user_name}</td>
            <td class="mono" style="font-size:11px;color:var(--muted)">${r.sku}</td>
            <td>${r.product}${r.notes?`<br><span style="font-size:10px;color:var(--muted)">${r.notes}</span>`:""}</td>
            <td>${r.product_category || "—"}</td>
            <td class="mono">${r.qty.toFixed(2)}</td>
            <td class="mono">${r.unit_price.toFixed(2)}</td>
            <td class="mono" style="color:${r.money_effect>=0?"var(--green)":"var(--danger)"};font-weight:700">${r.money_effect.toFixed(2)}</td>
            <td class="mono" style="color:${r.stock_effect>=0?"var(--green)":"var(--danger)"};font-weight:700">${r.stock_effect.toFixed(2)}</td>
            <td>${r.direction}</td>
            <td>${r.payment_method}</td>
            <td>${r.status}</td>
        </tr>`).join("")
        : `<tr><td colspan="17" style="text-align:center;color:var(--muted);padding:40px">No transactions in this period</td></tr>`;
    return;

    document.getElementById("tx-count").innerText   = data.total_rows;
    document.getElementById("tx-revenue").innerText = data.total_revenue.toFixed(2);
    document.getElementById("tx-qty").innerText     = data.total_qty.toFixed(2);
    document.getElementById("tx-discount").innerText= data.total_discount.toFixed(2);
    setPrintDates("ph-tx-dates", r.from, r.to);

    const payColor = (m) => {
        if(!m) return "var(--muted)";
        m = m.toLowerCase();
        if(m.includes("visa") || m.includes("card")) return "var(--blue)";
        if(m.includes("cash"))                        return "var(--green)";
        if(m.includes("consign"))                     return "var(--teal)";
        if(m.includes("transfer"))                    return "var(--purple)";
        if(m.includes("credit") || m.includes("exchange") || m.includes("refund")) return "var(--danger)";
        return "var(--sub)";
    };
    const statusColor = (s) => {
        if(s==="paid")      return "var(--green)";
        if(s==="unpaid")    return "var(--warn)";
        if(s==="partial")   return "var(--blue)";
        if(s==="consignment") return "var(--teal)";
        if(s==="refunded")  return "var(--danger)";
        if(s==="received")  return "var(--warn)";
        if(s==="posted")    return "var(--purple)";
        return "var(--muted)";
    };

    document.getElementById("tx-body").innerHTML = data.rows.length
        ? data.rows.map(r => {
            const isRef = r.row_type === "refund";
            const isReceipt = r.row_type === "receipt";
            const isExpense = r.row_type === "expense";
            const isOutflow = isRef || isReceipt || isExpense;
            const rowStyle = isRef
                ? 'style="background:rgba(255,77,109,.04);"'
                : isReceipt
                    ? 'style="background:rgba(255,181,71,.06);"'
                    : isExpense
                        ? 'style="background:rgba(168,85,247,.06);"'
                            : '';
            const numColor = isRef ? "var(--danger)" : isReceipt ? "var(--warn)" : isExpense ? "var(--purple)" : "var(--green)";
            const refBadge = isRef
                ? `<br><span style="font-size:9px;font-weight:800;letter-spacing:.5px;color:var(--danger);background:rgba(255,77,109,.15);padding:1px 5px;border-radius:4px">↩ REFUND</span>`
                : isReceipt
                    ? `<br><span style="font-size:9px;font-weight:800;letter-spacing:.5px;color:var(--warn);background:rgba(255,181,71,.16);padding:1px 5px;border-radius:4px">↑ RECEIVE</span>`
                    : isExpense
                        ? `<br><span style="font-size:9px;font-weight:800;letter-spacing:.5px;color:var(--purple);background:rgba(168,85,247,.16);padding:1px 5px;border-radius:4px">↓ EXPENSE</span>`
                            : "";
            return `<tr ${rowStyle}>
                <td class="mono" style="font-size:11px;white-space:nowrap">${r.date}</td>
                <td class="mono" style="font-size:11px;color:${isRef?"var(--danger)":isReceipt?"var(--warn)":isExpense?"var(--purple)":"var(--lime)"}">${r.invoice_number}${refBadge}</td>
                <td style="font-size:11px;color:${isRef?"var(--danger)":isReceipt?"var(--warn)":isExpense?"var(--purple)":"var(--sub)"}">${r.source}</td>
                <td class="name" style="white-space:nowrap">${r.customer}</td>
                <td style="font-size:12px;color:var(--muted);white-space:nowrap">${r.user_name}</td>
                <td class="mono" style="font-size:11px;color:var(--muted)">${r.sku}</td>
                <td style="font-weight:600;white-space:nowrap">${r.product}${isOutflow&&r.reason?`<br><span style="font-size:10px;color:var(--muted);font-weight:400">${r.reason}</span>`:""}</td>
                <td class="mono" style="color:${isRef?"var(--danger)":isReceipt?"var(--amber)":isExpense?"var(--purple)":"var(--blue)"};font-weight:700">${r.qty.toFixed(2)}</td>
                <td class="mono">${r.unit_price.toFixed(2)}</td>
                <td class="mono" style="color:${numColor};font-weight:700">${isOutflow?"−":""}${Math.abs(r.line_total).toFixed(2)}</td>
                <td class="mono" style="color:${r.discount>0?"var(--warn)":"var(--muted)"}">${r.discount>0?"-"+r.discount.toFixed(2):"—"}</td>
                <td class="mono" style="color:${r.discount_pct>0?"var(--warn)":"var(--muted)"}">${r.discount_pct>0?r.discount_pct.toFixed(1)+"%":"—"}</td>
                <td style="font-size:12px;font-weight:700;color:${payColor(r.payment_method)}">${r.payment_method}</td>
                <td class="mono" style="font-weight:700;color:${isOutflow?"var(--danger)":"inherit"}">${isOutflow?"−":""}${Math.abs(r.invoice_total).toFixed(2)}</td>
                <td><span style="font-size:11px;font-weight:700;padding:2px 8px;border-radius:20px;background:rgba(0,0,0,.2);color:${statusColor(r.status)}">${r.status}</span></td>
              </tr>`;
        }).join("")
        : `<tr><td colspan="15" style="text-align:center;color:var(--muted);padding:40px">No transactions in this period</td></tr>`;
}

/* ── SALES ── */
async function loadSales(){
    let r = getRange("sales-from","sales-to");
    let data = await fetchReportJson(`/reports/api/sales?date_from=${r.from}&date_to=${r.to}`);
    const statsRow = document.querySelector("#section-sales .stats-row");
    if (statsRow) {
        statsRow.innerHTML = `
            <div class="stat-card sc-blue"><div class="stat-label">Gross Sales</div><div class="stat-value sv-blue" id="s-gross">${data.gross_sales.toFixed(2)}</div></div>
            <div class="stat-card" style="border-color:rgba(255,77,109,.3);background:rgba(255,77,109,.04);position:relative;overflow:hidden;">
                <div style="position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,#ff4d6d,transparent)"></div>
                <div class="stat-label" style="color:#ff4d6d">Refunds</div>
                <div class="stat-value" style="color:#ff4d6d;font-family:var(--mono)" id="s-refunds">${data.refunds > 0 ? "−" + data.refunds.toFixed(2) : "0.00"}</div>
                <div style="font-size:11px;color:rgba(255,77,109,.6);margin-top:4px" id="s-refund-count">${data.refund_count} refund${data.refund_count !== 1 ? "s" : ""}</div>
            </div>
            <div class="stat-card sc-green"><div class="stat-label">Net Sales</div><div class="stat-value sv-green" id="s-net">${data.net_sales.toFixed(2)}</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Cash Collected</div><div class="stat-value sv-orange" id="s-collected">${data.cash_collected.toFixed(2)}</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Outstanding</div><div class="stat-value sv-danger" id="s-outstanding">${data.outstanding.toFixed(2)}</div></div>`;
    }
    const salesHead = document.querySelector("#sales-daily")?.closest("table")?.querySelector("thead tr");
    if (salesHead) {
        salesHead.innerHTML = "<th>Date</th><th>Gross Sales</th><th style='color:#ff4d6d'>Refunds</th><th>Net Sales</th><th>Cash Collected</th>";
    }
    setPrintDates("ph-sales-dates", data.date_from, data.date_to);
    document.getElementById("sales-daily").innerHTML = data.daily.length
        ? data.daily.map(d=>`<tr>
            <td class="mono">${d.date}</td>
            <td class="mono" style="color:var(--blue)">${d.gross_sales.toFixed(2)}</td>
            <td class="mono" style="color:#ff4d6d;font-weight:${d.refunds>0?700:400}">${d.refunds>0?"−"+d.refunds.toFixed(2):"—"}</td>
            <td class="mono" style="color:var(--green);font-weight:700">${d.net_sales.toFixed(2)}</td>
            <td class="mono" style="color:var(--orange)">${d.cash_collected.toFixed(2)}</td>
          </tr>`).join("")
        : `<tr><td colspan="5" style="text-align:center;color:var(--muted);padding:30px">No sales in this period</td></tr>`;

    let legacyMaxR = data.top_products.length ? data.top_products[0].revenue : 1;
    document.getElementById("sales-top").innerHTML = data.top_products.length
        ? data.top_products.map(p=>`<div class="bar-row">
            <div class="bar-label">${p.name}</div>
            <div class="bar-track"><div class="bar-fill" style="width:${(p.revenue/legacyMaxR*100).toFixed(1)}%;background:linear-gradient(90deg,var(--green),var(--lime))"></div></div>
            <div class="bar-val" style="color:var(--green)">${p.revenue.toFixed(0)}</div>
          </div>`).join("")
        : `<div style="color:var(--muted);font-size:13px">No data</div>`;

    let itemHtml = `
        <div class="table-title" style="margin-top:28px">Items Sold Detail - ${data.sold_item_records.length} lines</div>
        <div class="table-wrap">
        <div style="overflow-x:auto">
        <table><thead><tr><th>Date / Time</th><th>Source</th><th>Reference</th><th>Customer / Client</th><th>By</th><th>SKU</th><th>Product</th><th>Category</th><th style="text-align:right">Qty</th><th style="text-align:right">Unit Price</th><th style="text-align:right">Line Total</th><th>Payment</th><th>Status</th></tr></thead><tbody>`;
    if(data.sold_item_records.length){
        itemHtml += data.sold_item_records.map(row=>{
            const isRefund = row.line_type === "refund" || row.line_total < 0 || row.qty < 0;
            return `<tr style="${isRefund ? "background:rgba(255,77,109,.04)" : ""}">
                <td class="mono" style="font-size:11px;white-space:nowrap;color:var(--muted)">${row.datetime}</td>
                <td style="font-size:12px;color:${isRefund ? "#ff4d6d" : "var(--sub)"};font-weight:${isRefund ? 700 : 500}">${row.source}</td>
                <td class="mono" style="font-size:11px;color:${isRefund ? "#ff4d6d" : "var(--blue)"}">${row.reference}</td>
                <td class="name" style="font-size:12px;white-space:nowrap">${row.counterparty}</td>
                <td style="font-size:12px;color:var(--muted);white-space:nowrap">${row.user_name}</td>
                <td class="mono" style="font-size:11px;color:var(--muted)">${row.sku}</td>
                <td style="font-weight:600;white-space:nowrap">${row.product}</td>
                <td style="font-size:12px;color:var(--muted)">${row.category}</td>
                <td class="mono" style="text-align:right;color:${isRefund ? "#ff4d6d" : "var(--blue)"};font-weight:700">${row.qty.toFixed(2)}</td>
                <td class="mono" style="text-align:right">${row.unit_price.toFixed(2)}</td>
                <td class="mono" style="text-align:right;color:${isRefund ? "#ff4d6d" : "var(--green)"};font-weight:700">${row.line_total.toFixed(2)}</td>
                <td style="font-size:12px">${row.payment_method}</td>
                <td style="font-size:12px">${row.status}</td>
            </tr>`;
        }).join("");
    } else {
        itemHtml += `<tr><td colspan="13" style="text-align:center;color:var(--muted);padding:24px">No sold items in this period</td></tr>`;
    }
    itemHtml += `</tbody></table></div></div>`;

    let posHtml = `
        <div class="table-title" style="margin-top:28px">POS Invoices — ${data.pos_records.length} transactions</div>
        <div class="table-wrap">
        <table><thead><tr><th>Invoice #</th><th>Date / Time</th><th>Customer</th><th>By</th><th>Payment</th><th>Items</th><th style="text-align:right">Gross</th><th style="text-align:right">Collected</th><th style="text-align:right">Outstanding</th></tr></thead><tbody>`;
    if(data.pos_records.length){
        posHtml += data.pos_records.map(inv=>`
            <tr>
                <td class="mono" style="font-size:11px;color:var(--blue)">${inv.invoice_number}</td>
                <td class="mono" style="font-size:12px;color:var(--muted)">${inv.datetime}</td>
                <td class="name" style="font-size:12px">${inv.customer}</td>
                <td style="font-size:12px;color:var(--muted);white-space:nowrap">${inv.user_name}</td>
                <td style="font-size:12px">${inv.payment}</td>
                <td style="font-size:12px;color:var(--sub)">
                    ${inv.items.map(it=>`<span style="display:inline-block;background:var(--card2);border:1px solid var(--border2);border-radius:5px;padding:1px 7px;margin:2px;white-space:nowrap">${it.qty%1===0?it.qty.toFixed(0):it.qty.toFixed(2)} × ${it.name} <span style="color:var(--muted)">${it.total.toFixed(2)}</span></span>`).join("")}
                </td>
                <td class="mono" style="text-align:right">${inv.total.toFixed(2)}</td>
                <td class="mono" style="text-align:right;color:var(--green);font-weight:700">${inv.cash_collected.toFixed(2)}</td>
                <td class="mono" style="text-align:right;color:${inv.outstanding>0?"var(--warn)":"var(--muted)"}">${inv.outstanding>0?inv.outstanding.toFixed(2):"—"}</td>
            </tr>`).join("");
    } else {
        posHtml += `<tr><td colspan="9" style="text-align:center;color:var(--muted);padding:24px">No POS invoices</td></tr>`;
    }
    posHtml += `</tbody></table></div>`;

    const typeLabel = {cash:"Cash", full_payment:"Full Payment", consignment:"Consignment"};
    let b2bHtml = `
        <div class="table-title" style="margin-top:22px">B2B Invoices Collected — ${data.b2b_records.length} invoices</div>
        <div class="table-wrap">
        <table><thead><tr><th>Invoice #</th><th>Client</th><th>Issued</th><th>Collected</th><th>By</th><th>Type</th><th>Items</th><th style="text-align:right">Invoiced</th><th style="text-align:right">Collected</th><th style="text-align:right">Outstanding</th></tr></thead><tbody>`;
    if(data.b2b_records.length){
        b2bHtml += data.b2b_records.map(inv=>`
            <tr>
                <td class="mono" style="font-size:11px;color:var(--blue)">${inv.invoice_number}</td>
                <td class="name" style="font-size:13px">${inv.client}</td>
                <td class="mono" style="font-size:12px;color:var(--muted)">${inv.datetime}</td>
                <td class="mono" style="font-size:12px;color:var(--green)">${inv.collection_datetime}</td>
                <td style="font-size:12px;color:var(--muted);white-space:nowrap">${inv.user_name}</td>
                <td style="font-size:12px">${typeLabel[inv.invoice_type]||inv.invoice_type}</td>
                <td style="font-size:12px;color:var(--sub)">${inv.items.map(it=>`<span style="display:inline-block;background:var(--card2);border:1px solid var(--border2);border-radius:5px;padding:1px 7px;margin:2px;white-space:nowrap">${it.qty%1===0?it.qty.toFixed(0):it.qty.toFixed(2)} × ${it.name} <span style="color:var(--muted)">${it.total.toFixed(2)}</span></span>`).join("")}</td>
                <td class="mono" style="text-align:right">${inv.total.toFixed(2)}</td>
                <td class="mono" style="text-align:right;color:var(--green)">${inv.collected_in_period.toFixed(2)}</td>
                <td class="mono" style="text-align:right;color:${inv.balance_due>0?"var(--warn)":"var(--muted)"};font-weight:${inv.balance_due>0?700:400}">${inv.balance_due>0?inv.balance_due.toFixed(2):"—"}</td>
            </tr>`).join("");
    } else {
        b2bHtml += `<tr><td colspan="10" style="text-align:center;color:var(--muted);padding:24px">No B2B invoices collected in this period</td></tr>`;
    }
    b2bHtml += `</tbody></table></div>`;

    let b2bCollectionsHtml = `
        <div class="table-title" style="margin-top:22px">B2B Client Collections — ${data.b2b_payment_records.length} payment${data.b2b_payment_records.length!==1?"s":""}</div>
        <div class="table-wrap">
        <table><thead><tr><th>Reference</th><th>Client</th><th>Date / Time</th><th>By</th><th style="text-align:right">Amount</th><th>Notes</th></tr></thead><tbody>`;
    if(data.b2b_payment_records.length){
        b2bCollectionsHtml += data.b2b_payment_records.map(payment=>`
            <tr>
                <td class="mono" style="font-size:11px;color:var(--teal)">${payment.reference}</td>
                <td class="name" style="font-size:13px">${payment.client}</td>
                <td class="mono" style="font-size:12px;color:var(--muted)">${payment.datetime}</td>
                <td style="font-size:12px;color:var(--muted);white-space:nowrap">${payment.user_name}</td>
                <td class="mono" style="text-align:right;color:var(--green);font-weight:700">${payment.amount.toFixed(2)}</td>
                <td style="font-size:12px;color:var(--muted)">${payment.notes || "—"}</td>
            </tr>`).join("");
    } else {
        b2bCollectionsHtml += `<tr><td colspan="6" style="text-align:center;color:var(--muted);padding:24px">No B2B client payments in this period</td></tr>`;
    }
    b2bCollectionsHtml += `</tbody></table></div>`;

    let refHtml = "";
    if(data.refund_records && data.refund_records.length){
        refHtml = `
        <div style="margin-top:28px;display:flex;align-items:center;gap:14px;padding:14px 18px;background:rgba(255,77,109,.06);border:1px solid rgba(255,77,109,.2);border-radius:12px;">
            <div>
                <div style="font-size:13px;font-weight:700;color:#ff4d6d;letter-spacing:.3px">Refunds — ${data.refund_records.length} refund${data.refund_records.length!==1?"s":""}</div>
                <div style="font-size:11px;color:rgba(255,77,109,.6);margin-top:2px">Shown separately from sales and collections</div>
            </div>
            <div style="margin-left:auto;font-family:var(--mono);font-size:22px;font-weight:800;color:#ff4d6d">−${data.refunds.toFixed(2)}</div>
        </div>
        <div class="table-wrap" style="border-color:rgba(255,77,109,.18);">
        <table><thead style="background:rgba(255,77,109,.05)"><tr><th style="color:#ff4d6d">Ref #</th><th>Source</th><th>Counterparty</th><th>Date / Time</th><th>Processed By</th><th>Reason</th><th>Method</th><th style="text-align:right;color:#ff4d6d">Amount</th></tr></thead>
        <tbody>
            ${data.refund_records.map(row=>`<tr>
                <td class="mono" style="font-size:11px;color:#ff4d6d;font-weight:700">${row.refund_number}</td>
                <td>${row.source}</td>
                <td class="name">${row.counterparty}</td>
                <td class="mono" style="font-size:12px;color:var(--muted)">${row.datetime}</td>
                <td style="font-size:12px;color:var(--muted)">${row.processed_by}</td>
                <td style="font-size:12px;color:var(--sub)">${row.reason}</td>
                <td style="font-size:12px">${row.refund_method}</td>
                <td class="mono" style="text-align:right;font-weight:700;color:#ff4d6d">−${row.total.toFixed(2)}</td>
            </tr>`).join("")}
        </tbody></table></div>`;
    }
    document.getElementById("sales-records").innerHTML = itemHtml + posHtml + b2bHtml + b2bCollectionsHtml + refHtml;
    return;
}

/* ── B2B ── */
async function loadB2B(){
    let r = getRange("b2b-from","b2b-to");
    let data = await fetchReportJson(`/reports/api/b2b-statement?date_from=${r.from}&date_to=${r.to}`);
    document.getElementById("b-clients").innerText     = data.length;
    document.getElementById("b-invoiced").innerText    = data.reduce((s,c)=>s+c.total_invoiced,0).toFixed(2);
    document.getElementById("b-outstanding").innerText = data.reduce((s,c)=>s+c.outstanding,0).toFixed(2);
    setPrintDates("ph-b2b-dates", r.from, r.to);
    document.getElementById("b2b-body").innerHTML = data.length
        ? data.map(c=>`<tr>
            <td class="name">${c.name}</td>
            <td style="font-size:12px">${c.phone}</td>
            <td style="font-size:12px">${String(c.payment_terms || "-").replaceAll("_"," ")}</td>
            <td class="mono">${c.total_invoiced.toFixed(2)}</td>
            <td class="mono" style="color:var(--green)">${c.total_paid.toFixed(2)}</td>
            <td class="mono" style="color:${c.outstanding>0?"var(--warn)":"var(--muted)"};font-weight:${c.outstanding>0?700:400}">${c.outstanding>0?c.outstanding.toFixed(2):"—"}</td>
            <td class="mono" style="color:var(--muted)">${c.invoice_count}</td>
          </tr>`).join("")
        : `<tr><td colspan="7" style="text-align:center;color:var(--muted);padding:30px">No data for this period</td></tr>`;
}

/* ── INVENTORY ── */
async function loadInventory(){
    const mode = document.getElementById("inv-mode")?.value || "snapshot";
    const from = document.getElementById("inv-from")?.value || "";
    const to = document.getElementById("inv-to")?.value || "";
    const url = mode === "movement"
        ? `/reports/api/inventory?mode=movement&date_from=${from}&date_to=${to}`
        : "/reports/api/inventory?mode=snapshot";
    let data = await fetchReportJson(url);
    const filterBar = document.querySelector("#section-inventory .filter-bar");
    if (filterBar && !document.getElementById("inv-mode")) {
        filterBar.innerHTML = `
            <label>Mode</label>
            <select id="inv-mode">
                <option value="snapshot">Stock Snapshot</option>
                <option value="movement">Stock Movement</option>
            </select>
            <label>From</label><input type="date" id="inv-from">
            <label>To</label><input type="date" id="inv-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime" onclick="loadInventory()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('inventory')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>`;
        document.getElementById("inv-mode").value = mode;
        document.getElementById("inv-from").value = from || monthStart();
        document.getElementById("inv-to").value = to || today();
    }
    if (data.mode === "movement") {
        document.getElementById("inv-count").innerText = data.total_products;
        document.getElementById("inv-value").innerText = data.summary.stock_in.toFixed(2);
        document.getElementById("inv-low").innerText = data.summary.stock_out.toFixed(2);
        document.getElementById("ph-inv-dates").innerText = `Movement period: ${data.date_from}  →  ${data.date_to}`;
        document.getElementById("inv-body").closest("table").querySelector("thead").innerHTML = "<tr><th>SKU</th><th>Product</th><th>Category</th><th>Unit</th><th>Stock In</th><th>Stock Out</th><th>Receipts</th><th>Sales/Usage</th><th>Spoilage</th><th>Transfers In</th><th>Transfers Out</th><th>Adjustments Net</th><th>Net Movement</th></tr>";
        document.getElementById("inv-body").innerHTML = data.products.length ? data.products.map(p=>`<tr>
            <td class="mono" style="font-size:11px;color:var(--muted)">${p.sku}</td>
            <td class="name">${p.name}</td>
            <td>${p.category}</td>
            <td>${p.unit}</td>
            <td class="mono" style="color:var(--green)">${p.stock_in.toFixed(2)}</td>
            <td class="mono" style="color:var(--danger)">${p.stock_out.toFixed(2)}</td>
            <td class="mono">${p.receipts.toFixed(2)}</td>
            <td class="mono">${p.sales_usage.toFixed(2)}</td>
            <td class="mono">${p.spoilage.toFixed(2)}</td>
            <td class="mono">${p.transfers_in.toFixed(2)}</td>
            <td class="mono">${p.transfers_out.toFixed(2)}</td>
            <td class="mono">${p.adjustments_net.toFixed(2)}</td>
            <td class="mono" style="color:${p.net_movement>=0?"var(--green)":"var(--danger)"}">${p.net_movement.toFixed(2)}</td>
          </tr>`).join("") : `<tr><td colspan="13" style="text-align:center;color:var(--muted);padding:30px">No movement in this period</td></tr>`;
    } else {
        document.getElementById("inv-count").innerText = data.total_products;
        document.getElementById("inv-value").innerText = data.total_value.toFixed(2);
        document.getElementById("inv-low").innerText = data.low_count;
        document.getElementById("ph-inv-dates").innerText = `Snapshot as of ${today()}`;
        document.getElementById("inv-body").closest("table").querySelector("thead").innerHTML = "<tr><th>SKU</th><th>Product</th><th>Category</th><th>Stock</th><th>Unit</th><th>Price</th><th>Stock Value</th><th>Threshold</th><th>Last Move</th><th>Status</th></tr>";
        document.getElementById("inv-body").innerHTML = data.products.map(p=>`<tr>
            <td class="mono" style="font-size:11px;color:var(--muted)">${p.sku}</td>
            <td class="name">${p.name}</td>
            <td>${p.category}</td>
            <td class="mono" style="color:${p.low_stock?"var(--danger)":"var(--text)"};font-weight:700">${p.stock.toFixed(2)}</td>
            <td style="font-size:12px;color:var(--muted)">${p.unit}</td>
            <td class="mono">${p.price.toFixed(2)}</td>
            <td class="mono" style="color:var(--blue)">${p.value.toFixed(2)}</td>
            <td class="mono">${p.threshold.toFixed(2)}</td>
            <td class="mono">${p.last_move_at}</td>
            <td><span class="badge ${p.low_stock?"badge-low":"badge-ok"}">${p.low_stock?"Low Stock":"OK"}${p.dead_stock?" · Dead Stock":""}</span></td>
          </tr>`).join("");
    }
    return;
}

/* ── FARM ── */
async function loadFarm(){
    let r = getRange("farm-from","farm-to");
    let data = await fetchReportJson(`/reports/api/farm-intake?date_from=${r.from}&date_to=${r.to}`);
    setPrintDates("ph-farm-dates", r.from, r.to);
    const summaryRows = data.summary.length
        ? data.summary.map(row=>`<tr>
            <td class="name">${row.farm}</td>
            <td class="mono">${row.delivery_count}</td>
            <td class="mono">${row.line_count}</td>
            <td class="mono" style="color:var(--green)">${row.total_qty.toFixed(2)}</td>
            <td class="mono" style="color:var(--orange)">${Number(row.salary_cost || 0).toFixed(2)}</td>
            <td class="mono" style="color:var(--orange)">${Number(row.labor_cost || 0).toFixed(2)}</td>
            <td>${row.top_product}</td>
          </tr>`).join("")
        : `<tr><td colspan="7" style="text-align:center;color:var(--muted);padding:24px">No farm intake summary in this period</td></tr>`;
    const detailRows = data.detail.length
        ? data.detail.map(row=>`<tr>
            <td class="name">${row.farm}</td>
            <td class="mono">${row.date}</td>
            <td class="mono">${row.delivery_number}</td>
            <td class="mono">${row.sku}</td>
            <td>${row.product}</td>
            <td class="mono">${row.qty.toFixed(2)}</td>
            <td>${row.unit}</td>
            <td>${row.received_by}</td>
            <td>${row.user_name}</td>
          </tr>`).join("")
        : `<tr><td colspan="9" style="text-align:center;color:var(--muted);padding:24px">No farm intake detail in this period</td></tr>`;
    document.getElementById("farm-content").innerHTML = `
        <div class="stats-row">
            <div class="stat-card sc-blue"><div class="stat-label">Farms</div><div class="stat-value sv-blue">${data.totals.farm_count}</div></div>
            <div class="stat-card sc-green"><div class="stat-label">Deliveries</div><div class="stat-value sv-green">${data.totals.delivery_count}</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Line Items</div><div class="stat-value sv-orange">${data.totals.line_count}</div></div>
            <div class="stat-card sc-green"><div class="stat-label">Total Qty</div><div class="stat-value sv-green">${data.totals.total_qty.toFixed(2)}</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Salary & Wages</div><div class="stat-value sv-orange">${Number(data.totals.salary_cost || 0).toFixed(2)}</div></div>
        </div>
        <div class="table-wrap" style="margin-bottom:18px">
            <div class="table-title">Farm Intake Summary</div>
            <table><thead><tr><th>Farm</th><th>Deliveries</th><th>Line Items</th><th>Total Qty</th><th>Salary & Wages</th><th>Labor Cost</th><th>Top Product</th></tr></thead><tbody>${summaryRows}</tbody></table>
        </div>
        <div class="table-wrap">
            <div class="table-title">Farm Intake Detail</div>
            <table><thead><tr><th>Farm</th><th>Date</th><th>Delivery #</th><th>SKU</th><th>Product</th><th>Qty</th><th>Unit</th><th>Received By</th><th>Performed By</th></tr></thead><tbody>${detailRows}</tbody></table>
        </div>`;
    return;
    let summaryHtml = data.farms.map((farm,fi)=>{
        let color = fi===0?"var(--lime)":"var(--teal)";
        let maxQty = farm.products.length ? farm.products[0].total_qty : 1;
        return `<div class="table-wrap" style="margin-bottom:12px">
            <div class="table-title">
                <span>${fi===0?"🌿":"♻️"} ${farm.name}</span>
                <span>${farm.delivery_count} deliveries — ${farm.total_qty.toFixed(1)} total</span>
            </div>
            <div style="padding:14px 16px">
                ${farm.products.length
                    ? farm.products.map(p=>`<div class="bar-row">
                        <div class="bar-label">${p.name}</div>
                        <div class="bar-track"><div class="bar-fill" style="width:${(p.total_qty/maxQty*100).toFixed(1)}%;background:linear-gradient(90deg,${color},var(--green))"></div></div>
                        <div class="bar-val" style="color:${color}">${p.total_qty.toFixed(1)} ${p.unit}</div>
                      </div>`).join("")
                    : `<div style="color:var(--muted);font-size:13px">No deliveries in this period.</div>`}
            </div>
        </div>`;
    }).join("");
    let deliveriesHtml = `<div class="table-wrap">
        <div class="table-title">Delivery Records</div>
        <table><thead><tr><th>Delivery #</th><th>Farm</th><th>Date</th><th>Received By</th><th>Qty</th><th>Items</th><th>By</th><th>Notes</th></tr></thead><tbody>`;
    if(data.deliveries.length){
        deliveriesHtml += data.deliveries.map(d=>`<tr>
            <td class="mono" style="font-size:11px;color:var(--lime)">${d.delivery_number}</td>
            <td class="name">${d.farm}</td>
            <td class="mono" style="font-size:12px;color:var(--muted)">${d.delivery_date}</td>
            <td style="font-size:12px">${d.received_by}</td>
            <td class="mono" style="color:var(--green)">${d.total_qty.toFixed(2)}</td>
            <td class="mono">${d.total_items}</td>
            <td style="font-size:12px;color:var(--muted);white-space:nowrap">${d.user_name}</td>
            <td style="font-size:12px;color:var(--muted)">${d.notes||"—"}</td>
        </tr>`).join("");
    } else {
        deliveriesHtml += `<tr><td colspan="8" style="text-align:center;color:var(--muted);padding:30px">No deliveries in this period</td></tr>`;
    }
    deliveriesHtml += `</tbody></table></div>`;
    document.getElementById("farm-content").innerHTML = summaryHtml + deliveriesHtml;
}


/* ── SPOILAGE ── */
async function loadSpoilage(){
    let r = getRange("spl-from","spl-to");
    let data = await fetchReportJson(`/reports/api/spoilage?date_from=${r.from}&date_to=${r.to}`);
    document.getElementById("spl-count").innerText = data.total_count;
    document.getElementById("spl-qty").innerText   = data.total_qty.toFixed(2);
    setPrintDates("ph-spl-dates", r.from, r.to);
    let maxP = data.by_product.length ? data.by_product[0].qty : 1;
    document.getElementById("spl-by-product").innerHTML = data.by_product.length
        ? data.by_product.map(p=>`<div class="bar-row">
            <div class="bar-label">${p.name}</div>
            <div class="bar-track"><div class="bar-fill" style="width:${(p.qty/maxP*100).toFixed(1)}%;background:linear-gradient(90deg,var(--danger),var(--orange))"></div></div>
            <div class="bar-val" style="color:var(--danger)">${p.qty.toFixed(1)}</div>
          </div>`).join("")
        : `<div style="color:var(--muted);font-size:13px">No data</div>`;
    let maxReasonQty = data.by_reason.length ? data.by_reason[0].qty : 1;
    document.getElementById("spl-by-reason").innerHTML = data.by_reason.length
        ? data.by_reason.map(r=>`<div class="bar-row">
            <div class="bar-label">${r.reason}</div>
            <div class="bar-track"><div class="bar-fill" style="width:${(r.qty/maxReasonQty*100).toFixed(1)}%;background:linear-gradient(90deg,var(--warn),var(--orange))"></div></div>
            <div class="bar-val" style="color:var(--warn)">${r.qty.toFixed(1)}</div>
          </div>`).join("")
        : `<div style="color:var(--muted);font-size:13px">No data</div>`;
    document.getElementById("spl-body").innerHTML = data.records.length
        ? data.records.map(r=>`<tr>
            <td class="mono" style="font-size:11px;color:var(--danger)">${r.ref}</td>
            <td class="name">${r.product}</td>
            <td class="mono" style="color:var(--danger)">-${r.qty.toFixed(2)} ${r.unit}</td>
            <td style="font-size:12px">${r.reason}</td>
            <td style="font-size:12px;color:var(--muted)">${r.farm}</td>
            <td class="mono" style="font-size:12px">${r.date}</td>
            <td style="font-size:12px;color:var(--muted);white-space:nowrap">${r.user_name}</td>
            <td style="font-size:12px;color:var(--muted)">${r.notes}</td>
          </tr>`).join("")
        : `<tr><td colspan="8" style="text-align:center;color:var(--muted);padding:30px">No spoilage in this period</td></tr>`;
}

/* ── PRODUCTION ── */
async function loadProduction(){
    let r = getRange("prod-from","prod-to");
    let data = await fetchReportJson(`/reports/api/production?date_from=${r.from}&date_to=${r.to}`);
    document.getElementById("prod-proc").innerText   = data.total_processing;
    document.getElementById("prod-pkg").innerText    = data.total_packaging;
    document.getElementById("prod-drying").innerText = data.total_drying || 0;
    document.getElementById("prod-loss").innerText   = data.avg_loss_pct.toFixed(1)+"%";
    setPrintDates("ph-prod-dates", r.from, r.to);
    document.getElementById("prod-body").innerHTML = data.batches.length
        ? data.batches.map(b=>{
            const typeColor = (b.type==="Packaging") ? "var(--teal)" : (b.type==="Drying") ? "var(--warn)" : "var(--orange)";
            const typeBg    = (b.type==="Packaging") ? "rgba(45,212,191,.1)" : (b.type==="Drying") ? "rgba(245,158,11,.1)" : "rgba(251,146,60,.1)";
            return `<tr>
            <td class="mono" style="font-size:12px;color:${typeColor}">${b.batch_number}</td>
            <td><span style="font-size:11px;font-weight:700;padding:2px 8px;border-radius:20px;background:${typeBg};color:${typeColor}">${b.type}</span></td>
            <td class="name" style="font-size:12px">${b.recipe}</td>
            <td style="font-size:11px;color:var(--sub);max-width:130px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${b.inputs_str||"—"}</td>
            <td style="font-size:11px;color:var(--green);max-width:130px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${b.outputs_str||"—"}</td>
            <td class="mono" style="color:${b.waste_pct<10?"var(--green)":b.waste_pct<25?"var(--warn)":"var(--danger)"}">${b.waste_pct.toFixed(1)}%</td>
            <td class="mono" style="font-size:12px;color:var(--muted)">${b.date}</td>
            <td style="font-size:12px;color:var(--muted);white-space:nowrap">${b.user_name}</td>
          </tr>`;}).join("")
        : `<tr><td colspan="8" style="text-align:center;color:var(--muted);padding:30px">No batches in this period</td></tr>`;
}


/* ── HR ── */
async function loadHR(){
    let r = getRange("hr-from","hr-to");
    let period = document.getElementById("hr-period").value;
    let department = document.getElementById("hr-department").value.trim();
    let farmId = document.getElementById("hr-farm-id").value;
    let params = new URLSearchParams({date_from:r.from, date_to:r.to});
    if(period) params.set("period", period);
    if(department) params.set("department", department);
    if(farmId) params.set("farm_id", farmId);
    let data = await fetchReportJson(`/reports/api/hr?${params.toString()}`);
    let summary = data.summary;
    setPrintDates("ph-hr-dates", data.date_from, data.date_to);
    document.getElementById("hr-active").innerText = summary.active_employees;
    document.getElementById("hr-rate").innerText = Number(summary.attendance_rate || 0).toFixed(1) + "%";
    document.getElementById("hr-present").innerText = summary.present_days;
    document.getElementById("hr-absent").innerText = summary.absent_days;
    document.getElementById("hr-net").innerText = Number(summary.net_salary || 0).toFixed(2);
    document.getElementById("hr-unpaid").innerText = Number(summary.unpaid_salary || 0).toFixed(2);
    document.getElementById("hr-loans").innerText = Number(summary.total_outstanding_loans || 0).toFixed(2);
    document.getElementById("hr-deductions").innerText = Number(summary.deductions || 0).toFixed(2);
    document.getElementById("hr-paid-cash").innerText = Number(summary.paid_cash || 0).toFixed(2);
    document.getElementById("hr-daysoff").innerText = Number(summary.days_off_credit_balance || 0).toFixed(2) + "d";

    document.getElementById("hr-dept-body").innerHTML = data.by_department.length
        ? data.by_department.map(row=>`<tr>
            <td class="name">${row.department}</td>
            <td class="mono">${row.employees}</td>
            <td class="mono">${row.present_days}</td>
            <td class="mono" style="color:var(--danger)">${row.absent_days}</td>
            <td class="mono" style="color:var(--warn)">${row.late_days}</td>
            <td class="mono" style="color:var(--blue)">${row.leave_days}</td>
            <td class="mono" style="color:var(--green)">${Number(row.net_salary || 0).toFixed(2)}</td>
          </tr>`).join("")
        : `<tr><td colspan="7" style="text-align:center;color:var(--muted);padding:24px">No department records in this period</td></tr>`;

    document.getElementById("hr-farm-body").innerHTML = data.by_farm.length
        ? data.by_farm.map(row=>`<tr>
            <td class="name">${row.farm_name}</td>
            <td class="mono">${row.employees}</td>
            <td class="mono">${row.present_days}</td>
            <td class="mono" style="color:var(--danger)">${row.absent_days}</td>
            <td class="mono" style="color:var(--warn)">${row.late_days}</td>
            <td class="mono" style="color:var(--blue)">${row.leave_days}</td>
            <td class="mono" style="color:var(--green)">${Number(row.net_salary || 0).toFixed(2)}</td>
          </tr>`).join("")
        : `<tr><td colspan="7" style="text-align:center;color:var(--muted);padding:24px">No farm records in this period</td></tr>`;

    document.getElementById("hr-emp-body").innerHTML = data.employees.length
        ? data.employees.map(row=>`<tr>
            <td class="name">${row.employee}</td>
            <td>${row.phone}</td>
            <td>${row.position}</td>
            <td class="mono">${Number(row.base_salary || 0).toFixed(2)}</td>
            <td class="mono">${row.present_days}/${row.attendance_records} present · ${Number(row.attendance_rate || 0).toFixed(1)}%</td>
            <td class="mono">${row.days_worked}/${row.working_days}</td>
            <td class="mono" style="color:var(--orange)">${Number(row.outstanding_loan_balance || 0).toFixed(2)}</td>
            <td class="mono" style="color:var(--danger)">${Number(row.loan_deductions || 0).toFixed(2)}</td>
            <td class="mono" style="color:var(--danger)">${Number(row.day_deduction_days || 0).toFixed(2)}d / ${Number(row.day_deductions || 0).toFixed(2)}</td>
            <td class="mono" style="color:var(--danger)">${Number(row.manual_deductions || 0).toFixed(2)}</td>
            <td class="mono" style="color:var(--green)">${Number(row.net_salary || 0).toFixed(2)}</td>
            <td class="mono" style="color:var(--green)">${Number(row.paid_cash || 0).toFixed(2)}</td>
            <td class="mono" style="font-weight:700;color:${Number(row.days_off_credit_balance||0)<0?"var(--danger)":"var(--teal)"}">${Number(row.days_off_credit_balance || 0).toFixed(2)}</td>
            <td><span class="badge ${row.paid?"badge-ok":"badge-low"}">${row.paid?"Paid":"Unpaid"}</span></td>
          </tr>`).join("")
        : `<tr><td colspan="14" style="text-align:center;color:var(--muted);padding:30px">No employee rows in this period</td></tr>`;
}


/* ── UTILITIES ── */
async function loadUtilities(){
    let r = getRange("util-from","util-to");
    let data = await fetchReportJson(`/reports/api/utilities?date_from=${r.from}&date_to=${r.to}`);
    setPrintDates("ph-util-dates", data.date_from, data.date_to);

    document.getElementById("util-total-cost").innerText   = data.totals.cost.toFixed(2);
    document.getElementById("util-total-cats").innerText   = data.totals.categories;
    document.getElementById("util-total-carbon").innerText = data.totals.carbon_kg_co2e.toFixed(1) + " kg";

    let warn = document.getElementById("util-warning");
    if(data.warning){ warn.style.display = "block"; warn.innerText = data.warning; }
    else            { warn.style.display = "none"; }

    // Summary table
    let utilsBody = document.getElementById("util-summary-body");
    utilsBody.innerHTML = data.utilities.length
        ? data.utilities.map(u => `<tr>
            <td><strong>${u.name}</strong>${u.account_code?` <span style="color:var(--muted);font-size:11px">(${u.account_code})</span>`:""}</td>
            <td>${u.unit_name || "—"}</td>
            <td class="mono">${u.entries}</td>
            <td class="mono">${u.consumption.toLocaleString(undefined,{maximumFractionDigits:2})}</td>
            <td class="mono">${u.cost.toLocaleString(undefined,{minimumFractionDigits:2, maximumFractionDigits:2})}</td>
            <td class="mono">${u.cost_per_unit !== null ? u.cost_per_unit.toLocaleString(undefined,{minimumFractionDigits:2, maximumFractionDigits:4}) : "—"}</td>
            <td class="mono">${u.default_unit_price !== null ? u.default_unit_price.toLocaleString(undefined,{minimumFractionDigits:2, maximumFractionDigits:4}) : "—"}</td>
            <td class="mono" style="color:${u.carbon_kg_co2e>0?"var(--orange,#cc7a00)":"var(--muted)"}">${u.carbon_kg_co2e>0 ? u.carbon_kg_co2e.toLocaleString(undefined,{maximumFractionDigits:1}) : "—"}</td>
        </tr>`).join("")
        : `<tr><td colspan="8" style="text-align:center;color:var(--muted);padding:20px">No utility consumption recorded in this period.</td></tr>`;

    // Monthly trend — months along rows, utilities along columns
    let trendHead = document.getElementById("util-trend-head");
    let trendBody = document.getElementById("util-trend-body");
    if(data.utilities.length){
        trendHead.innerHTML = `<tr><th>Month</th>${data.utilities.map(u=>`<th>${u.name}<br><span style="color:var(--muted);font-size:10px;font-weight:500">${u.unit_name||""}</span></th>`).join("")}</tr>`;
        trendBody.innerHTML = data.trend.map(m => {
            let costByUtil = {};
            let consByUtil = {};
            m.items.forEach(it => { costByUtil[it.id] = it.cost; consByUtil[it.id] = it.consumption; });
            return `<tr>
                <td><strong>${m.label}</strong></td>
                ${data.utilities.map(u => {
                    let cost = costByUtil[u.id] || 0;
                    let cons = consByUtil[u.id] || 0;
                    let costStr = cost > 0 ? cost.toLocaleString(undefined,{minimumFractionDigits:2, maximumFractionDigits:2}) : "—";
                    let consStr = cons > 0 ? `<div style="color:var(--muted);font-size:11px">${cons.toLocaleString(undefined,{maximumFractionDigits:2})} ${u.unit_name||""}</div>` : "";
                    return `<td class="mono">${costStr}${consStr}</td>`;
                }).join("")}
            </tr>`;
        }).join("");
    } else {
        trendHead.innerHTML = `<tr><th>Month</th></tr>`;
        trendBody.innerHTML = `<tr><td style="text-align:center;color:var(--muted);padding:20px">No data.</td></tr>`;
    }

    // Per-farm breakdown
    let farmBody = document.getElementById("util-farm-body");
    farmBody.innerHTML = data.by_farm.length
        ? data.by_farm.map(r => `<tr>
            <td>${r.category}</td>
            <td>${r.farm_name}</td>
            <td class="mono">${r.consumption.toLocaleString(undefined,{maximumFractionDigits:2})}</td>
            <td>${r.unit_name || "—"}</td>
            <td class="mono">${r.cost.toLocaleString(undefined,{minimumFractionDigits:2, maximumFractionDigits:2})}</td>
        </tr>`).join("")
        : `<tr><td colspan="5" style="text-align:center;color:var(--muted);padding:20px">No farm-tagged utility expenses in this period.</td></tr>`;
}


/* ── P&L ── */
async function loadPL(){
    let r = getRange("pl-from","pl-to");
    let data = await fetchReportJson(`/reports/api/pl?date_from=${r.from}&date_to=${r.to}`);
    setPrintDates("ph-pl-dates", data.date_from, data.date_to);
    let isProfit = data.net_profit >= 0;

    const refLabel = {b2b:"B2B Sale", b2b_payment:"B2B Payment", pos:"POS Sale", consignment:"Consignment", payroll:"Payroll", spoilage:"Spoilage", manual:"Manual Entry"};

    function renderEntries(entries, color){
        if(!entries.length) return "";
        return `<div style="background:var(--bg);border-radius:8px;margin:4px 0 10px 24px;overflow:hidden">
            <table style="width:100%;border-collapse:collapse;font-size:12px">
                <thead><tr style="background:var(--card2)">
                    <th style="padding:6px 12px;text-align:left;color:var(--muted);font-weight:700;letter-spacing:.5px">Date</th>
                    <th style="padding:6px 12px;text-align:left;color:var(--muted);font-weight:700;letter-spacing:.5px">Type</th>
                    <th style="padding:6px 12px;text-align:left;color:var(--muted);font-weight:700;letter-spacing:.5px">Description</th>
                    <th style="padding:6px 12px;text-align:right;color:var(--muted);font-weight:700;letter-spacing:.5px">Amount (EGP)</th>
                </tr></thead>
                <tbody>
                ${entries.map(e=>`<tr style="border-top:1px solid var(--border)">
                    <td style="padding:7px 12px;font-family:var(--mono);color:var(--muted)">${e.date}</td>
                    <td style="padding:7px 12px"><span style="background:var(--card2);border-radius:4px;padding:1px 7px;font-size:11px;color:var(--sub)">${refLabel[e.ref_type]||e.ref_type}</span></td>
                    <td style="padding:7px 12px;color:var(--sub)">${e.description}</td>
                    <td style="padding:7px 12px;text-align:right;font-family:var(--mono);font-weight:700;color:${color}">${e.amount.toFixed(2)}</td>
                </tr>`).join("")}
                </tbody>
            </table>
        </div>`;
    }

    function renderAccountLine(item, color, expanded=false){
        let id = "pl-"+item.code.replace(/\\W/g,"");
        return `
            <div class="pl-row" style="cursor:pointer;user-select:none" onclick="togglePLDetail('${id}')">
                <span style="color:var(--sub);display:flex;align-items:center;gap:8px">
                    <span id="${id}-icon" style="color:var(--muted);font-size:11px;transition:transform .2s">${item.entries.length?"▶":""}</span>
                    ${item.code} — ${item.name}
                    <span style="font-size:11px;color:var(--muted)">(${item.entries.length} entries)</span>
                </span>
                <span class="mono" style="color:${color}">${item.amount.toFixed(2)}</span>
            </div>
            <div id="${id}" style="display:none">${renderEntries(item.entries, color)}</div>`;
    }

    document.getElementById("pl-content").innerHTML = `
        <div class="stats-row">
            <div class="stat-card sc-green"><div class="stat-label">Total Revenue</div><div class="stat-value sv-green">${data.total_revenue.toFixed(2)}</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Total Expenses</div><div class="stat-value sv-danger">${data.total_expense.toFixed(2)}</div></div>
            <div class="stat-card ${isProfit?"sc-green":"sc-danger"}">
                <div class="stat-label">Net ${isProfit?"Profit":"Loss"}</div>
                <div class="stat-value ${isProfit?"sv-green":"sv-danger"}">${Math.abs(data.net_profit).toFixed(2)}</div>
            </div>
        </div>
        ${data.warning ? `<div style="font-size:12px;color:var(--warn);margin-bottom:12px">${data.warning}</div>` : ``}
        <div style="font-size:12px;color:var(--muted);margin-bottom:12px">💡 Click any account line to expand its journal entries</div>
        <div class="pl-section">
            <div class="pl-header">Revenue</div>
            ${data.revenue_lines.map(r=>renderAccountLine(r,"var(--green)")).join("") || `<div class="pl-row"><span style="color:var(--muted)">No revenue entries</span><span></span></div>`}
            <div class="pl-row pl-total"><span>Total Revenue</span><span class="mono" style="color:var(--green)">${data.total_revenue.toFixed(2)}</span></div>
        </div>
        <div class="pl-section">
            <div class="pl-header">Expenses</div>
            ${data.expense_lines.map(e=>renderAccountLine(e,"var(--danger)")).join("") || `<div class="pl-row"><span style="color:var(--muted)">No expense entries</span><span></span></div>`}
            <div class="pl-row pl-total"><span>Total Expenses</span><span class="mono" style="color:var(--danger)">${data.total_expense.toFixed(2)}</span></div>
        </div>
        <div class="pl-section">
            <div class="pl-row pl-net" style="background:${isProfit?"rgba(0,255,157,.06)":"rgba(255,77,109,.06)"};border-top:2px solid ${isProfit?"var(--green)":"var(--danger)"}">
                <span>Net ${isProfit?"Profit":"Loss"}</span>
                <span class="mono" style="color:${isProfit?"var(--green)":"var(--danger)"};font-size:20px">${Math.abs(data.net_profit).toFixed(2)}</span>
            </div>
        </div>`;
}

const __rawReportLoaders = {
    sales: loadSales,
    transactions: loadTransactions,
    b2b: loadB2B,
    inventory: loadInventory,
    farm: loadFarm,
    spoilage: loadSpoilage,
    production: loadProduction,
    hr: loadHR,
    utilities: loadUtilities,
    pl: loadPL,
};

loadSales = () => runReportLoader("sales", __rawReportLoaders.sales);
loadTransactions = () => runReportLoader("transactions", __rawReportLoaders.transactions);
loadB2B = () => runReportLoader("b2b", __rawReportLoaders.b2b);
loadInventory = () => runReportLoader("inventory", __rawReportLoaders.inventory);
loadFarm = () => runReportLoader("farm", __rawReportLoaders.farm);
loadSpoilage = () => runReportLoader("spoilage", __rawReportLoaders.spoilage);
loadProduction = () => runReportLoader("production", __rawReportLoaders.production);
loadHR = () => runReportLoader("hr", __rawReportLoaders.hr);
loadUtilities = () => runReportLoader("utilities", __rawReportLoaders.utilities);
loadPL = () => runReportLoader("pl", __rawReportLoaders.pl);

function togglePLDetail(id){
    let el   = document.getElementById(id);
    let icon = document.getElementById(id+"-icon");
    if(!el) return;
    let open = el.style.display === "none" || el.style.display === "";
    el.style.display   = open ? "block" : "none";
    if(icon) icon.style.transform = open ? "rotate(90deg)" : "";
}

/* ── INIT: set default dates ── */
(function initDates(){
    let m = monthStart(), y = yearStart(), t = today();
    setEl("sales-from", m); setEl("sales-to", t);
    setEl("tx-from",    m); setEl("tx-to",    t);
    setEl("b2b-from",   m); setEl("b2b-to",   t);
    setEl("farm-from",  m); setEl("farm-to",   t);
    setEl("spl-from",   m); setEl("spl-to",    t);
    setEl("prod-from",  m); setEl("prod-to",   t);
    setEl("hr-from",    m); setEl("hr-to",     t);
    setEl("hr-period",  t.slice(0,7));
    setEl("util-from",  m); setEl("util-to",   t);
    setEl("pl-from",    y); setEl("pl-to",     t);
    setEl("animals-from", m); setEl("animals-to", t);
    const invMode = document.getElementById("inv-mode");
    const invFrom = document.getElementById("inv-from");
    const invTo = document.getElementById("inv-to");
    if(invMode) invMode.value = "snapshot";
    if(invFrom) invFrom.value = m;
    if(invTo) invTo.value = t;
})();

ensureTabMetadata();
</script>
</body>
</html>"""