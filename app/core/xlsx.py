"""Shared Excel writer for the list exports.

Same styling the inventory / receive exports already use — a green header row,
banded body rows, thin borders and auto-sized columns — kept in one place so new
exports don't each grow their own copy.
"""

import io


def to_xlsx(headers, rows, sheet_name="Report"):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:  # pragma: no cover - openpyxl ships in requirements.txt
        raise Exception("Run: pip install openpyxl --break-system-packages")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    hfill = PatternFill("solid", fgColor="2a7a2a")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5FAF5")

    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 4, 40)
    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
