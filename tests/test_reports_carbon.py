"""Carbon-footprint report (reports page) — contract tests.

The report slices one set of CarbonLog rows along five independent axes, so the
thing worth pinning is that every axis sums back to the same grand total and
that the derived numbers (scope split, intensity, previous-period delta,
target progress) are computed from real rows rather than hard-coded shapes.
"""

import asyncio
import io
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from tests.env_defaults import apply_test_environment_defaults

apply_test_environment_defaults()

import app.routers.reports as reports
from app.database import Base
from app.models.carbon import CarbonEmissionFactor, CarbonLog, CarbonTarget
from app.models.farm import Farm, FarmDelivery, FarmDeliveryItem
from app.models.product import Product
from app.models.production import BatchOutput, ProductionBatch
from app.models.user import User


class AsyncSessionAdapter:
    def __init__(self, session):
        self.session = session

    async def execute(self, statement, params=None):
        return self.session.execute(statement, params or {})


def run(coro):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()
        asyncio.set_event_loop(asyncio.new_event_loop())


async def read_streaming_response(response) -> bytes:
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk)
    return b"".join(chunks)


def make_session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(
        engine,
        tables=[
            Farm.__table__,
            User.__table__,
            Product.__table__,
            FarmDelivery.__table__,
            FarmDeliveryItem.__table__,
            ProductionBatch.__table__,
            BatchOutput.__table__,
            CarbonEmissionFactor.__table__,
            CarbonLog.__table__,
            CarbonTarget.__table__,
        ],
    )
    Session = sessionmaker(bind=engine, expire_on_commit=False)
    return Session()


def seed(session):
    farm = Farm(id=1, name="North Farm", is_active=1)
    user = User(id=1, name="Sara Ops", email="sara@example.com", password="x", role="manager")
    tomato = Product(id=1, sku="TOM", name="Tomato", price=10, unit="kg")
    crate = Product(id=2, sku="CRT", name="Herb bunch", price=5, unit="pcs", unit_weight_kg=Decimal("0.250"))
    session.add_all([farm, user, tomato, crate])

    diesel = CarbonEmissionFactor(
        id=1, source_type="energy", source_key="diesel_liter", label="Diesel fuel",
        factor_kg_co2e_per_unit=Decimal("2.68"), unit="litre", is_active=True,
        scope=1, methodology_source="DEFRA 2024", source_year=2024, region="Global default",
    )
    grid = CarbonEmissionFactor(
        id=2, source_type="energy", source_key="electricity_kwh", label="Grid electricity",
        factor_kg_co2e_per_unit=Decimal("0.45"), unit="kWh", is_active=True,
        scope=2, methodology_source="IEA 2024", source_year=2024, region="Egypt",
    )
    truck = CarbonEmissionFactor(
        id=3, source_type="transport", source_key="truck_km", label="Truck transport",
        factor_kg_co2e_per_unit=Decimal("0.21"), unit="km", is_active=True, scope=3,
    )
    retired = CarbonEmissionFactor(
        id=4, source_type="waste", source_key="old_waste", label="Retired waste factor",
        factor_kg_co2e_per_unit=Decimal("0.40"), unit="kg", is_active=False, scope=3,
    )
    session.add_all([diesel, grid, truck, retired])
    session.flush()

    # In-window logs — 2026-06
    session.add_all([
        CarbonLog(factor_id=1, farm_id=1, user_id=1, log_date=date(2026, 6, 2),
                  quantity=Decimal("100"), kg_co2e=Decimal("268"), ref_type="expense", ref_id=11),
        CarbonLog(factor_id=2, farm_id=1, user_id=1, log_date=date(2026, 6, 10),
                  quantity=Decimal("200"), kg_co2e=Decimal("90"), ref_type="expense", ref_id=12),
        CarbonLog(factor_id=3, farm_id=None, user_id=1, log_date=date(2026, 7, 3),
                  quantity=Decimal("50"), kg_co2e=Decimal("10.5"), ref_type="farm_delivery", ref_id=77),
        CarbonLog(factor_id=3, farm_id=1, user_id=None, log_date=date(2026, 7, 20),
                  quantity=Decimal("30"), kg_co2e=Decimal("6.3"), ref_type=None, ref_id=None),
    ])
    # Previous-period log — 2026-05 (same length window before 2026-06-01)
    session.add(
        CarbonLog(factor_id=1, farm_id=1, user_id=1, log_date=date(2026, 5, 20),
                  quantity=Decimal("50"), kg_co2e=Decimal("134"), ref_type="expense", ref_id=10)
    )

    # Mass-bearing activity for the intensity denominators
    delivery = FarmDelivery(id=1, delivery_number="FD-1", farm_id=1, delivery_date=date(2026, 6, 5))
    session.add(delivery)
    session.flush()
    session.add_all([
        FarmDeliveryItem(delivery_id=1, product_id=1, qty=Decimal("400"), unit="kg"),
        FarmDeliveryItem(delivery_id=1, product_id=2, qty=Decimal("80"), unit="pcs"),   # 80 × 0.25 = 20 kg
    ])
    batch = ProductionBatch(id=1, batch_number="B-1", status="completed",
                            created_at=datetime(2026, 6, 12, 9, 0, 0))
    session.add(batch)
    session.flush()
    session.add(BatchOutput(batch_id=1, product_id=1, qty=Decimal("150")))

    session.add(CarbonTarget(id=1, label="Q2 2026 budget", period_start=date(2026, 6, 1),
                             period_end=date(2026, 6, 30), target_kg_co2e=Decimal("500")))
    session.commit()


def build(session, d_from=date(2026, 6, 1), d_to=date(2026, 7, 31), **overrides):
    params = {"db": AsyncSessionAdapter(session), "d_from": d_from, "d_to": d_to}
    params.update(overrides)
    return run(reports._build_carbon_report(**params))


def test_totals_and_previous_period_delta():
    with make_session() as session:
        seed(session)
        data = build(session)

    assert data["date_from"] == "2026-06-01"
    assert data["date_to"] == "2026-07-31"
    assert data["days"] == 61
    assert data["totals"]["kg_co2e"] == 374.8          # 268 + 90 + 10.5 + 6.3
    assert data["totals"]["tonnes"] == 0.3748
    assert data["totals"]["entries"] == 4
    assert data["totals"]["factors_used"] == 3
    # Previous 61 days ends 2026-05-31 and captures the single May log
    assert data["totals"]["previous_from"] == "2026-04-01"
    assert data["totals"]["previous_to"] == "2026-05-31"
    assert data["totals"]["previous_kg"] == 134.0
    assert data["totals"]["delta_pct"] == 179.7
    assert data["warning"] is None


def test_every_axis_sums_back_to_the_same_total():
    with make_session() as session:
        seed(session)
        data = build(session)

    total = data["totals"]["kg_co2e"]
    for axis in ("by_scope", "by_source", "by_factor", "by_origin", "by_farm", "monthly"):
        assert round(sum(r["kg_co2e"] for r in data[axis]), 3) == total, axis


def test_scope_source_and_origin_breakdowns():
    with make_session() as session:
        seed(session)
        data = build(session)

    scopes = {r["label"]: r for r in data["by_scope"]}
    assert scopes["Scope 1"]["kg_co2e"] == 268.0
    assert scopes["Scope 2"]["kg_co2e"] == 90.0
    assert scopes["Scope 3"]["kg_co2e"] == 16.8
    assert scopes["Scope 1"]["pct"] == 71.5
    assert "Unclassified" not in scopes          # no unscoped factors were used

    sources = {r["label"]: r["kg_co2e"] for r in data["by_source"]}
    assert sources == {"Energy": 358.0, "Transport": 16.8}

    origins = {r["label"]: r["kg_co2e"] for r in data["by_origin"]}
    assert origins["Expense / utility"] == 358.0
    assert origins["Farm delivery (transport)"] == 10.5
    # ref_type NULL is reported as a manual entry rather than dropped
    assert origins["Manual entry"] == 6.3

    farms = {r["farm_name"]: r["kg_co2e"] for r in data["by_farm"]}
    assert farms == {"North Farm": 364.3, "Unassigned": 10.5}


def test_monthly_trend_splits_by_scope():
    with make_session() as session:
        seed(session)
        data = build(session)

    months = {m["month"]: m for m in data["monthly"]}
    assert list(months) == ["2026-06", "2026-07"]
    assert months["2026-06"]["scope_1"] == 268.0
    assert months["2026-06"]["scope_2"] == 90.0
    assert months["2026-06"]["scope_3"] == 0.0
    assert months["2026-07"]["scope_3"] == 16.8
    assert months["2026-07"]["entries"] == 2


def test_intensity_uses_mass_aware_denominators():
    with make_session() as session:
        seed(session)
        data = build(session)

    # 400 kg + (80 pcs × 0.25 kg/pc) = 420 kg intake, 150 kg completed output
    assert data["intensity"]["farm_intake_kg"] == 420.0
    assert data["intensity"]["production_output_kg"] == 150.0
    assert data["intensity"]["per_kg_intake"] == round(374.8 / 420, 4)
    assert data["intensity"]["per_kg_output"] == round(374.8 / 150, 4)


def test_target_progress_is_measured_over_the_targets_own_period():
    with make_session() as session:
        seed(session)
        data = build(session)

    target = data["targets"][0]
    # June-only actual (358), not the 374.8 of the wider report window
    assert target["actual_kg"] == 358.0
    assert target["target_kg"] == 500.0
    assert target["remaining_kg"] == 142.0
    assert target["progress_pct"] == 71.6
    assert target["status"] == "On track"


def test_entry_log_hotspots_and_methodology_appendix():
    with make_session() as session:
        seed(session)
        data = build(session)

    assert data["entries_total"] == 4
    assert [e["kg_co2e"] for e in data["top_entries"]] == [268.0, 90.0, 10.5, 6.3]

    top = data["top_entries"][0]
    assert top["factor"] == "Diesel fuel"
    assert top["scope"] == 1
    assert top["unit"] == "litre"
    assert top["factor_value"] == 2.68
    assert top["origin_label"] == "Expense / utility"
    assert top["farm_name"] == "North Farm"
    assert top["user"] == "Sara Ops"

    unassigned = next(e for e in data["entries"] if e["ref_id"] is None)
    assert unassigned["user"] == "—"
    assert unassigned["origin_label"] == "Manual entry"

    # Inactive factors stay out of the appendix; active ones keep their provenance
    labels = [m["label"] for m in data["methodology"]]
    assert "Retired waste factor" not in labels
    assert labels == ["Diesel fuel", "Grid electricity", "Truck transport"]
    assert data["methodology"][1]["methodology_source"] == "IEA 2024"
    assert data["methodology"][1]["region"] == "Egypt"


def test_empty_period_returns_zeroed_report_with_warning():
    with make_session() as session:
        seed(session)
        data = build(session, d_from=date(2026, 1, 1), d_to=date(2026, 1, 31))

    assert data["totals"]["kg_co2e"] == 0.0
    assert data["totals"]["entries"] == 0
    assert data["totals"]["delta_pct"] is None
    assert data["intensity"]["per_kg_intake"] is None
    assert data["by_source"] == []
    assert data["monthly"] == []
    assert [s["label"] for s in data["by_scope"]] == ["Scope 1", "Scope 2", "Scope 3"]
    assert data["warning"] is not None


def test_excel_export_builds_every_sheet():
    with make_session() as session:
        seed(session)
        response = run(reports.export_carbon(
            date_from="2026-06-01", date_to="2026-07-31", db=AsyncSessionAdapter(session),
        ))
        body = run(read_streaming_response(response))

    wb = openpyxl.load_workbook(io.BytesIO(body))
    assert wb.sheetnames == [
        "Summary", "Targets", "By Scope", "By Source", "By Factor",
        "By Origin", "By Farm", "Monthly Trend", "Emission Log", "Methodology",
    ]
    # title + generated + 2 metadata + spacer + header, then one row per entry
    log = wb["Emission Log"]
    assert log.max_row == 6 + 4
    assert log.cell(row=6, column=1).value == "Date"
    assert log.cell(row=7, column=2).value == "Truck transport"   # newest entry first
    assert "carbon_footprint_2026-06-01_to_2026-07-31.xlsx" in response.headers["Content-Disposition"]


def test_daily_series_is_dropped_for_long_ranges():
    with make_session() as session:
        seed(session)
        short = build(session, d_from=date(2026, 6, 1), d_to=date(2026, 6, 30))
        long_range = build(session, d_from=date(2025, 9, 1), d_to=date(2026, 7, 31))

    assert [d["date"] for d in short["daily"]] == ["2026-06-02", "2026-06-10"]
    assert long_range["daily"] == []
