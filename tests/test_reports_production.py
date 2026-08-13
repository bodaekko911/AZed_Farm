"""Production report — item detail must stay in its own columns.

The report used to flatten each batch's materials into one string per cell
("30kg Tomato, 40pcs Herb"), which put the item and the quantity in the same
column and truncated at 130px. These tests pin the structured shape that
replaced it: per-line product/qty/unit fields, period roll-ups of what was
consumed and produced, and a flat one-row-per-material list.
"""

import asyncio
import io
from datetime import date, datetime, timezone
from decimal import Decimal

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from tests.env_defaults import apply_test_environment_defaults

apply_test_environment_defaults()

import app.routers.reports as reports
from app.database import Base
from app.models.drying import (
    DryingBatch,
    DryingBatchStage,
    DryingBatchStageInput,
    DryingBatchStageOutput,
)
from app.models.product import Product
from app.models.production import BatchInput, BatchOutput, ProductionBatch, Recipe
from app.models.user import User


class AsyncSessionAdapter:
    def __init__(self, session):
        self.session = session

    async def execute(self, statement, params=None):
        return self.session.execute(statement, params or {})


def run(coro):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()
        asyncio.set_event_loop(asyncio.new_event_loop())


async def read_streaming_response(response) -> bytes:
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk)
    return b"".join(chunks)


def make_session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine, tables=[
        User.__table__, Product.__table__, Recipe.__table__,
        ProductionBatch.__table__, BatchInput.__table__, BatchOutput.__table__,
        DryingBatch.__table__, DryingBatchStage.__table__,
        DryingBatchStageInput.__table__, DryingBatchStageOutput.__table__,
    ])
    Session = sessionmaker(bind=engine, expire_on_commit=False)
    return Session()


def seed(session):
    session.add_all([
        User(id=1, name="Ali Operator", email="ali@farm.example", password="x", role="staff"),
        Product(id=1, sku="TOM", name="Tomato", price=Decimal("20"), unit="kg"),
        Product(id=2, sku="SAU", name="Tomato Sauce", price=Decimal("60"), unit="kg"),
        Product(id=3, sku="JAR", name="Glass Jar", price=Decimal("3"), unit="pcs"),
        Recipe(id=1, name="Sauce Recipe"),
    ])
    session.flush()

    # Processing batch — two inputs, one output
    session.add(ProductionBatch(id=1, batch_number="PRD-0001", recipe_id=1, user_id=1,
                                status="completed", waste_pct=Decimal("12.5"), notes="First run",
                                created_at=datetime(2026, 8, 3, tzinfo=timezone.utc)))
    session.flush()
    session.add_all([
        BatchInput(batch_id=1, product_id=1, qty=Decimal("100")),
        BatchInput(batch_id=1, product_id=3, qty=Decimal("40")),
        BatchOutput(batch_id=1, product_id=2, qty=Decimal("87.5")),
    ])

    # Packaging run — reuses Tomato so the roll-up has something to add up
    session.add(ProductionBatch(id=2, batch_number="PKG-0002", recipe_id=None, user_id=1,
                                status="completed", waste_pct=Decimal("0"),
                                created_at=datetime(2026, 8, 5, tzinfo=timezone.utc)))
    session.flush()
    session.add_all([
        BatchInput(batch_id=2, product_id=1, qty=Decimal("50")),
        BatchOutput(batch_id=2, product_id=2, qty=Decimal("50")),
    ])
    session.commit()


def build(session, d_from=None, d_to=None, **kw):
    return run(reports._build_production_report(
        AsyncSessionAdapter(session),
        d_from=d_from or datetime(2026, 8, 1, tzinfo=timezone.utc),
        d_to=d_to or datetime(2026, 8, 31, tzinfo=timezone.utc),
        **kw,
    ))


def test_batch_lines_keep_product_qty_and_unit_apart():
    with make_session() as session:
        seed(session)
        data = build(session)

    batch = next(b for b in data["batches"] if b["batch_number"] == "PRD-0001")
    assert batch["inputs"] == [
        {"product_id": 1, "product": "Tomato", "sku": "TOM", "qty": 100.0, "unit": "kg"},
        {"product_id": 3, "product": "Glass Jar", "sku": "JAR", "qty": 40.0, "unit": "pcs"},
    ]
    assert batch["outputs"] == [
        {"product_id": 2, "product": "Tomato Sauce", "sku": "SAU", "qty": 87.5, "unit": "kg"},
    ]


def test_fractional_quantities_are_not_rounded_to_whole_units():
    """The old string builder used :.0f, so 87.5kg of sauce printed as '88kg'."""
    with make_session() as session:
        seed(session)
        data = build(session)

    batch = next(b for b in data["batches"] if b["batch_number"] == "PRD-0001")
    assert batch["outputs"][0]["qty"] == 87.5


def test_consumed_and_produced_rollups():
    with make_session() as session:
        seed(session)
        data = build(session)

    consumed = {c["product"]: c for c in data["consumed"]}
    assert consumed["Tomato"]["qty"] == 150.0        # 100 + 50 across two batches
    assert consumed["Tomato"]["unit"] == "kg"
    assert consumed["Tomato"]["batches"] == 2
    assert consumed["Glass Jar"]["qty"] == 40.0
    assert consumed["Glass Jar"]["batches"] == 1

    produced = {p["product"]: p for p in data["produced"]}
    assert produced["Tomato Sauce"]["qty"] == 137.5   # 87.5 + 50
    assert produced["Tomato Sauce"]["batches"] == 2

    # Sorted heaviest first so the biggest movers lead
    assert [c["product"] for c in data["consumed"]] == ["Tomato", "Glass Jar"]


def test_flat_item_list_has_one_row_per_material():
    with make_session() as session:
        seed(session)
        data = build(session)

    assert len(data["items"]) == 5                   # 3 lines + 2 lines
    tomato_in = [i for i in data["items"] if i["product"] == "Tomato" and i["direction"] == "Input"]
    assert {i["batch_number"] for i in tomato_in} == {"PRD-0001", "PKG-0002"}
    assert all({"batch_number", "type", "date", "direction", "product", "qty", "unit"} <= set(i)
               for i in data["items"])


def test_rollups_cover_the_whole_period_not_just_the_visible_page():
    """Pagination trims the batch table; the roll-ups must still total
    everything, otherwise page 2 would silently vanish from the summary."""
    with make_session() as session:
        seed(session)
        data = build(session, skip=0, limit=1)

    assert len(data["batches"]) == 1
    assert data["total_batches"] == 2
    consumed = {c["product"]: c["qty"] for c in data["consumed"]}
    assert consumed["Tomato"] == 150.0               # both batches, not just the shown one
    assert len(data["items"]) == 5


def test_legacy_string_columns_still_populated():
    """The Excel batch sheet and any older consumer still read *_str."""
    with make_session() as session:
        seed(session)
        data = build(session)

    batch = next(b for b in data["batches"] if b["batch_number"] == "PRD-0001")
    assert "Tomato" in batch["inputs_str"]
    assert "Glass Jar" in batch["inputs_str"]
    assert "87.5" in batch["outputs_str"]


def test_excel_export_splits_items_into_their_own_sheets():
    with make_session() as session:
        seed(session)
        response = run(reports.export_production(
            date_from="2026-08-01", date_to="2026-08-31", db=AsyncSessionAdapter(session),
        ))
        body = run(read_streaming_response(response))

    wb = openpyxl.load_workbook(io.BytesIO(body))
    assert wb.sheetnames == ["Batches", "Batch Items", "Materials Consumed", "Products Produced"]

    items = wb["Batch Items"]
    header_row = next(r for r in range(1, 15) if items.cell(row=r, column=1).value == "Batch #")
    headers = [items.cell(row=header_row, column=c).value for c in range(1, 9)]
    assert headers == ["Batch #", "Type", "Date", "Direction", "Item", "SKU", "Qty", "Unit"]
    assert items.max_row == header_row + 5           # one row per material line

    # Item and quantity land in different columns, as numbers not text
    first_item = items.cell(row=header_row + 1, column=5).value
    first_qty = items.cell(row=header_row + 1, column=7).value
    assert isinstance(first_item, str)
    assert isinstance(first_qty, (int, float))
